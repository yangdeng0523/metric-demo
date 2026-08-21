"""统一指标维度管理平台 Demo - API 服务

对齐需求文档 6 章 API 设计：
  Base URL: /api/v1（同时以 /api 挂载一份，兼容旧版本前端路径）
  统一响应: {code, message, data}；分页: page/page_size；错误码: 0=成功/400/404/409/500

功能覆盖（需求文档 1.2 目标，P0/P1/P2 全部落地）：
  P0 指标定义与管理  原子/派生/复合指标 定义、编辑、查询（CRUD + 状态 + 引用校验）
  P0 维度定义与管理  维度及维度属性 统一定义和管理（CRUD）
  P0 派生规则引擎    原子指标 + 修饰词(时间周期/统计维度/筛选条件) -> 派生指标，动态 SQL
  P1 统一指标查询    指标 + 维度组合，自动生成并执行 SQL；SQL 透明预览；Excel 导出
  P1 逻辑模型映射    物理表 -> 逻辑模型，多表 JOIN 定义与 SQL 预览
  P2 血缘追溯        物理字段 -> 原子 -> 派生 -> 复合，全链路影响分析/根因追溯
  P2 可视化看板      前端基于查询结果渲染图表
"""
import datetime as dt
import hmac
import io
import json
import re
import secrets
import threading
import time
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from sqlalchemy import func, or_, text

from models import (
    get_session, engine, ensure_schema, STATUS_DRAFT, STATUS_PUBLISHED,
    SubjectDomain, BusinessProcess, BusinessProcessField, Dimension, DimensionAttribute,
    AtomicMetric, DerivedMetric, CompositeMetric, MetricModifier, LogicalModel,
    DownstreamModel, DownstreamApp, Dataset, AppDatasetGrant, ApiCallLog,
    MetricVersion, Approval, QualityRule, Alert, TaskInstance, Schedule,
    EntityTag, CodeRule, CERT_LEVELS, COMPARE_TYPES, MODIFIER_TYPES,
)
from sql_generator import SQLGenerator, MetricNotFoundError, _safe_ident, GRANULARITY_FMT

AGG_FUNCTIONS = ("SUM", "COUNT", "AVG", "MAX", "MIN", "COUNT_DISTINCT")
TIME_PERIODS = ("1d", "7d", "30d", "90d", "ytd", "custom")
FILTER_OPS = ("=", "!=", ">", ">=", "<", "<=", "IN", "NOT IN", "BETWEEN", "LIKE")
COMPARE_LABELS = {"none": "无", "yoy": "同比", "mom": "环比",
                  "yoy_mom": "同比+环比", "cumulative": "累计"}
MODIFIER_TYPE_LABELS = {"time_period": "时间周期", "business_filter": "业务限定",
                        "granularity": "统计粒度"}
CERT_LABELS = {"UNVERIFIED": "未认证", "COMMON": "普通", "CERTIFIED": "认证",
               "QUALITY": "优质"}
DATASET_SOURCES = ("downstream_model", "metric_query")
APP_STATUSES = ("ENABLED", "DISABLED")

router = APIRouter()
openapi_router = APIRouter()
gen = SQLGenerator()
FRONTEND_DIR = Path(__file__).resolve().parent.parent / "frontend"

# 轻量迁移：旧库启动时补齐新增表/列（修饰词库、同环比、Owner/认证/口径文档），幂等
ensure_schema()


# ---------------------------------------------------------------------------
# 统一响应结构：{code, message, data}（文档 6.1）
# ---------------------------------------------------------------------------

def ok(data=None):
    return {"code": 0, "message": "ok", "data": data}


def _get_or_404(s, model, obj_id: int, label: str):
    obj = s.query(model).get(obj_id)
    if not obj:
        raise HTTPException(404, f"{label}不存在: id={obj_id}")
    return obj


def _check_code(s, model, code: str, exclude_id: Optional[int] = None):
    q = s.query(model).filter_by(code=code)
    if exclude_id is not None:
        q = q.filter(model.id != exclude_id)
    if q.first():
        raise HTTPException(409, f"编码已存在: {code}")


def _check_status_arg(status: str):
    if status not in (STATUS_DRAFT, STATUS_PUBLISHED, "ARCHIVED"):
        raise HTTPException(400, f"非法状态: {status}，可选 DRAFT/PUBLISHED/ARCHIVED")


def _check_filters(s, filters: list):
    """筛选条件（业务限定）白名单校验：字段名合法、操作符支持、值形态正确"""
    for f in filters or []:
        if not isinstance(f, dict) or "field" not in f or "op" not in f:
            raise HTTPException(400, f"筛选条件格式错误: {f}，需含 field/op/value")
        if f["op"] not in FILTER_OPS:
            raise HTTPException(400, f"不支持的操作符: {f['op']}")
        if not f["field"] or not str(f["field"]).replace("_", "").isalnum():
            raise HTTPException(400, f"非法字段名: {f['field']}")
        if f["op"] in ("IN", "NOT IN"):
            vals = f.get("value") if isinstance(f.get("value"), list) else [f.get("value")]
            if not vals or vals == [None]:
                raise HTTPException(400, f"{f['op']} 值不能为空")
        elif f["op"] == "BETWEEN":
            vals = f.get("value") if isinstance(f.get("value"), list) else [f.get("value")]
            if not isinstance(vals, list) or len(vals) != 2:
                raise HTTPException(400, "BETWEEN 需要两个值")


def _check_dims(s, dim_codes: list):
    for dc in dim_codes or []:
        if not s.query(Dimension).filter_by(code=dc).first():
            raise HTTPException(400, f"维度不存在: {dc}")


def _check_cert_level(cert_level: str):
    if cert_level and cert_level not in CERT_LEVELS:
        raise HTTPException(400, f"非法认证等级: {cert_level}，可选 {CERT_LEVELS}")


def _check_compare_type(compare_type: str):
    if compare_type not in COMPARE_TYPES:
        raise HTTPException(400, f"非法同环比类型: {compare_type}，可选 {COMPARE_TYPES}")


def _modifier_dict(m: MetricModifier) -> dict:
    return {"id": m.id, "modifier_type": m.modifier_type,
            "type_label": MODIFIER_TYPE_LABELS.get(m.modifier_type, m.modifier_type),
            "code": m.code, "name": m.name, "config": m.config or {},
            "description": m.description,
            "created_at": (m.created_at.strftime("%Y-%m-%d %H:%M:%S")
                           if m.created_at else ""),
            "updated_at": (m.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                           if m.updated_at else "")}


def _check_modifiers(s, modifier_codes: list, atomic_process=None):
    """修饰词库引用校验 + 解析：返回 (time_period, dim_codes, filters)。
    派生指标引用修饰词即不复用内嵌口径（库为准）；删除被引用修饰词返回 409"""
    if not modifier_codes:
        return None, None, None
    seen = set()
    period, dims, filters = None, [], []
    for code in modifier_codes:
        if code in seen:
            continue
        seen.add(code)
        m = s.query(MetricModifier).filter_by(code=code).first()
        if not m:
            raise HTTPException(404, f"修饰词不存在: {code}")
        cfg = m.config or {}
        if m.modifier_type == "time_period":
            p = cfg.get("period")
            if p not in TIME_PERIODS:
                raise HTTPException(400, f"修饰词 {m.code} 的时间周期非法: {p}")
            period = p
        elif m.modifier_type == "business_filter":
            _check_filters(s, cfg.get("filters") or [])
            filters.extend(cfg.get("filters") or [])
        elif m.modifier_type == "granularity":
            _check_dims(s, cfg.get("dim_codes") or [])
            dims.extend(cfg.get("dim_codes") or [])
        else:
            raise HTTPException(400, f"修饰词类型非法: {m.modifier_type}")
    return period, dims, filters


def _modifiers_refs(s, modifier_codes: list):
    """修饰词引用详情（供派生指标详情/列表展示）"""
    out = []
    for code in modifier_codes or []:
        m = s.query(MetricModifier).filter_by(code=code).first()
        if m:
            out.append(_modifier_dict(m))
    return out


def _field_dict(f):
    return {"id": f.id, "code": f.code, "name": f.name, "data_type": f.data_type}


def _check_physical_field(s, process, physical_field: str):
    """原子指标度量字段必须属于业务过程已定义字段（过程未定义字段时宽松兼容）"""
    fields = s.query(BusinessProcessField).filter_by(process_id=process.id).all()
    if fields and not any(f.code == physical_field for f in fields):
        options = "、".join(f.code for f in fields)
        raise HTTPException(
            400, f"物理字段 {physical_field} 不存在于业务过程 {process.name} 的字段列表中，可选: {options}")


def _refs_in_downstream(s, metric_code=None, dim_code=None):
    """下游模型 JSON 引用（无外键），返回命中的下游模型编码"""
    hits = []
    for dm in s.query(DownstreamModel).all():
        for it in (dm.metrics or []):
            if metric_code and it.get("metric_code") == metric_code:
                hits.append(dm.code)
                break
            if dim_code and dim_code in (it.get("dim_codes") or []):
                hits.append(dm.code)
                break
    return hits


def _refs_in_datasets(s, metric_code=None, downstream_id=None):
    hits = []
    for d in s.query(Dataset).all():
        if metric_code and metric_code in (d.metric_codes or []):
            hits.append(d.code)
        elif (downstream_id is not None
              and d.source_type == "downstream_model"
              and d.source_model_id == downstream_id):
            hits.append(d.code)
    return hits


def _metric_sql(code: str, start_date=None, end_date=None):
    """生成指标 SQL（口径透明）；custom 周期无参数时用最近 7 天示例"""
    try:
        mtype, _mname, sql, params = gen.generate(code, None, start_date, end_date)
    except ValueError:
        end = dt.date.today()
        start = end - dt.timedelta(days=7)
        mtype, _mname, sql, params = gen.generate(
            code, None, start.isoformat(), end.isoformat())
    return mtype, sql, {k: str(v) for k, v in params.items()}


def _page_clamped(page: int) -> int:
    """分页参数安全化：page 至少为 1，防止负偏移"""
    return max(1, int(page or 1))


def _page_size_clamped(page_size: int) -> int:
    """分页参数安全化：单页上限 200，防止超大页拖垮查询"""
    return min(200, max(1, int(page_size or 20)))


def _like_escape(kw: str) -> str:
    """LIKE 通配符转义，避免用户输入的 %/_ 放大匹配范围"""
    return (kw or "").replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


# ---------------------------------------------------------------------------
# 治理与运维辅助：编码规范 / 版本快照 / 标签 / 告警
# ---------------------------------------------------------------------------

ENTITY_MODELS = {
    "atomic_metric": AtomicMetric,
    "derived_metric": DerivedMetric,
    "composite_metric": CompositeMetric,
    "dimension": Dimension,
    "logical_model": LogicalModel,
    "downstream_model": DownstreamModel,
}


def _check_code_rule(s, entity_type: str, code: str):
    """编码规范校验：内置规则（seed 写入 meta_code_rule）命中时按正则校验"""
    rule = s.query(CodeRule).filter_by(entity_type=entity_type).first()
    if rule and rule.pattern:
        if not re.fullmatch(rule.pattern, code or ""):
            raise HTTPException(
                400, f"编码不符合规范 {rule.pattern}（示例: {rule.example}）: {code}")


def _entity_by_type(s, entity_type: str, entity_id: int):
    model = ENTITY_MODELS.get(entity_type)
    return s.query(model).get(entity_id) if model else None


def _next_version_no(s, entity_type: str, entity_id: int) -> str:
    """下一个版本号：v1/v2/...（按该实体已有版本数自增）"""
    n = (s.query(MetricVersion).filter_by(entity_type=entity_type,
                                          entity_id=entity_id).count()) + 1
    return f"v{n}"


def _entity_snapshot(obj) -> dict:
    """实体全字段快照（JSON 可序列化：日期转 isoformat）"""
    snap = {}
    for c in obj.__table__.columns:
        v = getattr(obj, c.name)
        if isinstance(v, (dt.date, dt.datetime)):
            v = v.isoformat()
        snap[c.name] = v
    return snap


def _snapshot_version(s, entity_type: str, obj, change_type: str = "update",
                      change_note: str = ""):
    """变更前存档（版本快照）：实体被修改/状态变更/审批通过前调用"""
    s.add(MetricVersion(
        entity_type=entity_type, entity_id=obj.id,
        version_no=_next_version_no(s, entity_type, obj.id),
        snapshot=json.dumps(_entity_snapshot(obj), ensure_ascii=False),
        change_type=change_type, change_note=change_note))


def _with_tags(s, entity_type: str, entity_ids: list):
    """批量附带标签：返回 {entity_id: [tag, ...]}（单次查询）"""
    if not entity_ids:
        return {}
    rows = (s.query(EntityTag).filter_by(entity_type=entity_type)
            .filter(EntityTag.entity_id.in_(entity_ids)).all())
    out = {}
    for r in rows:
        out.setdefault(r.entity_id, []).append(r.tag)
    return out


def _set_tags(s, entity_type: str, entity_id: int, tags: list):
    """全量替换实体标签（先删后插）"""
    s.query(EntityTag).filter_by(entity_type=entity_type,
                                 entity_id=entity_id).delete()
    for t in tags or []:
        t = str(t).strip()
        if t:
            s.add(EntityTag(entity_type=entity_type, entity_id=entity_id, tag=t))


def _purge_entity_artifacts(s, entity_type: str, entity_id: int):
    """删除实体时清理其治理/质量/运维关联数据（版本、审批、标签、质量规则、
    告警、任务实例、数据集删除时的调用日志），避免 SQLite 复用自增 id 后
    旧记录挂到新实体上"""
    for model in (MetricVersion, Approval, EntityTag, QualityRule):
        s.query(model).filter_by(entity_type=entity_type,
                                 entity_id=entity_id).delete()
    # 告警按来源关联（质量规则告警 source_id=规则 id；任务告警 source_id=实例 id）
    if entity_type == "downstream_model":
        rule_ids = [r.id for r in
                    s.query(QualityRule.id).filter_by(entity_id=entity_id)]
        inst_ids = [i.id for i in
                    s.query(TaskInstance.id)
                    .filter_by(entity_type="downstream_model", entity_id=entity_id)]
        if rule_ids:
            s.query(Alert).filter(
                Alert.source_type == "quality",
                Alert.source_id.in_(rule_ids)).delete(synchronize_session=False)
        if inst_ids:
            s.query(Alert).filter(
                Alert.source_type == "task",
                Alert.source_id.in_(inst_ids)).delete(synchronize_session=False)
        s.query(TaskInstance).filter_by(entity_type=entity_type,
                                        entity_id=entity_id).delete()


def _new_alert(s, level: str, source_type: str, source_id, message: str):
    s.add(Alert(level=level, source_type=source_type, source_id=source_id,
                message=message))


# ---------------------------------------------------------------------------
# Pydantic 请求模型
# ---------------------------------------------------------------------------

class DomainIn(BaseModel):
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    description: str = Field("", max_length=2000)
    sort_order: int = 0


class ProcessIn(BaseModel):
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    domain_id: int
    physical_table: str = Field(..., max_length=128)
    date_field: str = Field("order_date", max_length=64)
    description: str = Field("", max_length=2000)


class AtomicIn(BaseModel):
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    process_id: int
    agg_function: str = Field(..., max_length=32)
    physical_field: str = Field(..., max_length=128)
    data_type: str = Field("DECIMAL", max_length=32)
    unit: str = Field("", max_length=64)
    owner: str = Field("", max_length=64)
    cert_level: str = Field("UNVERIFIED", max_length=16)
    biz_definition: str = Field("", max_length=10000)
    description: str = Field("", max_length=2000)
    status: str = STATUS_DRAFT


class DimensionIn(BaseModel):
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    domain_id: int
    physical_table: str = Field(..., max_length=128)
    join_field: str = Field(..., max_length=128)
    name_field: str = Field(..., max_length=128)
    description: str = Field("", max_length=2000)


class AttrIn(BaseModel):
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    physical_field: str = Field(..., max_length=128)
    data_type: str = "STRING"


class ProcessFieldIn(BaseModel):
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    data_type: str = Field("STRING", max_length=32)


class DerivedIn(BaseModel):
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    atomic_code: str = Field(..., max_length=64)
    time_period: str = "7d"
    dim_codes: list = []
    filters: list = []
    modifier_codes: list = []            # 引用修饰词库（时间周期/业务限定/统计粒度）
    compare_type: str = "none"           # none/yoy/mom/yoy_mom/cumulative 同环比自动派生
    owner: str = Field("", max_length=64)
    cert_level: str = Field("UNVERIFIED", max_length=16)
    biz_definition: str = Field("", max_length=10000)
    description: str = Field("", max_length=2000)
    status: str = STATUS_DRAFT


class CompositeIn(BaseModel):
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    expression: str = Field(..., max_length=1000)
    ref_codes: list
    data_type: str = Field("DECIMAL", max_length=32)
    unit: str = Field("", max_length=64)
    owner: str = Field("", max_length=64)
    cert_level: str = Field("UNVERIFIED", max_length=16)
    biz_definition: str = Field("", max_length=10000)
    description: str = Field("", max_length=2000)
    status: str = STATUS_DRAFT


class LogicalModelIn(BaseModel):
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    domain_id: int
    physical_table: str = Field(..., max_length=128)
    join_type: str = "SINGLE"
    join_config: list = []
    description: str = Field("", max_length=2000)


class StatusIn(BaseModel):
    status: str


class ModifierIn(BaseModel):
    modifier_type: str = Field(..., max_length=32)
    code: str = Field(..., max_length=64)
    name: str = Field(..., max_length=128)
    config: dict = {}
    description: str = Field("", max_length=2000)


class QueryRequest(BaseModel):
    metric_code: Optional[str] = None            # 单指标（兼容旧请求）
    metric_codes: Optional[list] = None          # 多指标（优先）
    dim_codes: list = []
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    granularity: str = "day"                     # day/week/month 日期粒度


# ===========================================================================
# 5.1 主题域管理
# ===========================================================================

@router.post("/domains", tags=["主题域"])
def create_domain(body: DomainIn):
    s = get_session()
    try:
        _check_code(s, SubjectDomain, body.code)
        d = SubjectDomain(**body.dict())
        s.add(d)
        s.commit()
        return ok({"id": d.id, "code": d.code})
    finally:
        s.close()


@router.get("/domains", tags=["主题域"])
def list_domains(page: int = 1, page_size: int = 20, keyword: str = ""):
    s = get_session()
    try:
        q = s.query(SubjectDomain)
        if keyword:
            q = q.filter(or_(SubjectDomain.code.like(f"%{_like_escape(keyword)}%", escape="\\"),
                             SubjectDomain.name.like(f"%{_like_escape(keyword)}%", escape="\\")))
        total = q.count()
        rows = (q.order_by(SubjectDomain.sort_order)
                .offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        items = [{
            "id": d.id, "code": d.code, "name": d.name,
            "description": d.description, "sort_order": d.sort_order,
            "process_count": s.query(BusinessProcess).filter_by(domain_id=d.id).count(),
            "dimension_count": s.query(Dimension).filter_by(domain_id=d.id).count(),
        } for d in rows]
        return ok({"items": items, "total": total, "page": page, "page_size": page_size})
    finally:
        s.close()


@router.get("/domains/{domain_id}", tags=["主题域"])
def get_domain(domain_id: int):
    s = get_session()
    try:
        d = _get_or_404(s, SubjectDomain, domain_id, "主题域")
        return ok({
            "id": d.id, "code": d.code, "name": d.name,
            "description": d.description, "sort_order": d.sort_order,
            "processes": [{"id": p.id, "code": p.code, "name": p.name}
                          for p in s.query(BusinessProcess).filter_by(domain_id=d.id).all()],
            "dimensions": [{"id": dm.id, "code": dm.code, "name": dm.name}
                           for dm in s.query(Dimension).filter_by(domain_id=d.id).all()],
        })
    finally:
        s.close()


@router.put("/domains/{domain_id}", tags=["主题域"])
def update_domain(domain_id: int, body: DomainIn):
    s = get_session()
    try:
        d = _get_or_404(s, SubjectDomain, domain_id, "主题域")
        _check_code(s, SubjectDomain, body.code, exclude_id=domain_id)
        for k, v in body.dict().items():
            setattr(d, k, v)
        s.commit()
        return ok({"id": d.id})
    finally:
        s.close()


@router.delete("/domains/{domain_id}", tags=["主题域"])
def delete_domain(domain_id: int):
    """删除校验：有下属业务过程/维度则禁止（用户故事 5.1）"""
    s = get_session()
    try:
        d = _get_or_404(s, SubjectDomain, domain_id, "主题域")
        n1 = s.query(BusinessProcess).filter_by(domain_id=domain_id).count()
        n2 = s.query(Dimension).filter_by(domain_id=domain_id).count()
        if n1 or n2:
            raise HTTPException(409, f"主题域 {d.name} 下有 {n1} 个业务过程、{n2} 个维度，禁止删除")
        s.delete(d)
        s.commit()
        return ok({"deleted": domain_id})
    finally:
        s.close()


# ===========================================================================
# 5.2 业务过程管理
# ===========================================================================

@router.post("/processes", tags=["业务过程"])
def create_process(body: ProcessIn):
    s = get_session()
    try:
        _check_code(s, BusinessProcess, body.code)
        _check_code_rule(s, "business_process", body.code)
        _get_or_404(s, SubjectDomain, body.domain_id, "主题域")
        p = BusinessProcess(**body.dict())
        s.add(p)
        s.commit()
        return ok({"id": p.id, "code": p.code})
    finally:
        s.close()


@router.get("/processes", tags=["业务过程"])
def list_processes(domain_id: Optional[int] = None, keyword: str = ""):
    s = get_session()
    try:
        q = s.query(BusinessProcess)
        if domain_id:
            q = q.filter_by(domain_id=domain_id)
        if keyword:
            q = q.filter(or_(BusinessProcess.code.like(f"%{_like_escape(keyword)}%", escape="\\"),
                             BusinessProcess.name.like(f"%{_like_escape(keyword)}%", escape="\\")))
        rows = q.order_by(BusinessProcess.id).all()
        items = [{
            "id": p.id, "code": p.code, "name": p.name,
            "domain_id": p.domain_id, "domain_name": p.domain_name,
            "physical_table": p.physical_table, "date_field": p.date_field,
            "description": p.description,
            "atomic_count": s.query(AtomicMetric).filter_by(process_id=p.id).count(),
            "fields": [_field_dict(f) for f in p.fields],
            "fields_count": len(p.fields),
        } for p in rows]
        return ok(items)
    finally:
        s.close()


@router.get("/processes/{process_id}", tags=["业务过程"])
def get_process(process_id: int):
    """详情含下属原子指标列表（用户故事：查看过程下所有原子指标）"""
    s = get_session()
    try:
        p = _get_or_404(s, BusinessProcess, process_id, "业务过程")
        atomics = [{"id": a.id, "code": a.code, "name": a.name,
                    "agg_function": a.agg_function, "physical_field": a.physical_field}
                   for a in s.query(AtomicMetric).filter_by(process_id=process_id).all()]
        return ok({
            "id": p.id, "code": p.code, "name": p.name,
            "domain_id": p.domain_id, "domain_name": p.domain_name,
            "physical_table": p.physical_table, "date_field": p.date_field,
            "description": p.description,
            "fields": [_field_dict(f) for f in p.fields],
            "atomics": atomics,
        })
    finally:
        s.close()


@router.put("/processes/{process_id}", tags=["业务过程"])
def update_process(process_id: int, body: ProcessIn):
    s = get_session()
    try:
        p = _get_or_404(s, BusinessProcess, process_id, "业务过程")
        _check_code(s, BusinessProcess, body.code, exclude_id=process_id)
        _check_code_rule(s, "business_process", body.code)
        _get_or_404(s, SubjectDomain, body.domain_id, "主题域")
        for k, v in body.dict().items():
            setattr(p, k, v)
        s.commit()
        return ok({"id": p.id})
    finally:
        s.close()


@router.delete("/processes/{process_id}", tags=["业务过程"])
def delete_process(process_id: int):
    """删除校验：有下属原子指标则禁止删除"""
    s = get_session()
    try:
        p = _get_or_404(s, BusinessProcess, process_id, "业务过程")
        n = s.query(AtomicMetric).filter_by(process_id=process_id).count()
        if n:
            raise HTTPException(409, f"业务过程 {p.name} 下有 {n} 个原子指标，禁止删除")
        s.delete(p)
        s.commit()
        return ok({"deleted": process_id})
    finally:
        s.close()


# ---- 业务过程字段 ----

@router.get("/processes/{process_id}/fields", tags=["业务过程"])
def list_process_fields(process_id: int):
    s = get_session()
    try:
        _get_or_404(s, BusinessProcess, process_id, "业务过程")
        rows = (s.query(BusinessProcessField)
                .filter_by(process_id=process_id)
                .order_by(BusinessProcessField.id).all())
        return ok([_field_dict(f) for f in rows])
    finally:
        s.close()


@router.post("/processes/{process_id}/fields", tags=["业务过程"])
def add_process_field(process_id: int, body: ProcessFieldIn):
    s = get_session()
    try:
        p = _get_or_404(s, BusinessProcess, process_id, "业务过程")
        if not body.code or not str(body.code).replace("_", "").isalnum():
            raise HTTPException(400, "字段名只能包含字母、数字、下划线")
        if any(f.code == body.code for f in p.fields):
            raise HTTPException(409, f"字段已存在: {body.code}")
        f = BusinessProcessField(process_id=process_id, **body.dict())
        s.add(f)
        s.commit()
        return ok({"id": f.id, "code": f.code})
    finally:
        s.close()


@router.put("/business-process-fields/{field_id}", tags=["业务过程"])
def update_process_field(field_id: int, body: ProcessFieldIn):
    s = get_session()
    try:
        f = _get_or_404(s, BusinessProcessField, field_id, "业务过程字段")
        if not str(body.code).replace("_", "").isalnum():
            raise HTTPException(400, "字段名只能包含字母、数字、下划线")
        dup = (s.query(BusinessProcessField)
               .filter_by(process_id=f.process_id, code=body.code)
               .filter(BusinessProcessField.id != field_id).first())
        if dup:
            raise HTTPException(409, f"字段已存在: {body.code}")
        if body.code != f.code:
            # 被原子指标引用的字段禁止改名（级联引用为物理列名，需先调整原子指标）
            n = (s.query(AtomicMetric)
                 .filter_by(process_id=f.process_id, physical_field=f.code).count())
            if n:
                raise HTTPException(
                    409, f"字段 {f.code} 被 {n} 个原子指标引用，请先调整原子指标后再改名")
        for k, v in body.dict().items():
            setattr(f, k, v)
        s.commit()
        return ok({"id": f.id})
    finally:
        s.close()


@router.delete("/business-process-fields/{field_id}", tags=["业务过程"])
def delete_process_field(field_id: int):
    s = get_session()
    try:
        f = _get_or_404(s, BusinessProcessField, field_id, "业务过程字段")
        n = (s.query(AtomicMetric)
             .filter_by(process_id=f.process_id, physical_field=f.code).count())
        if n:
            raise HTTPException(
                409, f"字段 {f.code} 被 {n} 个原子指标引用，禁止删除（请先调整原子指标）")
        s.delete(f)
        s.commit()
        return ok({"deleted": field_id})
    finally:
        s.close()


@router.post("/processes/{process_id}/fields/sync", tags=["业务过程"])
def sync_process_fields(process_id: int):
    """从物理表一键导入真实列（pragma_table_info，跳过代理主键 id），已存在字段跳过（幂等）"""
    s = get_session()
    try:
        p = _get_or_404(s, BusinessProcess, process_id, "业务过程")
        rows = s.execute(text("SELECT name FROM pragma_table_info(:t) WHERE pk = 0"),
                         {"t": p.physical_table}).fetchall()
        columns = sorted(r[0] for r in rows)
        if not columns:
            raise HTTPException(400, f"物理表 {p.physical_table} 不存在或无列，无法同步")
        existing = {f.code for f in p.fields}
        added = []
        for col in columns:
            if col in existing:
                continue
            s.add(BusinessProcessField(process_id=process_id, code=col,
                                       name=col, data_type="STRING"))
            added.append(col)
        s.commit()
        return ok({"added": added, "skipped": len(columns) - len(added),
                   "total": len(columns)})
    finally:
        s.close()


# ===========================================================================
# 5.3 原子指标管理
# ===========================================================================

@router.post("/atomic-metrics", tags=["原子指标"])
def create_atomic(body: AtomicIn):
    s = get_session()
    try:
        _check_code(s, AtomicMetric, body.code)
        _check_code_rule(s, "atomic_metric", body.code)
        process = _get_or_404(s, BusinessProcess, body.process_id, "业务过程")
        if body.agg_function not in AGG_FUNCTIONS:
            raise HTTPException(400, f"非法聚合方式: {body.agg_function}，可选 {AGG_FUNCTIONS}")
        if body.status != STATUS_DRAFT:
            raise HTTPException(400, "新建指标只能为草稿（DRAFT），发布须提交审批流")
        _check_cert_level(body.cert_level)
        _check_physical_field(s, process, body.physical_field)
        a = AtomicMetric(**body.dict())
        s.add(a)
        s.commit()
        return ok({"id": a.id, "code": a.code})
    finally:
        s.close()


@router.get("/atomic-metrics", tags=["原子指标"])
def list_atomic_metrics(process_id: Optional[int] = None, status: Optional[str] = None,
                        keyword: str = "", page: int = 1, page_size: int = 20):
    s = get_session()
    try:
        q = s.query(AtomicMetric)
        if process_id:
            q = q.filter_by(process_id=process_id)
        if status:
            q = q.filter_by(status=status)
        if keyword:
            q = q.filter(or_(AtomicMetric.code.like(f"%{_like_escape(keyword)}%", escape="\\"),
                             AtomicMetric.name.like(f"%{_like_escape(keyword)}%", escape="\\")))
        total = q.count()
        rows = (q.order_by(AtomicMetric.id)
                .offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        tags = _with_tags(s, "atomic_metric", [m.id for m in rows])
        items = [{
            "id": m.id, "code": m.code, "name": m.name,
            "process_id": m.process_id, "process_name": m.process_name,
            "physical_table": m.process.physical_table,
            "agg_function": m.agg_function, "physical_field": m.physical_field,
            "data_type": m.data_type, "unit": m.unit, "status": m.status,
            "owner": m.owner, "cert_level": m.cert_level,
            "biz_definition": m.biz_definition,
            "description": m.description, "tags": tags.get(m.id, []),
        } for m in rows]
        return ok({"items": items, "total": total, "page": page, "page_size": page_size})
    finally:
        s.close()


@router.get("/atomic-metrics/{metric_id}", tags=["原子指标"])
def get_atomic(metric_id: int):
    """详情含引用此指标的派生指标列表（用户故事：查看被哪些派生引用）"""
    s = get_session()
    try:
        m = _get_or_404(s, AtomicMetric, metric_id, "原子指标")
        derived = s.query(DerivedMetric).filter_by(atomic_id=metric_id).all()
        return ok({
            "id": m.id, "code": m.code, "name": m.name,
            "process_id": m.process_id, "process_name": m.process_name,
            "physical_table": m.process.physical_table, "date_field": m.process.date_field,
            "agg_function": m.agg_function, "physical_field": m.physical_field,
            "data_type": m.data_type, "unit": m.unit, "status": m.status,
            "owner": m.owner, "cert_level": m.cert_level,
            "biz_definition": m.biz_definition,
            "description": m.description,
            "tags": [t.tag for t in s.query(EntityTag)
                     .filter_by(entity_type="atomic_metric", entity_id=m.id).all()],
            "derived_refs": [{"id": d.id, "code": d.code, "name": d.name,
                              "time_period": d.time_period} for d in derived],
        })
    finally:
        s.close()


@router.put("/atomic-metrics/{metric_id}", tags=["原子指标"])
def update_atomic(metric_id: int, body: AtomicIn):
    s = get_session()
    try:
        m = _get_or_404(s, AtomicMetric, metric_id, "原子指标")
        _check_code(s, AtomicMetric, body.code, exclude_id=metric_id)
        _check_code_rule(s, "atomic_metric", body.code)
        process = _get_or_404(s, BusinessProcess, body.process_id, "业务过程")
        if body.agg_function not in AGG_FUNCTIONS:
            raise HTTPException(400, f"非法聚合方式: {body.agg_function}")
        _check_cert_level(body.cert_level)
        _check_physical_field(s, process, body.physical_field)
        _snapshot_version(s, "atomic_metric", m, "update", "编辑更新")
        # 状态不随编辑改变（发布走审批流，归档走状态接口）
        for k, v in body.dict().items():
            if k != "status":
                setattr(m, k, v)
        s.commit()
        return ok({"id": m.id})
    finally:
        s.close()


@router.delete("/atomic-metrics/{metric_id}", tags=["原子指标"])
def delete_atomic(metric_id: int):
    """删除校验：被派生指标引用则禁止删除"""
    s = get_session()
    try:
        m = _get_or_404(s, AtomicMetric, metric_id, "原子指标")
        n = s.query(DerivedMetric).filter_by(atomic_id=metric_id).count()
        if n:
            raise HTTPException(409, f"原子指标 {m.name} 被 {n} 个派生指标引用，禁止删除")
        comp_using = [c.code for c in s.query(CompositeMetric).all()
                      if m.code in (c.ref_codes or [])]
        if comp_using:
            raise HTTPException(409, f"原子指标 {m.name} 被复合指标引用: {', '.join(comp_using)}，禁止删除")
        ds = _refs_in_downstream(s, metric_code=m.code)
        qs = _refs_in_datasets(s, metric_code=m.code)
        if ds or qs:
            raise HTTPException(409, f"原子指标 {m.name} 被下游模型/数据集引用，禁止删除")
        _purge_entity_artifacts(s, "atomic_metric", metric_id)
        s.delete(m)
        s.commit()
        return ok({"deleted": metric_id})
    finally:
        s.close()


def _apply_status_transition(s, entity_type, m, target: str):
    """指标状态机：DRAFT/PUBLISHED/ARCHIVED。
    发布（->PUBLISHED）只能走审批流；归档可直切；归档后可重新启用为草稿"""
    _check_status_arg(target)
    cur = m.status
    if cur == target:
        return m.status
    if target == STATUS_PUBLISHED:
        raise HTTPException(400, "发布不能直接改状态：请先提交发布审批（草稿 → 审核 → 发布）")
    if target == "ARCHIVED" and cur not in (STATUS_DRAFT, STATUS_PUBLISHED):
        raise HTTPException(400, f"当前状态 {cur} 不能归档")
    if target == STATUS_DRAFT and cur != "ARCHIVED":
        raise HTTPException(400, "仅已归档指标可重新启用为草稿")
    _snapshot_version(s, entity_type, m, "status",
                      f"状态变更: {cur} -> {target}")
    m.status = target
    return m.status


@router.post("/atomic-metrics/{metric_id}/status", tags=["原子指标"])
def change_atomic_status(metric_id: int, body: StatusIn):
    """发布/归档原子指标：发布必须走审批流，直接置 PUBLISHED 返回 400"""
    s = get_session()
    try:
        m = _get_or_404(s, AtomicMetric, metric_id, "原子指标")
        _apply_status_transition(s, "atomic_metric", m, body.status)
        s.commit()
        return ok({"id": m.id, "code": m.code, "status": m.status})
    finally:
        s.close()


@router.post("/derived-metrics/{metric_id}/status", tags=["派生指标"])
def change_derived_status(metric_id: int, body: StatusIn):
    """发布/归档派生指标：发布必须走审批流"""
    s = get_session()
    try:
        m = _get_or_404(s, DerivedMetric, metric_id, "派生指标")
        _apply_status_transition(s, "derived_metric", m, body.status)
        s.commit()
        return ok({"id": m.id, "code": m.code, "status": m.status})
    finally:
        s.close()


@router.post("/composite-metrics/{metric_id}/status", tags=["复合指标"])
def change_composite_status(metric_id: int, body: StatusIn):
    """发布/归档复合指标：发布必须走审批流"""
    s = get_session()
    try:
        m = _get_or_404(s, CompositeMetric, metric_id, "复合指标")
        _apply_status_transition(s, "composite_metric", m, body.status)
        s.commit()
        return ok({"id": m.id, "code": m.code, "status": m.status})
    finally:
        s.close()


# ===========================================================================
# 5.4 维度管理（含维度属性）
# ===========================================================================

@router.post("/dimensions", tags=["维度"])
def create_dimension(body: DimensionIn):
    s = get_session()
    try:
        _check_code(s, Dimension, body.code)
        _check_code_rule(s, "dimension", body.code)
        _get_or_404(s, SubjectDomain, body.domain_id, "主题域")
        d = Dimension(**body.dict())
        s.add(d)
        s.commit()
        return ok({"id": d.id, "code": d.code})
    finally:
        s.close()


@router.get("/dimensions", tags=["维度"])
def list_dimensions(domain_id: Optional[int] = None, keyword: str = ""):
    s = get_session()
    try:
        q = s.query(Dimension)
        if domain_id:
            q = q.filter_by(domain_id=domain_id)
        if keyword:
            q = q.filter(or_(Dimension.code.like(f"%{_like_escape(keyword)}%", escape="\\"),
                             Dimension.name.like(f"%{_like_escape(keyword)}%", escape="\\")))
        dims = q.order_by(Dimension.id).all()
        tags = _with_tags(s, "dimension", [d.id for d in dims])
        items = [{
            "id": d.id, "code": d.code, "name": d.name,
            "domain_id": d.domain_id, "domain_name": d.domain_name,
            "physical_table": d.physical_table,
            "join_field": d.join_field, "name_field": d.name_field,
            "description": d.description,
            "tags": tags.get(d.id, []),
            "attributes": [{"id": a.id, "code": a.code, "name": a.name,
                            "physical_field": a.physical_field,
                            "data_type": a.data_type}
                           for a in d.attributes],
        } for d in dims]
        return ok(items)
    finally:
        s.close()


@router.get("/dimensions/{dim_id}", tags=["维度"])
def get_dimension(dim_id: int):
    """详情含维度属性列表 + 使用该维度的派生指标"""
    s = get_session()
    try:
        d = _get_or_404(s, Dimension, dim_id, "维度")
        return ok({
            "id": d.id, "code": d.code, "name": d.name,
            "domain_id": d.domain_id, "domain_name": d.domain_name,
            "physical_table": d.physical_table,
            "join_field": d.join_field, "name_field": d.name_field,
            "description": d.description,
            "tags": [t.tag for t in s.query(EntityTag)
                     .filter_by(entity_type="dimension", entity_id=d.id).all()],
            "attributes": [{"id": a.id, "code": a.code, "name": a.name,
                            "physical_field": a.physical_field,
                            "data_type": a.data_type}
                           for a in d.attributes],
            "used_by": [{"code": dm.code, "name": dm.name}
                        for dm in s.query(DerivedMetric).all()
                        if d.code in (dm.dim_codes or [])],
        })
    finally:
        s.close()


@router.put("/dimensions/{dim_id}", tags=["维度"])
def update_dimension(dim_id: int, body: DimensionIn):
    s = get_session()
    try:
        d = _get_or_404(s, Dimension, dim_id, "维度")
        _check_code(s, Dimension, body.code, exclude_id=dim_id)
        _check_code_rule(s, "dimension", body.code)
        _get_or_404(s, SubjectDomain, body.domain_id, "主题域")
        _snapshot_version(s, "dimension", d, "update", "编辑更新")
        for k, v in body.dict().items():
            setattr(d, k, v)
        s.commit()
        return ok({"id": d.id})
    finally:
        s.close()


@router.delete("/dimensions/{dim_id}", tags=["维度"])
def delete_dimension(dim_id: int):
    """删除校验：被派生指标引用则拒绝删除"""
    s = get_session()
    try:
        d = _get_or_404(s, Dimension, dim_id, "维度")
        using = [dm.code for dm in s.query(DerivedMetric).all()
                 if d.code in (dm.dim_codes or [])]
        if using:
            raise HTTPException(409, f"维度 {d.name} 被派生指标引用: {', '.join(using)}，禁止删除")
        # 数据集以该维度作为统计维度时禁止删除（物化表按维度列展开）
        ref_ds = [ds.code for ds in s.query(Dataset).all()
                  if d.code in (ds.dim_codes or [])]
        if ref_ds:
            raise HTTPException(409, f"维度 {d.name} 被数据集引用: {', '.join(ref_ds)}，禁止删除")
        ds = _refs_in_downstream(s, dim_code=d.code)
        if ds:
            raise HTTPException(409, f"维度 {d.name} 被下游模型引用: {', '.join(ds)}，禁止删除")
        _purge_entity_artifacts(s, "dimension", dim_id)
        s.delete(d)
        s.commit()
        return ok({"deleted": dim_id})
    finally:
        s.close()


# ---- 维度属性 ----

@router.post("/dimensions/{dim_id}/attributes", tags=["维度属性"])
def add_attribute(dim_id: int, body: AttrIn):
    s = get_session()
    try:
        d = _get_or_404(s, Dimension, dim_id, "维度")
        if any(a.code == body.code for a in d.attributes):
            raise HTTPException(409, f"属性编码已存在: {body.code}")
        a = DimensionAttribute(dimension_id=dim_id, **body.dict())
        s.add(a)
        s.commit()
        return ok({"id": a.id, "code": a.code})
    finally:
        s.close()


@router.put("/dimension-attributes/{attr_id}", tags=["维度属性"])
def update_attribute(attr_id: int, body: AttrIn):
    s = get_session()
    try:
        a = _get_or_404(s, DimensionAttribute, attr_id, "维度属性")
        for k, v in body.dict().items():
            setattr(a, k, v)
        s.commit()
        return ok({"id": a.id})
    finally:
        s.close()


@router.delete("/dimension-attributes/{attr_id}", tags=["维度属性"])
def delete_attribute(attr_id: int):
    s = get_session()
    try:
        a = _get_or_404(s, DimensionAttribute, attr_id, "维度属性")
        s.delete(a)
        s.commit()
        return ok({"deleted": attr_id})
    finally:
        s.close()


# ===========================================================================
# 5.5 派生指标管理（P0 派生规则引擎）
# ===========================================================================

@router.post("/derived-metrics", tags=["派生指标"])
def create_derived(body: DerivedIn):
    """派生指标 = 原子指标 + 时间周期 + 统计粒度 + 业务限定（修饰词）"""
    s = get_session()
    try:
        _check_code(s, DerivedMetric, body.code)
        _check_code_rule(s, "derived_metric", body.code)
        atomic = s.query(AtomicMetric).filter_by(code=body.atomic_code).first()
        if not atomic:
            raise HTTPException(404, f"原子指标不存在: {body.atomic_code}")
        if body.status != STATUS_DRAFT:
            raise HTTPException(400, "新建指标只能为草稿（DRAFT），发布须提交审批流")
        _check_cert_level(body.cert_level)
        _check_compare_type(body.compare_type)
        # 修饰词库：引用了修饰词则以其为准，内嵌周期/粒度/限定仅作展示回填
        mod_period, mod_dims, mod_filters = _check_modifiers(s, body.modifier_codes)
        if mod_period is not None:
            time_period, dim_codes, filters = mod_period, mod_dims, mod_filters
        else:
            if body.time_period not in TIME_PERIODS:
                raise HTTPException(400, f"非法时间周期: {body.time_period}，可选 {TIME_PERIODS}")
            _check_filters(s, body.filters)
            _check_dims(s, body.dim_codes)
            time_period, dim_codes, filters = body.time_period, body.dim_codes, body.filters
        m = DerivedMetric(code=body.code, name=body.name, atomic_id=atomic.id,
                          time_period=time_period, dim_codes=dim_codes,
                          filters=filters, modifier_codes=list(dict.fromkeys(body.modifier_codes)),
                          compare_type=body.compare_type, owner=body.owner,
                          cert_level=body.cert_level,
                          biz_definition=body.biz_definition,
                          status=body.status, description=body.description)
        s.add(m)
        s.commit()
        return ok({"id": m.id, "code": m.code})
    finally:
        s.close()


@router.get("/derived-metrics", tags=["派生指标"])
def list_derived(atomic_id: Optional[int] = None, keyword: str = "",
                 page: int = 1, page_size: int = 20):
    s = get_session()
    try:
        q = s.query(DerivedMetric)
        if atomic_id:
            q = q.filter_by(atomic_id=atomic_id)
        if keyword:
            q = q.filter(or_(DerivedMetric.code.like(f"%{_like_escape(keyword)}%", escape="\\"),
                             DerivedMetric.name.like(f"%{_like_escape(keyword)}%", escape="\\")))
        total = q.count()
        rows = (q.order_by(DerivedMetric.id)
                .offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        tags = _with_tags(s, "derived_metric", [m.id for m in rows])
        items = [{
            "id": m.id, "code": m.code, "name": m.name,
            "atomic_id": m.atomic_id, "atomic_code": m.atomic.code,
            "atomic_name": m.atomic.name,
            "time_period": m.time_period, "dim_codes": m.dim_codes or [],
            "filters": m.filters or [], "modifier_codes": m.modifier_codes or [],
            "compare_type": m.compare_type or "none",
            "owner": m.owner, "cert_level": m.cert_level,
            "biz_definition": m.biz_definition,
            "status": m.status,
            "description": m.description, "tags": tags.get(m.id, []),
        } for m in rows]
        return ok({"items": items, "total": total, "page": page, "page_size": page_size})
    finally:
        s.close()


@router.get("/derived-metrics/{metric_id}", tags=["派生指标"])
def get_derived(metric_id: int):
    """详情：原子信息 + 维度信息 + 动态生成 SQL（用户故事：查看生成的 SQL）"""
    s = get_session()
    try:
        m = _get_or_404(s, DerivedMetric, metric_id, "派生指标")
        dims = []
        for code in (m.dim_codes or []):
            d = s.query(Dimension).filter_by(code=code).first()
            if d:
                dims.append({"code": d.code, "name": d.name})
        refs = [{"code": c.code, "name": c.name}
                for c in s.query(CompositeMetric).all() if m.code in (c.ref_codes or [])]
        _type, sql, params = _metric_sql(m.code)
        return ok({
            "id": m.id, "code": m.code, "name": m.name,
            "atomic": {"code": m.atomic.code, "name": m.atomic.name,
                       "agg_function": m.atomic.agg_function,
                       "physical_field": m.atomic.physical_field,
                       "table": m.atomic.process.physical_table,
                       "date_field": m.atomic.process.date_field,
                       "unit": m.atomic.unit},
            "time_period": m.time_period, "dims": dims,
            "filters": m.filters or [],
            "modifier_codes": m.modifier_codes or [],
            "modifiers": _modifiers_refs(s, m.modifier_codes),
            "compare_type": m.compare_type or "none",
            "owner": m.owner, "cert_level": m.cert_level,
            "biz_definition": m.biz_definition,
            "status": m.status,
            "description": m.description, "composite_refs": refs,
            "tags": [t.tag for t in s.query(EntityTag)
                     .filter_by(entity_type="derived_metric", entity_id=m.id).all()],
            "generated_sql": sql, "sql_params": params,
        })
    finally:
        s.close()


@router.get("/derived-metrics/{metric_id}/sql-preview", tags=["派生指标"])
def derived_sql_preview(metric_id: int):
    """预览派生指标自动生成的 SQL（口径透明可审计）"""
    s = get_session()
    try:
        m = _get_or_404(s, DerivedMetric, metric_id, "派生指标")
        _type, sql, params = _metric_sql(m.code)
        return ok({"metric_code": m.code, "metric_name": m.name, "type": _type,
                   "sql": sql, "params": params})
    finally:
        s.close()


@router.put("/derived-metrics/{metric_id}", tags=["派生指标"])
def update_derived(metric_id: int, body: DerivedIn):
    s = get_session()
    try:
        m = _get_or_404(s, DerivedMetric, metric_id, "派生指标")
        _check_code(s, DerivedMetric, body.code, exclude_id=metric_id)
        _check_code_rule(s, "derived_metric", body.code)
        atomic = s.query(AtomicMetric).filter_by(code=body.atomic_code).first()
        if not atomic:
            raise HTTPException(404, f"原子指标不存在: {body.atomic_code}")
        _check_cert_level(body.cert_level)
        _check_compare_type(body.compare_type)
        mod_period, mod_dims, mod_filters = _check_modifiers(s, body.modifier_codes)
        if mod_period is not None:
            time_period, dim_codes, filters = mod_period, mod_dims, mod_filters
        else:
            if body.time_period not in TIME_PERIODS:
                raise HTTPException(400, f"非法时间周期: {body.time_period}")
            _check_filters(s, body.filters)
            _check_dims(s, body.dim_codes)
            time_period, dim_codes, filters = body.time_period, body.dim_codes, body.filters
        _snapshot_version(s, "derived_metric", m, "update", "编辑更新")
        m.atomic_id = atomic.id
        # 状态不随编辑改变（发布走审批流，归档走状态接口）
        for k in ("code", "name", "time_period", "dim_codes", "filters",
                  "modifier_codes", "compare_type", "owner", "cert_level",
                  "biz_definition", "description"):
            setattr(m, k, body.dict()[k])
        m.modifier_codes = list(dict.fromkeys(m.modifier_codes or []))
        s.commit()
        return ok({"id": m.id})
    finally:
        s.close()


@router.delete("/derived-metrics/{metric_id}", tags=["派生指标"])
def delete_derived(metric_id: int):
    """删除校验：被复合指标引用则拒绝删除"""
    s = get_session()
    try:
        m = _get_or_404(s, DerivedMetric, metric_id, "派生指标")
        using = [c.code for c in s.query(CompositeMetric).all()
                 if m.code in (c.ref_codes or [])]
        if using:
            raise HTTPException(409, f"派生指标 {m.name} 被复合指标引用: {', '.join(using)}，禁止删除")
        ds = _refs_in_downstream(s, metric_code=m.code)
        qs = _refs_in_datasets(s, metric_code=m.code)
        if ds or qs:
            raise HTTPException(409, f"派生指标 {m.name} 被下游模型/数据集引用，禁止删除")
        _purge_entity_artifacts(s, "derived_metric", metric_id)
        s.delete(m)
        s.commit()
        return ok({"deleted": metric_id})
    finally:
        s.close()


# ===========================================================================
# 5.6 复合指标管理
# ===========================================================================

def _check_refs(s, ref_codes: list):
    if not ref_codes:
        raise HTTPException(400, "复合指标必须引用至少一个指标（派生或原子）")
    for rc in ref_codes:
        dm = s.query(DerivedMetric).filter_by(code=rc).first()
        am = s.query(AtomicMetric).filter_by(code=rc).first()
        if not dm and not am:
            raise HTTPException(400, f"引用的指标不存在: {rc}")


_EXPR_TOKEN_RE = re.compile(r"[a-zA-Z_][a-zA-Z0-9_]*|\d+(\.\d+)?|[+\-*/()]|\s+")


def _check_expression(expression: str, ref_codes: list):
    """复合指标表达式白名单校验：仅允许引用指标 code、数字与 + - * / ( )。
    防止表达式里混入表名/字段名/函数等注入 SQL 生成器"""
    pos = 0
    while pos < len(expression):
        m = _EXPR_TOKEN_RE.match(expression, pos)
        if not m:
            raise HTTPException(400, f"计算表达式包含非法字符: {expression[pos]!r}")
        tok = m.group()
        if (not tok.isspace() and not tok.isdigit()
                and not re.fullmatch(r"[+\-*/()]", tok) and tok not in ref_codes):
            raise HTTPException(400, f"计算表达式包含未定义的标识符: {tok}")
        pos = m.end()
    for ref in ref_codes:
        if ref not in expression:
            raise HTTPException(400, f"计算表达式未引用指标 {ref}")


@router.post("/composite-metrics", tags=["复合指标"])
def create_composite(body: CompositeIn):
    s = get_session()
    try:
        _check_code(s, CompositeMetric, body.code)
        _check_code_rule(s, "composite_metric", body.code)
        _check_refs(s, body.ref_codes)
        _check_expression(body.expression, body.ref_codes)
        if body.status != STATUS_DRAFT:
            raise HTTPException(400, "新建指标只能为草稿（DRAFT），发布须提交审批流")
        _check_cert_level(body.cert_level)
        m = CompositeMetric(**body.dict())
        s.add(m)
        s.commit()
        return ok({"id": m.id, "code": m.code})
    finally:
        s.close()


@router.get("/composite-metrics", tags=["复合指标"])
def list_composites(keyword: str = "", page: int = 1, page_size: int = 20):
    s = get_session()
    try:
        q = s.query(CompositeMetric)
        if keyword:
            q = q.filter(or_(CompositeMetric.code.like(f"%{_like_escape(keyword)}%", escape="\\"),
                             CompositeMetric.name.like(f"%{_like_escape(keyword)}%", escape="\\")))
        total = q.count()
        rows = (q.order_by(CompositeMetric.id)
                .offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        tags = _with_tags(s, "composite_metric", [m.id for m in rows])
        items = [{
            "id": m.id, "code": m.code, "name": m.name,
            "expression": m.expression, "ref_codes": m.ref_codes or [],
            "data_type": m.data_type, "unit": m.unit, "status": m.status,
            "owner": m.owner, "cert_level": m.cert_level,
            "biz_definition": m.biz_definition,
            "description": m.description, "tags": tags.get(m.id, []),
        } for m in rows]
        return ok({"items": items, "total": total, "page": page, "page_size": page_size})
    finally:
        s.close()


@router.get("/composite-metrics/{metric_id}", tags=["复合指标"])
def get_composite(metric_id: int):
    """详情含引用的派生指标信息、计算表达式与生成 SQL"""
    s = get_session()
    try:
        m = _get_or_404(s, CompositeMetric, metric_id, "复合指标")
        refs = []
        for code in (m.ref_codes or []):
            dm = s.query(DerivedMetric).filter_by(code=code).first()
            am = s.query(AtomicMetric).filter_by(code=code).first()
            if dm:
                refs.append({"code": dm.code, "name": dm.name, "type": "derived",
                             "atomic_code": dm.atomic.code,
                             "time_period": dm.time_period})
            elif am:
                refs.append({"code": am.code, "name": am.name, "type": "atomic",
                             "agg_function": am.agg_function,
                             "physical_field": am.physical_field,
                             "time_period": "custom（查询窗口）"})
        _type, sql, params = _metric_sql(m.code)
        return ok({
            "id": m.id, "code": m.code, "name": m.name,
            "expression": m.expression, "ref_codes": m.ref_codes or [],
            "refs": refs,
            "data_type": m.data_type, "unit": m.unit, "status": m.status,
            "description": m.description,
            "tags": [t.tag for t in s.query(EntityTag)
                     .filter_by(entity_type="composite_metric", entity_id=m.id).all()],
            "generated_sql": sql, "sql_params": params,
        })
    finally:
        s.close()


@router.get("/composite-metrics/{metric_id}/sql-preview", tags=["复合指标"])
def composite_sql_preview(metric_id: int):
    s = get_session()
    try:
        m = _get_or_404(s, CompositeMetric, metric_id, "复合指标")
        _type, sql, params = _metric_sql(m.code)
        return ok({"metric_code": m.code, "metric_name": m.name,
                   "type": _type, "sql": sql, "params": params})
    finally:
        s.close()


@router.put("/composite-metrics/{metric_id}", tags=["复合指标"])
def update_composite(metric_id: int, body: CompositeIn):
    s = get_session()
    try:
        m = _get_or_404(s, CompositeMetric, metric_id, "复合指标")
        _check_code(s, CompositeMetric, body.code, exclude_id=metric_id)
        _check_code_rule(s, "composite_metric", body.code)
        _check_refs(s, body.ref_codes)
        _check_expression(body.expression, body.ref_codes)
        _check_cert_level(body.cert_level)
        _snapshot_version(s, "composite_metric", m, "update", "编辑更新")
        # 状态不随编辑改变（发布走审批流，归档走状态接口）
        for k, v in body.dict().items():
            if k != "status":
                setattr(m, k, v)
        s.commit()
        return ok({"id": m.id})
    finally:
        s.close()


@router.delete("/composite-metrics/{metric_id}", tags=["复合指标"])
def delete_composite(metric_id: int):
    s = get_session()
    try:
        m = _get_or_404(s, CompositeMetric, metric_id, "复合指标")
        qs = _refs_in_datasets(s, metric_code=m.code)
        if qs:
            raise HTTPException(409, f"复合指标 {m.name} 被数据集引用: {', '.join(qs)}，禁止删除")
        _purge_entity_artifacts(s, "composite_metric", metric_id)
        s.delete(m)
        s.commit()
        return ok({"deleted": metric_id})
    finally:
        s.close()


# ===========================================================================
# 5.6b 修饰词库：时间周期 / 业务限定 / 统计粒度 独立成库（可复用，不写死在派生指标）
# ===========================================================================

@router.post("/modifiers", tags=["修饰词库"])
def create_modifier(body: ModifierIn):
    s = get_session()
    try:
        _check_code(s, MetricModifier, body.code)
        _check_code_rule(s, "modifier", body.code)
        if body.modifier_type not in MODIFIER_TYPES:
            raise HTTPException(400, f"非法修饰词类型: {body.modifier_type}，可选 {MODIFIER_TYPES}")
        if body.modifier_type == "time_period":
            if (body.config or {}).get("period") not in TIME_PERIODS:
                raise HTTPException(400, f"时间周期修饰词 config.period 非法，可选 {TIME_PERIODS}")
        elif body.modifier_type == "business_filter":
            _check_filters(s, (body.config or {}).get("filters") or [])
        elif body.modifier_type == "granularity":
            _check_dims(s, (body.config or {}).get("dim_codes") or [])
        m = MetricModifier(modifier_type=body.modifier_type, code=body.code,
                           name=body.name, config=body.config or {},
                           description=body.description)
        s.add(m)
        s.commit()
        return ok({"id": m.id, "code": m.code})
    finally:
        s.close()


@router.get("/modifiers", tags=["修饰词库"])
def list_modifiers(modifier_type: Optional[str] = None, keyword: str = "",
                   page: int = 1, page_size: int = 100):
    s = get_session()
    try:
        q = s.query(MetricModifier)
        if modifier_type:
            if modifier_type not in MODIFIER_TYPES:
                raise HTTPException(400, f"非法修饰词类型: {modifier_type}")
            q = q.filter_by(modifier_type=modifier_type)
        if keyword:
            q = q.filter(or_(MetricModifier.code.like(f"%{_like_escape(keyword)}%", escape="\\"),
                             MetricModifier.name.like(f"%{_like_escape(keyword)}%", escape="\\")))
        total = q.count()
        rows = (q.order_by(MetricModifier.modifier_type, MetricModifier.id)
                .offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        # 使用量：被多少派生指标引用
        used = {}
        for dm in s.query(DerivedMetric).all():
            for code in (dm.modifier_codes or []):
                used[code] = used.get(code, 0) + 1
        items = []
        for m in rows:
            d = _modifier_dict(m)
            d["used_by"] = used.get(m.code, 0)
            items.append(d)
        return ok({"items": items, "total": total, "page": page, "page_size": page_size})
    finally:
        s.close()


@router.put("/modifiers/{modifier_id}", tags=["修饰词库"])
def update_modifier(modifier_id: int, body: ModifierIn):
    s = get_session()
    try:
        m = _get_or_404(s, MetricModifier, modifier_id, "修饰词")
        _check_code(s, MetricModifier, body.code, exclude_id=modifier_id)
        _check_code_rule(s, "modifier", body.code)
        if body.modifier_type not in MODIFIER_TYPES:
            raise HTTPException(400, f"非法修饰词类型: {body.modifier_type}")
        if body.modifier_type == "time_period":
            if (body.config or {}).get("period") not in TIME_PERIODS:
                raise HTTPException(400, f"时间周期修饰词 config.period 非法")
        elif body.modifier_type == "business_filter":
            _check_filters(s, (body.config or {}).get("filters") or [])
        elif body.modifier_type == "granularity":
            _check_dims(s, (body.config or {}).get("dim_codes") or [])
        for k, v in body.dict().items():
            setattr(m, k, v)
        m.config = body.config or {}
        s.commit()
        return ok({"id": m.id})
    finally:
        s.close()


@router.delete("/modifiers/{modifier_id}", tags=["修饰词库"])
def delete_modifier(modifier_id: int):
    """删除保护：被派生指标引用的修饰词禁止删除"""
    s = get_session()
    try:
        m = _get_or_404(s, MetricModifier, modifier_id, "修饰词")
        using = [dm.code for dm in s.query(DerivedMetric).all()
                 if m.code in (dm.modifier_codes or [])]
        if using:
            raise HTTPException(409, f"修饰词 {m.name} 被派生指标引用: {', '.join(using)}，禁止删除")
        s.delete(m)
        s.commit()
        return ok({"deleted": modifier_id})
    finally:
        s.close()


# ===========================================================================
# 5.7 统一指标查询（核心：动态 SQL 生成 + 执行）
# ===========================================================================

@router.post("/query", tags=["统一指标查询"])
def query(body: QueryRequest):
    codes = body.metric_codes or ([body.metric_code] if body.metric_code else None)
    if not codes:
        raise HTTPException(400, "必须指定指标（metric_codes 或 metric_code）")
    if body.granularity not in GRANULARITY_FMT:
        raise HTTPException(400, f"不支持的日期粒度: {body.granularity}")
    try:
        meta, cols, rows, sql = gen.execute_multi(
            codes, body.dim_codes, body.start_date, body.end_date, body.granularity)
    except MetricNotFoundError as e:
        raise HTTPException(404, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))

    # 按指标分别汇总，避免多指标把金额与笔数加在一起
    n_dims = len(body.dim_codes)
    metric_summaries = []
    for i, (code, name, mtype) in enumerate(
            zip(codes, meta["metric_names"], meta["metric_types"])):
        col_i = 1 + n_dims + i
        values = [r[col_i] for r in rows
                  if col_i < len(r) and isinstance(r[col_i], (int, float))]
        metric_summaries.append({
            "code": code, "name": name, "type": mtype,
            "total": round(sum(values), 2) if values else None,
            "avg": round(sum(values) / len(values), 2) if values else None,
        })
    only = metric_summaries[0] if len(metric_summaries) == 1 else None
    summary = {
        "metric_names": meta["metric_names"], "metric_types": meta["metric_types"],
        "granularity": meta["granularity"], "row_count": len(rows),
        "metrics": metric_summaries,
        "total": only["total"] if only else None,
        "avg": only["avg"] if only else None,
    }
    return ok({"summary": summary, "columns": cols, "rows": rows, "sql": sql})


@router.get("/sql-preview", tags=["指标查询"])
def sql_preview(metric_codes: str, dim_codes: str = "",
                start_date: Optional[str] = None, end_date: Optional[str] = None,
                granularity: str = "day"):
    """只生成 SQL 不执行（口径透明：任何查询可查看生成逻辑）"""
    codes = [c for c in metric_codes.split(",") if c]
    dims = [d for d in dim_codes.split(",") if d] if dim_codes else []
    if not codes:
        raise HTTPException(400, "metric_codes 不能为空")
    if granularity not in GRANULARITY_FMT:
        raise HTTPException(400, f"不支持的日期粒度: {granularity}")
    try:
        mtypes, mnames, sql, params = gen.generate_multi(
            codes, dims, start_date, end_date, granularity)
    except (MetricNotFoundError, ValueError) as e:
        raise HTTPException(400, str(e))
    return ok({"metric_codes": codes, "metric_names": mnames,
               "metric_types": mtypes, "granularity": granularity,
               "sql": sql, "params": params})


def _export_excel(cols, rows, title="指标查询结果"):
    """查询结果导出 Excel（openpyxl 生成 .xlsx）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F54EB")
    for r in rows:
        ws.append(r)
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width * 1.9 + 4, 42)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/query/export", tags=["指标查询"])
def export_query(metric_codes: str, dim_codes: str = "",
                 start_date: Optional[str] = None, end_date: Optional[str] = None,
                 granularity: str = "day"):
    """导出查询结果为 Excel（.xlsx 下载）"""
    codes = [c for c in metric_codes.split(",") if c]
    dims = [d for d in dim_codes.split(",") if d] if dim_codes else []
    if not codes:
        raise HTTPException(400, "metric_codes 不能为空")
    try:
        _meta, cols, rows, _sql = gen.execute_multi(
            codes, dims, start_date, end_date, granularity)
    except (MetricNotFoundError, ValueError) as e:
        raise HTTPException(400, str(e))
    buf = _export_excel(cols, rows, "指标查询")
    filename = f"metric_{codes[0]}_multi.xlsx"
    return StreamingResponse(
        buf,
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition": f"attachment; filename={filename}"})


# ===========================================================================
# 5.8 逻辑模型管理（P1：物理表 -> 逻辑模型，屏蔽底层差异）
# ===========================================================================

def _logical_model_sql(m: LogicalModel) -> str:
    return gen.logical_model_sql(m)


@router.post("/logical-models", tags=["逻辑模型"])
def create_logical_model(body: LogicalModelIn):
    s = get_session()
    try:
        _check_code(s, LogicalModel, body.code)
        _check_code_rule(s, "logical_model", body.code)
        _get_or_404(s, SubjectDomain, body.domain_id, "主题域")
        m = LogicalModel(**body.dict())
        try:
            _logical_model_sql(m)
        except ValueError as e:
            raise HTTPException(400, str(e))
        s.add(m)
        s.commit()
        return ok({"id": m.id, "code": m.code})
    finally:
        s.close()


@router.get("/logical-models", tags=["逻辑模型"])
def list_logical_models(keyword: str = ""):
    s = get_session()
    try:
        q = s.query(LogicalModel)
        if keyword:
            q = q.filter(or_(LogicalModel.code.like(f"%{_like_escape(keyword)}%", escape="\\"),
                             LogicalModel.name.like(f"%{_like_escape(keyword)}%", escape="\\")))
        lms = q.order_by(LogicalModel.id).all()
        tags = _with_tags(s, "logical_model", [m.id for m in lms])
        items = [{
            "id": m.id, "code": m.code, "name": m.name,
            "domain_id": m.domain_id, "domain_name": m.domain_name,
            "physical_table": m.physical_table, "join_type": m.join_type,
            "join_config": m.join_config or [], "description": m.description,
            "tags": tags.get(m.id, []),
            "generated_sql": _logical_model_sql(m),
        } for m in lms]
        return ok(items)
    finally:
        s.close()


@router.get("/logical-models/{model_id}", tags=["逻辑模型"])
def get_logical_model(model_id: int):
    s = get_session()
    try:
        m = _get_or_404(s, LogicalModel, model_id, "逻辑模型")
        return ok({
            "id": m.id, "code": m.code, "name": m.name,
            "domain_id": m.domain_id, "domain_name": m.domain_name,
            "physical_table": m.physical_table, "join_type": m.join_type,
            "join_config": m.join_config or [], "description": m.description,
            "tags": [t.tag for t in s.query(EntityTag)
                     .filter_by(entity_type="logical_model", entity_id=m.id).all()],
            "generated_sql": _logical_model_sql(m),
        })
    finally:
        s.close()


@router.put("/logical-models/{model_id}", tags=["逻辑模型"])
def update_logical_model(model_id: int, body: LogicalModelIn):
    s = get_session()
    try:
        m = _get_or_404(s, LogicalModel, model_id, "逻辑模型")
        _check_code(s, LogicalModel, body.code, exclude_id=model_id)
        _check_code_rule(s, "logical_model", body.code)
        _get_or_404(s, SubjectDomain, body.domain_id, "主题域")
        _snapshot_version(s, "logical_model", m, "update", "编辑更新")
        for k, v in body.dict().items():
            setattr(m, k, v)
        try:
            _logical_model_sql(m)
        except ValueError as e:
            raise HTTPException(400, str(e))
        s.commit()
        return ok({"id": m.id})
    finally:
        s.close()


@router.delete("/logical-models/{model_id}", tags=["逻辑模型"])
def delete_logical_model(model_id: int):
    s = get_session()
    try:
        m = _get_or_404(s, LogicalModel, model_id, "逻辑模型")
        n = s.query(DownstreamModel).filter_by(source_model_id=model_id).count()
        if n:
            raise HTTPException(409, f"逻辑模型 {m.name} 被 {n} 个下游模型引用，禁止删除")
        _purge_entity_artifacts(s, "logical_model", model_id)
        s.delete(m)
        s.commit()
        return ok({"deleted": model_id})
    finally:
        s.close()


# ===========================================================================
# 5.9 下游模型管理：基于逻辑模型 + 指标集合生成指标汇总模型，支持物化
# 定义 SQL 生成见 SQLGenerator.generate_downstream_sql（sql_generator.py）
# ===========================================================================

class DownstreamIn(BaseModel):
    code: str
    name: str
    source_model_id: int
    metrics: list
    granularity: str = "day"
    description: str = ""


@router.post("/downstream-models", tags=["下游模型"])
def create_downstream(body: DownstreamIn):
    s = get_session()
    try:
        _check_code(s, DownstreamModel, body.code)
        lm = _get_or_404(s, LogicalModel, body.source_model_id, "逻辑模型")
        m = DownstreamModel(**body.dict())
        try:
            sql, params = gen.generate_downstream_sql(m, lm)
            m.definition_sql = sql
        except (ValueError, MetricNotFoundError) as e:
            raise HTTPException(400, str(e))
        s.add(m)
        s.commit()
        return ok({"id": m.id, "definition_sql": m.definition_sql,
                   "params": params})
    finally:
        s.close()


@router.get("/downstream-models", tags=["下游模型"])
def list_downstream(page: int = 1, page_size: int = 20, keyword: str = ""):
    s = get_session()
    try:
        q = s.query(DownstreamModel).order_by(DownstreamModel.id.desc())
        if keyword:
            like = f"%{keyword}%"
            q = q.filter(or_(DownstreamModel.code.like(_like_escape(like), escape="\\"),
                             DownstreamModel.name.like(_like_escape(like), escape="\\")))
        total = q.count()
        rows = (q.offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        items = [{
            "id": m.id, "code": m.code, "name": m.name,
            "source_model_id": m.source_model_id,
            "source_model_name": m.source_model_name,
            "metrics": m.metrics or [], "granularity": m.granularity,
            "materialized": bool(m.materialized),
            "physical_table": m.physical_table, "row_count": m.row_count,
            "description": m.description,
        } for m in rows]
        return ok({"total": total, "page": page, "page_size": page_size,
                   "items": items})
    finally:
        s.close()


@router.get("/downstream-models/{model_id}", tags=["下游模型"])
def get_downstream(model_id: int):
    s = get_session()
    try:
        m = _get_or_404(s, DownstreamModel, model_id, "下游模型")
        return ok({
            "id": m.id, "code": m.code, "name": m.name,
            "source_model_id": m.source_model_id,
            "source_model_name": m.source_model_name,
            "metrics": m.metrics or [], "granularity": m.granularity,
            "definition_sql": m.definition_sql,
            "materialized": bool(m.materialized),
            "physical_table": m.physical_table, "row_count": m.row_count,
            "description": m.description,
        })
    finally:
        s.close()


@router.put("/downstream-models/{model_id}", tags=["下游模型"])
def update_downstream(model_id: int, body: DownstreamIn):
    s = get_session()
    try:
        m = _get_or_404(s, DownstreamModel, model_id, "下游模型")
        _check_code(s, DownstreamModel, body.code, exclude_id=model_id)
        lm = _get_or_404(s, LogicalModel, body.source_model_id, "逻辑模型")
        for k, v in body.dict().items():
            setattr(m, k, v)
        try:
            sql, _ = gen.generate_downstream_sql(m, lm)
            m.definition_sql = sql
        except (ValueError, MetricNotFoundError) as e:
            raise HTTPException(400, str(e))
        # 定义变更后旧物化表过期：落地表删除并复位状态
        if m.materialized and m.physical_table:
            tbl = _safe_ident(m.physical_table)
            with engine.begin() as conn:
                conn.execute(text(f"DROP TABLE IF EXISTS {tbl}"))
            m.materialized = 0
            m.physical_table = None
            m.row_count = None
        s.commit()
        return ok({"id": m.id})
    finally:
        s.close()


@router.delete("/downstream-models/{model_id}", tags=["下游模型"])
def delete_downstream(model_id: int):
    s = get_session()
    try:
        m = _get_or_404(s, DownstreamModel, model_id, "下游模型")
        qs = _refs_in_datasets(s, downstream_id=model_id)
        if qs:
            raise HTTPException(409, f"下游模型 {m.name} 被数据集引用: {', '.join(qs)}，禁止删除")
        if m.materialized and m.physical_table:
            tbl = _safe_ident(m.physical_table)
            with engine.begin() as conn:
                conn.execute(text(f"DROP TABLE IF EXISTS {tbl}"))
        _purge_entity_artifacts(s, "downstream_model", model_id)
        s.query(Schedule).filter_by(entity_id=model_id).delete()
        s.delete(m)
        s.commit()
        return ok({"deleted": model_id})
    finally:
        s.close()


@router.post("/downstream-models/{model_id}/materialize", tags=["下游模型"])
def materialize_downstream(model_id: int):
    """物化：CREATE TABLE dl_{code} AS <定义 SQL>；重复执行 = 重建刷新（幂等）。
    记录任务实例，成功后自动跑该模型全部启用质量规则（trigger=auto）"""
    s = get_session()
    try:
        m = _get_or_404(s, DownstreamModel, model_id, "下游模型")
        m_id, m_code = m.id, m.code
        inst = TaskInstance(task_type="materialize", entity_type="downstream_model",
                            entity_id=m_id, entity_code=m_code,
                            trigger="manual", status="RUNNING")
        s.add(inst)
        try:
            sql, params = gen.generate_downstream_sql(m)
            tbl = f"dl_{_safe_ident(m.code)}"
            m.definition_sql = sql
            with engine.begin() as conn:
                conn.execute(text(f"DROP TABLE IF EXISTS {tbl}"))
                conn.execute(text(f"CREATE TABLE {tbl} AS {sql}"), params)
                # 物化表按日期桶建索引，支撑下游查询/质量检查按时间过滤
                conn.execute(text(
                    f"CREATE INDEX IF NOT EXISTS ix_{_safe_ident(m.code)}_bucket "
                    f"ON {tbl} (date_bucket)"))
            with engine.connect() as conn:
                n = conn.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
            m.materialized = 1
            m.physical_table = tbl
            m.row_count = n
            inst.status = "SUCCESS"
            inst.detail = {"physical_table": tbl, "row_count": n}
            inst.finished_at = dt.datetime.now()
            _run_quality_checks(s, m_id)  # 物化成功自动质量检查（内部兜底，不抛错）
        except Exception as e:  # noqa: BLE001 - 失败写 FAILED 实例 + 告警
            s.rollback()
            inst = TaskInstance(task_type="materialize", entity_type="downstream_model",
                                entity_id=m_id, entity_code=m_code,
                                trigger="manual", status="FAILED",
                                error=str(e), finished_at=dt.datetime.now())
            s.add(inst)
            s.flush()  # 先落 id，供告警 source_id 引用
            _new_alert(s, "error", "task", inst.id,
                       f"物化任务失败 {m_code}: {e}")
            s.commit()
            if isinstance(e, ValueError):
                raise HTTPException(400, str(e))
            raise HTTPException(500, "物化失败，请查看服务端日志")
        s.commit()
        return ok({"physical_table": tbl, "row_count": n,
                   "task_instance_id": inst.id, "status": inst.status})
    finally:
        s.close()


def _reimport_default_start():
    """重导默认起点：3 个月前的当月 1 日（近 3 个月）"""
    today = dt.date.today()
    y, m = today.year, today.month - 3
    if m <= 0:
        m += 12
        y -= 1
    return dt.date(y, m, 1)


def _reimport_range(start_date: Optional[str] = None,
                    end_date: Optional[str] = None):
    """解析重导时间范围：默认近 3 个月（3 个月前当月 1 日 ~ 今天）；
    校验日期格式与起止顺序，非法抛 400"""
    try:
        start = dt.date.fromisoformat(start_date) if start_date else _reimport_default_start()
        end = dt.date.fromisoformat(end_date) if end_date else dt.date.today()
    except ValueError:
        raise HTTPException(400, "日期格式非法，需为 YYYY-MM-DD")
    if start > end:
        raise HTTPException(400, "开始日期不能晚于结束日期")
    return start, end


def _do_reimport(s, m, sb: str, eb: str):
    """按最新上游定义重算物化表区间数据（DELETE + INSERT 同一事务，原子）。
    重新生成定义 SQL = 读取上游逻辑模型/指标/维度最新口径；
    返回 (deleted, inserted, total)；上游定义非法时抛 ValueError"""
    sql, params = gen.generate_downstream_sql(m)
    m.definition_sql = sql
    tbl = _safe_ident(m.physical_table)
    params["r_s"], params["r_e"] = sb, eb
    with engine.begin() as conn:
        deleted = conn.execute(text(
            f"DELETE FROM {tbl} WHERE date_bucket >= :r_s AND date_bucket <= :r_e"),
            {"r_s": sb, "r_e": eb}).rowcount
        inserted = conn.execute(text(
            f"INSERT INTO {tbl} SELECT * FROM ( {sql} ) t "
            f"WHERE date_bucket >= :r_s AND date_bucket <= :r_e"), params).rowcount
    with engine.connect() as conn:
        total = conn.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
    m.row_count = total
    return deleted, inserted, total


@router.post("/downstream-models/{model_id}/reimport", tags=["下游模型"])
def reimport_downstream(model_id: int, start_date: Optional[str] = None,
                        end_date: Optional[str] = None):
    """重导：上游逻辑模型指标/维度更新上线后，按时间范围重建物化表数据。
    物化表区间内先 DELETE 再按最新定义重算 INSERT（同一事务，原子）；
    默认范围 = 近 3 个月（3 个月前当月 1 日 ~ 今天），可用参数覆盖。
    记录任务实例，成功后自动跑该模型全部启用质量规则（trigger=auto）"""
    s = get_session()
    try:
        m = _get_or_404(s, DownstreamModel, model_id, "下游模型")
        if not m.materialized or not m.physical_table:
            raise HTTPException(400, "请先物化，再执行数据重导")
        m_id, m_code = m.id, m.code
        start, end = _reimport_range(start_date, end_date)
        # 按模型粒度生成桶边界字符串（日/周/月），保证区间匹配 date_bucket
        fmt = GRANULARITY_FMT.get(m.granularity, "%Y-%m-%d")
        sb, eb = start.strftime(fmt), end.strftime(fmt)
        inst = TaskInstance(task_type="reimport", entity_type="downstream_model",
                            entity_id=m_id, entity_code=m_code,
                            trigger="manual", status="RUNNING")
        s.add(inst)
        try:
            deleted, inserted, total = _do_reimport(s, m, sb, eb)
            inst.status = "SUCCESS"
            inst.detail = {"start_date": start.isoformat(),
                           "end_date": end.isoformat(),
                           "deleted": deleted, "inserted": inserted,
                           "total_rows": total}
            inst.finished_at = dt.datetime.now()
            _run_quality_checks(s, m_id)  # 重导成功自动质量检查
        except Exception as e:  # noqa: BLE001 - 失败写 FAILED 实例 + 告警
            s.rollback()
            inst = TaskInstance(task_type="reimport", entity_type="downstream_model",
                                entity_id=m_id, entity_code=m_code,
                                trigger="manual", status="FAILED",
                                error=str(e), finished_at=dt.datetime.now())
            s.add(inst)
            s.flush()  # 先落 id，供告警 source_id 引用
            _new_alert(s, "error", "task", inst.id,
                       f"重导任务失败 {m_code}: {e}")
            s.commit()
            if isinstance(e, ValueError):
                raise HTTPException(400, str(e))
            raise HTTPException(500, "重导失败，请查看服务端日志")
        s.commit()
        return ok({"physical_table": _safe_ident(m.physical_table),
                   "start_date": start.isoformat(),
                   "end_date": end.isoformat(), "deleted": deleted,
                   "inserted": inserted, "total_rows": total,
                   "task_instance_id": inst.id, "status": inst.status})
    finally:
        s.close()


@router.post("/downstream-models/{model_id}/preview", tags=["下游模型"])
def preview_downstream(model_id: int, limit: int = 100):
    """执行定义 SQL 预览（不落地），返回前 limit 行"""
    s = get_session()
    try:
        m = _get_or_404(s, DownstreamModel, model_id, "下游模型")
        try:
            sql, params = gen.generate_downstream_sql(m)
        except (ValueError, MetricNotFoundError) as e:
            raise HTTPException(400, str(e))
        result = s.execute(text(sql), params)
        cols = list(result.keys())
        rows = [list(r) for r in result.fetchmany(limit)]
        return ok({"columns": cols, "rows": rows, "row_count": len(rows)})
    finally:
        s.close()


@router.get("/downstream-models/{model_id}/data", tags=["下游模型"])
def downstream_data(model_id: int, page: int = 1, page_size: int = 100):
    """查询物化表数据（分页）"""
    s = get_session()
    try:
        m = _get_or_404(s, DownstreamModel, model_id, "下游模型")
        if not m.materialized or not m.physical_table:
            raise HTTPException(400, "下游模型尚未物化，请先执行物化")
        tbl = _safe_ident(m.physical_table)
        pg, pgs = _page_clamped(page), _page_size_clamped(page_size)
        total = s.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
        rows_sql = text(
            f"SELECT * FROM {tbl} ORDER BY date_bucket "
            f"LIMIT {pgs} OFFSET {(pg - 1) * pgs}")
        result = s.execute(rows_sql)
        return ok({"total": total, "page": pg, "page_size": pgs,
                   "columns": list(result.keys()),
                   "rows": [list(r) for r in result.fetchall()]})
    finally:
        s.close()


# ===========================================================================
# 5.10 下游应用 + 数据集 + 开放 API（下游消费面）
# 数据集双源：downstream_model（读物化表 dl_xxx）/ metric_query（动态 SQL 实时计算）
# 开放 API：X-App-Key + X-App-Secret 请求头认证；openapi_router 仅挂载 /openapi
# ===========================================================================

class DownstreamAppIn(BaseModel):
    code: str
    name: str
    description: str = ""
    status: str = "ENABLED"


class DatasetIn(BaseModel):
    code: str
    name: str
    source_type: str
    source_model_id: Optional[int] = None
    metric_codes: list = []
    dim_codes: list = []
    granularity: str = "day"
    description: str = ""


class GrantIn(BaseModel):
    app_id: int


# ---- 下游应用 CRUD ---------------------------------------------------------

@router.post("/downstream-apps", tags=["下游应用"])
def create_downstream_app(body: DownstreamAppIn):
    s = get_session()
    try:
        _check_code(s, DownstreamApp, body.code)
        if body.status not in APP_STATUSES:
            raise HTTPException(400, f"非法状态: {body.status}")
        app = DownstreamApp(code=body.code, name=body.name,
                            description=body.description, status=body.status,
                            appkey=secrets.token_hex(10),
                            appsecret=secrets.token_urlsafe(24))
        s.add(app)
        s.commit()
        return ok({"id": app.id, "appkey": app.appkey, "appsecret": app.appsecret})
    finally:
        s.close()


@router.get("/downstream-apps", tags=["下游应用"])
def list_downstream_apps(page: int = 1, page_size: int = 20, keyword: str = ""):
    s = get_session()
    try:
        q = s.query(DownstreamApp).order_by(DownstreamApp.id.desc())
        if keyword:
            like = f"%{keyword}%"
            q = q.filter(or_(DownstreamApp.code.like(_like_escape(like), escape="\\"),
                             DownstreamApp.name.like(_like_escape(like), escape="\\")))
        total = q.count()
        rows = (q.offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        items = []
        for a in rows:
            items.append({
                "id": a.id, "code": a.code, "name": a.name,
                "appkey": a.appkey,
                "status": a.status, "description": a.description,
                "call_count": s.query(ApiCallLog).filter_by(app_id=a.id).count(),
                "dataset_count": s.query(AppDatasetGrant).filter_by(app_id=a.id).count(),
                "created_at": (a.created_at.strftime("%Y-%m-%d %H:%M")
                               if a.created_at else ""),
            })
        return ok({"items": items, "total": total, "page": page, "page_size": page_size})
    finally:
        s.close()


@router.get("/downstream-apps/{app_id}", tags=["下游应用"])
def get_downstream_app(app_id: int):
    s = get_session()
    try:
        a = _get_or_404(s, DownstreamApp, app_id, "下游应用")
        ds = (s.query(Dataset)
              .join(AppDatasetGrant, AppDatasetGrant.dataset_id == Dataset.id)
              .filter(AppDatasetGrant.app_id == a.id).all())
        return ok({
            "id": a.id, "code": a.code, "name": a.name,
            "appkey": a.appkey,
            "status": a.status, "description": a.description,
            "dataset_ids": [d.id for d in ds],
            "datasets": [{"id": d.id, "code": d.code, "name": d.name} for d in ds],
        })
    finally:
        s.close()


@router.put("/downstream-apps/{app_id}", tags=["下游应用"])
def update_downstream_app(app_id: int, body: DownstreamAppIn):
    s = get_session()
    try:
        a = _get_or_404(s, DownstreamApp, app_id, "下游应用")
        _check_code(s, DownstreamApp, body.code, exclude_id=app_id)
        if body.status not in APP_STATUSES:
            raise HTTPException(400, f"非法状态: {body.status}")
        for k, v in body.dict().items():
            setattr(a, k, v)
        s.commit()
        return ok({"id": a.id})
    finally:
        s.close()


@router.post("/downstream-apps/{app_id}/reset-secret", tags=["下游应用"])
def reset_app_secret(app_id: int):
    """重置 AppSecret（AppKey 不变），旧密钥立即失效"""
    s = get_session()
    try:
        a = _get_or_404(s, DownstreamApp, app_id, "下游应用")
        a.appsecret = secrets.token_urlsafe(24)
        s.commit()
        return ok({"appkey": a.appkey, "appsecret": a.appsecret})
    finally:
        s.close()


@router.delete("/downstream-apps/{app_id}", tags=["下游应用"])
def delete_downstream_app(app_id: int):
    s = get_session()
    try:
        a = _get_or_404(s, DownstreamApp, app_id, "下游应用")
        s.query(ApiCallLog).filter_by(app_id=a.id).delete()
        s.query(AppDatasetGrant).filter_by(app_id=a.id).delete()
        s.delete(a)
        s.commit()
        return ok({"deleted": app_id})
    finally:
        s.close()


# ---- 数据集 CRUD + 授权 ----------------------------------------------------

def _validate_dataset(s, body):
    if body.source_type not in DATASET_SOURCES:
        raise HTTPException(400, f"非法数据源类型: {body.source_type}")
    if body.granularity not in GRANULARITY_FMT:
        raise HTTPException(400, f"非法日期粒度: {body.granularity}")
    if body.source_type == "downstream_model":
        if not body.source_model_id:
            raise HTTPException(400, "downstream_model 类型必须指定来源下游模型")
        _get_or_404(s, DownstreamModel, body.source_model_id, "下游模型")
    else:
        for c in body.metric_codes:
            try:
                gen.find_metric(c)
            except MetricNotFoundError:
                raise HTTPException(400, f"指标不存在: {c}")


@router.post("/datasets", tags=["数据集"])
def create_dataset(body: DatasetIn):
    s = get_session()
    try:
        _check_code(s, Dataset, body.code)
        _validate_dataset(s, body)
        d = Dataset(**body.dict())
        s.add(d)
        s.commit()
        return ok({"id": d.id, "code": d.code})
    finally:
        s.close()


@router.get("/datasets", tags=["数据集"])
def list_datasets(page: int = 1, page_size: int = 20, keyword: str = ""):
    s = get_session()
    try:
        q = s.query(Dataset).order_by(Dataset.id.desc())
        if keyword:
            like = f"%{keyword}%"
            q = q.filter(or_(Dataset.code.like(_like_escape(like), escape="\\"), Dataset.name.like(_like_escape(like), escape="\\")))
        total = q.count()
        rows = (q.offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        items = []
        for d in rows:
            granted_apps = (s.query(DownstreamApp)
                            .join(AppDatasetGrant,
                                  AppDatasetGrant.app_id == DownstreamApp.id)
                            .filter(AppDatasetGrant.dataset_id == d.id).all())
            items.append({
                "id": d.id, "code": d.code, "name": d.name,
                "source_type": d.source_type,
                "source_model_id": d.source_model_id,
                "source_model_name": d.source_model_name,
                "metric_codes": d.metric_codes or [],
                "dim_codes": d.dim_codes or [],
                "granularity": d.granularity, "description": d.description,
                "granted_app_ids": [a.id for a in granted_apps],
                "granted_app_count": len(granted_apps),
            })
        return ok({"items": items, "total": total, "page": page, "page_size": page_size})
    finally:
        s.close()


@router.get("/datasets/{dataset_id}", tags=["数据集"])
def get_dataset(dataset_id: int):
    s = get_session()
    try:
        d = _get_or_404(s, Dataset, dataset_id, "数据集")
        granted_apps = (s.query(DownstreamApp)
                        .join(AppDatasetGrant,
                              AppDatasetGrant.app_id == DownstreamApp.id)
                        .filter(AppDatasetGrant.dataset_id == d.id).all())
        return ok({
            "id": d.id, "code": d.code, "name": d.name,
            "source_type": d.source_type,
            "source_model_id": d.source_model_id,
            "source_model_name": d.source_model_name,
            "metric_codes": d.metric_codes or [],
            "dim_codes": d.dim_codes or [],
            "granularity": d.granularity, "description": d.description,
            "granted_app_ids": [a.id for a in granted_apps],
            "granted_apps": [{"id": a.id, "code": a.code, "name": a.name}
                             for a in granted_apps],
        })
    finally:
        s.close()


@router.put("/datasets/{dataset_id}", tags=["数据集"])
def update_dataset(dataset_id: int, body: DatasetIn):
    s = get_session()
    try:
        d = _get_or_404(s, Dataset, dataset_id, "数据集")
        _check_code(s, Dataset, body.code, exclude_id=dataset_id)
        _validate_dataset(s, body)
        for k, v in body.dict().items():
            setattr(d, k, v)
        s.commit()
        return ok({"id": d.id})
    finally:
        s.close()


@router.delete("/datasets/{dataset_id}", tags=["数据集"])
def delete_dataset(dataset_id: int):
    s = get_session()
    try:
        d = _get_or_404(s, Dataset, dataset_id, "数据集")
        s.query(AppDatasetGrant).filter_by(dataset_id=d.id).delete()
        s.delete(d)
        s.commit()
        return ok({"deleted": dataset_id})
    finally:
        s.close()


@router.post("/datasets/{dataset_id}/grant", tags=["数据集"])
def grant_dataset(dataset_id: int, body: GrantIn):
    """授权数据集给下游应用（幂等，重复授权自动去重）"""
    s = get_session()
    try:
        d = _get_or_404(s, Dataset, dataset_id, "数据集")
        a = _get_or_404(s, DownstreamApp, body.app_id, "下游应用")
        exists = (s.query(AppDatasetGrant)
                  .filter_by(dataset_id=d.id, app_id=a.id).first())
        if not exists:
            s.add(AppDatasetGrant(dataset_id=d.id, app_id=a.id))
            s.commit()
        return ok({"dataset_id": d.id, "app_id": a.id, "granted": not exists})
    finally:
        s.close()


@router.delete("/datasets/{dataset_id}/grant/{app_id}", tags=["数据集"])
def revoke_dataset(dataset_id: int, app_id: int):
    """撤销授权：应用将无法再调用该数据集"""
    s = get_session()
    try:
        g = (s.query(AppDatasetGrant)
             .filter_by(dataset_id=dataset_id, app_id=app_id).first())
        if g:
            s.delete(g)
            s.commit()
        return ok({"revoked": bool(g)})
    finally:
        s.close()


# ---- 调用监控（管理端视角） ------------------------------------------------

@router.get("/openapi/stats", tags=["开放 API"])
def openapi_stats():
    s = get_session()
    try:
        total = s.query(ApiCallLog).count()
        total_rows = s.query(ApiCallLog).with_entities(
            func.coalesce(func.sum(ApiCallLog.row_count), 0)).scalar() or 0
        by_app = []
        for a in s.query(DownstreamApp).all():
            q = s.query(ApiCallLog).filter_by(app_id=a.id)
            by_app.append({
                "app_id": a.id, "app_code": a.code, "app_name": a.name,
                "calls": q.count(),
                "rows": q.with_entities(
                    func.coalesce(func.sum(ApiCallLog.row_count), 0)).scalar() or 0,
            })
        by_dataset = []
        for d in s.query(Dataset).all():
            q = s.query(ApiCallLog).filter_by(dataset_id=d.id)
            by_dataset.append({
                "dataset_id": d.id, "dataset_code": d.code, "dataset_name": d.name,
                "calls": q.count(),
                "rows": q.with_entities(
                    func.coalesce(func.sum(ApiCallLog.row_count), 0)).scalar() or 0,
            })
        return ok({"total_calls": total, "total_rows": total_rows,
                   "by_app": by_app, "by_dataset": by_dataset})
    finally:
        s.close()


@router.get("/openapi/logs", tags=["开放 API"])
def openapi_logs(page: int = 1, page_size: int = 20,
                 app_id: Optional[int] = None):
    s = get_session()
    try:
        q = s.query(ApiCallLog).order_by(ApiCallLog.id.desc())
        if app_id:
            q = q.filter_by(app_id=app_id)
        total = q.count()
        rows = (q.offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        items = []
        for log in rows:
            app = s.query(DownstreamApp).get(log.app_id)
            ds = s.query(Dataset).get(log.dataset_id)
            items.append({
                "id": log.id,
                "app_code": app.code if app else str(log.app_id),
                "app_name": app.name if app else "",
                "dataset_code": ds.code if ds else str(log.dataset_id),
                "dataset_name": ds.name if ds else "",
                "row_count": log.row_count, "duration_ms": log.duration_ms,
                "status": log.status,
                "called_at": (log.called_at.strftime("%Y-%m-%d %H:%M:%S")
                              if log.called_at else ""),
            })
        return ok({"items": items, "total": total, "page": page, "page_size": page_size})
    finally:
        s.close()


# ---- 开放 API（下游消费面：独立路由，仅挂载 /openapi） ----------------------

def _auth_app(s, request: Request):
    """AppKey + AppSecret 认证（常数时间比较，防时序攻击）"""
    key = request.headers.get("X-App-Key", "")
    secret = request.headers.get("X-App-Secret", "")
    if not key or not secret:
        raise HTTPException(401, "缺少认证头: X-App-Key / X-App-Secret")
    app = s.query(DownstreamApp).filter_by(appkey=key).first()
    if not app:
        raise HTTPException(401, "认证失败: AppKey/AppSecret 不匹配")
    stored, given = app.appsecret or "", secret or ""
    if len(stored) != len(given) or not hmac.compare_digest(stored, given):
        raise HTTPException(401, "认证失败: AppKey/AppSecret 不匹配")
    if app.status != "ENABLED":
        raise HTTPException(401, f"应用已停用: {app.code}")
    return app


def _log_call(s, app_id, dataset_id, row_count, duration_ms, status):
    s.add(ApiCallLog(app_id=app_id, dataset_id=dataset_id,
                     row_count=row_count, duration_ms=duration_ms, status=status))


@openapi_router.get("/v1/datasets")
def openapi_datasets(request: Request):
    """当前应用（认证后）有权限的数据集列表"""
    s = get_session()
    try:
        app = _auth_app(s, request)
        datasets = (s.query(Dataset)
                    .join(AppDatasetGrant, AppDatasetGrant.dataset_id == Dataset.id)
                    .filter(AppDatasetGrant.app_id == app.id).all())
        return ok({"datasets": [{
            "code": d.code, "name": d.name, "source_type": d.source_type,
            "granularity": d.granularity,
            "metric_codes": d.metric_codes or [],
            "dim_codes": d.dim_codes or [],
            "description": d.description,
        } for d in datasets]})
    finally:
        s.close()


@openapi_router.get("/v1/datasets/{code}/data")
def openapi_dataset_data(code: str, request: Request,
                         page: int = 1, page_size: int = 100,
                         start_date: Optional[str] = None,
                         end_date: Optional[str] = None):
    """数据集数据调用：物化表直读 / 指标实时计算，分页返回
    downstream_model 源：读物化表 dl_{code}（未物化 -> 400）
    metric_query 源：execute_multi 实时计算，支持可选日期覆盖
    每次调用写入 ApiCallLog（行数/耗时/状态）"""
    s = get_session()
    t0 = time.time()
    app = ds = None
    try:
        app = _auth_app(s, request)
        ds = s.query(Dataset).filter_by(code=code).first()
        if not ds:
            raise HTTPException(404, f"数据集不存在: {code}")
        granted = (s.query(AppDatasetGrant)
                   .filter_by(app_id=app.id, dataset_id=ds.id).first())
        if not granted:
            raise HTTPException(403, f"应用 {app.code} 未获授权访问数据集 {code}")
        if ds.source_type == "downstream_model":
            dm = _get_or_404(s, DownstreamModel, ds.source_model_id, "下游模型")
            if not dm.materialized or not dm.physical_table:
                raise HTTPException(400, "数据集来源下游模型尚未物化，请先执行物化")
            tbl = _safe_ident(dm.physical_table)
            pg, pgs = _page_clamped(page), _page_size_clamped(page_size)
            total = s.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
            result = s.execute(text(
                f"SELECT * FROM {tbl} ORDER BY date_bucket "
                f"LIMIT {pgs} OFFSET {(pg - 1) * pgs}"))
            ret = {"columns": list(result.keys()),
                   "rows": [list(r) for r in result.fetchall()],
                   "total": total, "page": pg, "page_size": pgs}
        else:
            meta, columns, rows, sql = gen.execute_multi(
                ds.metric_codes or [], ds.dim_codes or [],
                start_date, end_date, ds.granularity)
            pg, pgs = _page_clamped(page), _page_size_clamped(page_size)
            start = (pg - 1) * pgs
            ret = {"columns": columns, "rows": rows[start:start + pgs],
                   "total": len(rows), "page": pg, "page_size": pgs,
                   "sql": sql}
        _log_call(s, app.id, ds.id, len(ret["rows"]),
                  int((time.time() - t0) * 1000), "success")
        s.commit()
        return ok(ret)
    except MetricNotFoundError as e:
        s.rollback()
        if app is not None and ds is not None:
            try:
                _log_call(s, app.id, ds.id, 0, int((time.time() - t0) * 1000),
                          "error:400")
                s.commit()
            except Exception:
                s.rollback()
        raise HTTPException(400, str(e))
    except HTTPException as e:
        s.rollback()
        if app is not None and ds is not None:
            try:
                _log_call(s, app.id, ds.id, 0, int((time.time() - t0) * 1000),
                          f"error:{e.status_code}")
                s.commit()
            except Exception:
                s.rollback()
        raise
    finally:
        s.close()


# ===========================================================================
# 血缘追溯（P2）：物理表/字段 -> 原子 -> 派生 -> 复合
# ===========================================================================

def _atomic_node(m: AtomicMetric):
    return {"id": f"atomic:{m.code}", "type": "atomic", "label": m.name, "code": m.code}


def _physical_nodes(a: AtomicMetric, nodes: list, edges: list):
    p = a.process
    nodes.append({"id": f"table:{p.physical_table}", "type": "table",
                  "label": p.physical_table, "code": p.physical_table})
    nodes.append({"id": f"field:{p.physical_table}.{a.physical_field}",
                  "type": "field", "label": a.physical_field,
                  "code": a.physical_field})
    edges.append({"from": f"field:{p.physical_table}.{a.physical_field}",
                  "to": f"atomic:{a.code}"})
    edges.append({"from": f"table:{p.physical_table}",
                  "to": f"field:{p.physical_table}.{a.physical_field}"})


def _lineage_upstream(s, code: str, nodes: list, edges: list, seen: set):
    """向上追溯（根因分析）：复合 <- 派生 <- 原子 <- 物理表/字段"""
    comp = s.query(CompositeMetric).filter_by(code=code).first()
    if comp:
        nodes.append({"id": f"composite:{code}", "type": "composite",
                      "label": comp.name, "code": code})
        for ref in comp.ref_codes or []:
            if ("composite", code, "derived", ref) in seen:
                continue
            seen.add(("composite", code, "derived", ref))
            dm = s.query(DerivedMetric).filter_by(code=ref).first()
            am = s.query(AtomicMetric).filter_by(code=ref).first()
            if dm:
                nodes.append({"id": f"derived:{ref}", "type": "derived",
                              "label": dm.name, "code": ref})
                edges.append({"from": f"derived:{ref}", "to": f"composite:{code}"})
                _lineage_upstream(s, ref, nodes, edges, seen)
            elif am:
                nodes.append(_atomic_node(am))
                edges.append({"from": f"atomic:{ref}", "to": f"composite:{code}"})
                _physical_nodes(am, nodes, edges)

    der = s.query(DerivedMetric).filter_by(code=code).first()
    if der:
        if not any(n["id"] == f"derived:{code}" for n in nodes):
            nodes.append({"id": f"derived:{code}", "type": "derived",
                          "label": der.name, "code": code})
        a = der.atomic
        nodes.append(_atomic_node(a))
        edges.append({"from": f"atomic:{a.code}", "to": f"derived:{code}"})
        _physical_nodes(a, nodes, edges)

    atom = s.query(AtomicMetric).filter_by(code=code).first()
    if atom:
        if not any(n["id"] == f"atomic:{code}" for n in nodes):
            nodes.append(_atomic_node(atom))
        _physical_nodes(atom, nodes, edges)


def _lineage_downstream(s, code: str, nodes: list, edges: list):
    """向下影响（影响分析）：原子/派生 -> 下游派生/复合"""
    has_atomic = bool(s.query(AtomicMetric).filter_by(code=code).first())
    composites = s.query(CompositeMetric).all()  # 提到循环外，避免 N+1 重复全表查询
    for dm in s.query(DerivedMetric).all():
        if has_atomic and code == dm.atomic.code:
            if not any(n["id"] == f"derived:{dm.code}" for n in nodes):
                nodes.append({"id": f"derived:{dm.code}", "type": "derived",
                              "label": dm.name, "code": dm.code})
            edges.append({"from": f"atomic:{code}", "to": f"derived:{dm.code}"})
        elif code == dm.code and not any(n["id"] == f"derived:{dm.code}" for n in nodes):
            nodes.append({"id": f"derived:{dm.code}", "type": "derived",
                          "label": dm.name, "code": dm.code})
        for cm in composites:
            # 复合引用派生：派生节点已出现时连边
            if dm.code in (cm.ref_codes or []) and \
                    any(n["id"] == f"derived:{dm.code}" for n in nodes):
                if not any(n["id"] == f"composite:{cm.code}" for n in nodes):
                    nodes.append({"id": f"composite:{cm.code}", "type": "composite",
                                  "label": cm.name, "code": cm.code})
                edges.append({"from": f"derived:{dm.code}",
                              "to": f"composite:{cm.code}"})
            # 复合直接引用原子：原子节点出现时连边
            if has_atomic and code in (cm.ref_codes or []) and \
                    any(n["id"] == f"atomic:{code}" for n in nodes):
                if not any(n["id"] == f"composite:{cm.code}" for n in nodes):
                    nodes.append({"id": f"composite:{cm.code}", "type": "composite",
                                  "label": cm.name, "code": cm.code})
                edges.append({"from": f"atomic:{code}",
                              "to": f"composite:{cm.code}"})


@router.get("/lineage/tables", tags=["血缘"])
def lineage_tables():
    """表级血缘（全量）：物理表 -> 逻辑模型 -> 下游模型 -> 物化表
    逻辑模型的物理表关联由 physical_table + join_config 推导"""
    s = get_session()
    try:
        nodes, edges = [], []
        seen_tables = set()

        def add_table(tbl: str):
            if tbl and tbl not in seen_tables:
                seen_tables.add(tbl)
                nodes.append({"id": f"table:{tbl}", "type": "table",
                              "label": tbl, "code": tbl})

        # 指标口径涉及的物理表
        for p in s.query(BusinessProcess).all():
            add_table(p.physical_table)
        for d in s.query(Dimension).all():
            add_table(d.physical_table)

        # 逻辑模型：物理表 -> 模型（含 join_config 关联表）
        for lm in s.query(LogicalModel).all():
            nid = f"model:{lm.code}"
            nodes.append({"id": nid, "type": "logical_model",
                          "label": lm.name, "code": lm.code})
            add_table(lm.physical_table)
            edges.append({"from": f"table:{lm.physical_table}", "to": nid})
            for j in lm.join_config or []:
                tbl = j.get("table", "")
                add_table(tbl)
                edges.append({"from": f"table:{tbl}", "to": nid})

        # 下游模型：逻辑模型 -> 下游模型 -> 物化表
        for dm in s.query(DownstreamModel).all():
            nid = f"downstream:{dm.code}"
            nodes.append({"id": nid, "type": "downstream_model",
                          "label": dm.name, "code": dm.code})
            edges.append({"from": f"model:{dm.source_model.code}", "to": nid})
            if dm.materialized and dm.physical_table:
                add_table(dm.physical_table)
                edges.append({"from": nid, "to": f"table:{dm.physical_table}"})

        return ok({"nodes": nodes, "edges": edges})
    finally:
        s.close()


@router.get("/lineage/{code}", tags=["血缘"])
def get_lineage(code: str):
    """全链路血缘：物理表/字段 -> 原子 -> 派生 -> 复合（影响分析 + 根因追溯）"""
    s = get_session()
    try:
        nodes, edges = [], []
        _lineage_upstream(s, code, nodes, edges, set())
        if not nodes:
            raise HTTPException(404, f"指标不存在: {code}")
        _lineage_downstream(s, code, nodes, edges)

        seen_ids, uniq_nodes = set(), []
        for n in nodes:
            if n["id"] not in seen_ids:
                seen_ids.add(n["id"])
                uniq_nodes.append(n)
        uniq_edges, seen_e = [], set()
        for e in edges:
            k = (e["from"], e["to"])
            if k not in seen_e:
                seen_e.add(k)
                uniq_edges.append(e)
        return ok({"nodes": uniq_nodes, "edges": uniq_edges})
    finally:
        s.close()


# ===========================================================================
# 5.11 任务重导：对象下游血缘 + 重导执行计划生成 + 确认执行
# 下游模型只使用原子/派生指标（复合指标不进下游），故对象类型限定 4 类
# ===========================================================================

REIMPORT_OBJECT_TYPES = ("atomic_metric", "derived_metric", "dimension", "logical_model")


def _resolve_reimport_object(s, object_type: str, object_id: int):
    """定位重导对象，返回 {type, code, name}；对象类型非法/不存在时抛 4xx"""
    if object_type not in REIMPORT_OBJECT_TYPES:
        raise HTTPException(400, f"不支持的对象类型: {object_type}，"
                                 f"可选: {', '.join(REIMPORT_OBJECT_TYPES)}")
    if object_type == "atomic_metric":
        obj = _get_or_404(s, AtomicMetric, object_id, "原子指标")
    elif object_type == "derived_metric":
        obj = _get_or_404(s, DerivedMetric, object_id, "派生指标")
    elif object_type == "dimension":
        obj = _get_or_404(s, Dimension, object_id, "维度")
    else:
        obj = _get_or_404(s, LogicalModel, object_id, "逻辑模型")
    return {"type": object_type, "id": object_id, "code": obj.code, "name": obj.name}


def _find_impacted_downstreams(s, object_type: str, object_id: int):
    """反查受影响的下游模型（任务血缘）：对象 ->(派生中介)-> 逻辑模型 -> 下游模型。
    返回 (object_node, [{"dm": DownstreamModel, "chain": [节点...]}, ...])；
    chain 节点形如 {type, code, name}，type ∈ 对象类型/downstream"""
    object_node = _resolve_reimport_object(s, object_type, object_id)

    if object_type == "logical_model":
        # 逻辑模型：source_model_id 外键直接反查其下游模型
        hits = [{"dm": ds, "chain": [object_node]}
                for ds in s.query(DownstreamModel)
                .filter_by(source_model_id=object_id).all()]
        for h in hits:
            h["chain"].append({"type": "downstream",
                               "code": h["dm"].code, "name": h["dm"].name})
        return object_node, hits

    # 命中编码：下游 metrics[].metric_code 匹配即受影响；
    # 原子指标/维度还会经派生指标间接命中（派生被下游引用）
    hit_codes = set()
    if object_type == "atomic_metric":
        hit_codes.add(object_node["code"])
        for dm in s.query(DerivedMetric).filter_by(atomic_id=object_id).all():
            hit_codes.add(dm.code)
    elif object_type == "derived_metric":
        hit_codes.add(object_node["code"])
    elif object_type == "dimension":
        for dm in s.query(DerivedMetric).all():
            if object_node["code"] in (dm.dim_codes or []):
                hit_codes.add(dm.code)

    results = []
    for ds in s.query(DownstreamModel).all():
        used, mid_node = False, None
        for it in (ds.metrics or []):
            mc = it.get("metric_code")
            if mc in hit_codes:
                used = True
                # 指标是派生且非对象本身 -> 血缘链补派生中介节点
                if mc != object_node["code"] and \
                        object_type in ("atomic_metric", "dimension"):
                    dm = s.query(DerivedMetric).filter_by(code=mc).first()
                    if dm:
                        mid_node = {"type": "derived_metric",
                                    "code": dm.code, "name": dm.name}
                break
            if object_type == "dimension" and \
                    object_node["code"] in (it.get("dim_codes") or []):
                used = True
                break
        if not used:
            continue
        chain = [object_node]
        if mid_node:
            chain.append(mid_node)
        if object_type != "logical_model" and ds.source_model:
            chain.append({"type": "logical_model", "code": ds.source_model.code,
                          "name": ds.source_model.name})
        chain.append({"type": "downstream", "code": ds.code, "name": ds.name})
        results.append({"dm": ds, "chain": chain})
    return object_node, results


@router.get("/reimport/impact", tags=["任务重导"])
def reimport_impact(object_type: str, object_id: int):
    """对象下游任务血缘：给定模型/字段（原子/派生指标、维度、逻辑模型），
    反查所有受影响的下游模型及血缘链，供任务重导页展示"""
    s = get_session()
    try:
        object_node, hits = _find_impacted_downstreams(s, object_type, object_id)
        return ok({
            "object": object_node,
            "downstreams": [{
                "id": h["dm"].id, "code": h["dm"].code, "name": h["dm"].name,
                "source_model_code": h["dm"].source_model.code
                if h["dm"].source_model else None,
                "granularity": h["dm"].granularity,
                "materialized": bool(h["dm"].materialized),
                "physical_table": h["dm"].physical_table,
                "row_count": h["dm"].row_count,
                "chain": h["chain"],
            } for h in hits],
        })
    finally:
        s.close()


class ReimportPlanIn(BaseModel):
    object_type: str
    object_id: int
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    downstream_ids: Optional[list] = None


@router.post("/reimport/plan", tags=["任务重导"])
def reimport_plan(body: ReimportPlanIn):
    """生成重导执行计划：列出受影响下游模型与预估影响行数
    （默认近 3 个月；downstream_ids 可过滤只计划部分模型）"""
    s = get_session()
    try:
        _, hits = _find_impacted_downstreams(s, body.object_type, body.object_id)
        start, end = _reimport_range(body.start_date, body.end_date)
        items = []
        for h in hits:
            m = h["dm"]
            if body.downstream_ids and m.id not in body.downstream_ids:
                continue
            # 预估删除行数：物化表区间内现有行数（未物化无法预估）
            est = None
            if m.materialized and m.physical_table:
                fmt = GRANULARITY_FMT.get(m.granularity, "%Y-%m-%d")
                sb, eb = start.strftime(fmt), end.strftime(fmt)
                with engine.connect() as conn:
                    est = conn.execute(text(
                        f"SELECT COUNT(*) FROM {_safe_ident(m.physical_table)} "
                        f"WHERE date_bucket >= :r_s AND date_bucket <= :r_e"),
                        {"r_s": sb, "r_e": eb}).scalar()
            items.append({
                "id": m.id, "code": m.code, "name": m.name,
                "granularity": m.granularity,
                "materialized": bool(m.materialized),
                "row_count": m.row_count, "estimated_deleted": est,
            })
        return ok({"start_date": start.isoformat(), "end_date": end.isoformat(),
                   "items": items})
    finally:
        s.close()


class ReimportExecuteIn(BaseModel):
    downstream_ids: list
    start_date: Optional[str] = None
    end_date: Optional[str] = None


@router.post("/reimport/plan/execute", tags=["任务重导"])
def reimport_plan_execute(body: ReimportExecuteIn):
    """确认执行重导计划：逐个下游模型独立事务重导（一个失败不阻断其余）；
    每个模型写任务实例，成功后自动质量检查"""
    if not body.downstream_ids:
        raise HTTPException(400, "请至少选择一个下游模型")
    s = get_session()
    try:
        start, end = _reimport_range(body.start_date, body.end_date)
        results = []
        for mid in body.downstream_ids:
            m = _get_or_404(s, DownstreamModel, mid, "下游模型")
            m_id, m_code = m.id, m.code
            if not m.materialized or not m.physical_table:
                results.append({"id": mid, "code": m.code, "status": "skipped",
                                "deleted": None, "inserted": None,
                                "total_rows": None, "message": "请先物化，再执行数据重导"})
                continue
            fmt = GRANULARITY_FMT.get(m.granularity, "%Y-%m-%d")
            sb, eb = start.strftime(fmt), end.strftime(fmt)
            inst = TaskInstance(task_type="reimport", entity_type="downstream_model",
                                entity_id=m_id, entity_code=m_code,
                                trigger="manual", status="RUNNING")
            s.add(inst)
            try:
                deleted, inserted, total = _do_reimport(s, m, sb, eb)
                inst.status = "SUCCESS"
                inst.detail = {"start_date": start.isoformat(),
                               "end_date": end.isoformat(),
                               "deleted": deleted, "inserted": inserted,
                               "total_rows": total}
                inst.finished_at = dt.datetime.now()
                _run_quality_checks(s, m_id)  # 重导成功自动质量检查
                s.commit()  # 每模型独立提交，失败不阻断其余
                results.append({"id": mid, "code": m.code, "status": "ok",
                                "deleted": deleted, "inserted": inserted,
                                "total_rows": total, "message": None})
            except Exception as e:  # noqa: BLE001 - 单模型失败不阻断批量执行
                s.rollback()
                inst = TaskInstance(task_type="reimport", entity_type="downstream_model",
                                    entity_id=m_id, entity_code=m_code,
                                    trigger="manual", status="FAILED",
                                    error=str(e), finished_at=dt.datetime.now())
                s.add(inst)
                s.flush()  # 先落 id，供告警 source_id 引用
                _new_alert(s, "error", "task", inst.id,
                           f"重导任务失败 {m_code}: {e}")
                s.commit()
                results.append({"id": mid, "code": m.code, "status": "error",
                                "deleted": None, "inserted": None,
                                "total_rows": None, "message": str(e)})
        return ok({"start_date": start.isoformat(), "end_date": end.isoformat(),
                   "results": results})
    finally:
        s.close()


# ===========================================================================
# 治理与生命周期：版本快照/回滚、审批发布流、变更影响评估、标签、编码规范
# ===========================================================================

class ApprovalIn(BaseModel):
    entity_type: str
    entity_id: int
    comment: str = ""


class ApprovalReviewIn(BaseModel):
    comment: str = ""


class EntityTagsIn(BaseModel):
    entity_type: str
    entity_id: int
    tags: list = []


APPROVAL_ENTITY_TYPES = ("atomic_metric", "derived_metric", "composite_metric")


@router.get("/metric-versions", tags=["治理"])
def list_metric_versions(entity_type: str, entity_id: int,
                         page: int = 1, page_size: int = 20):
    """实体版本历史：每次变更前存档的快照列表（最新在前）"""
    s = get_session()
    try:
        q = (s.query(MetricVersion).filter_by(entity_type=entity_type,
                                              entity_id=entity_id)
             .order_by(MetricVersion.id.desc()))
        total = q.count()
        rows = (q.offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        items = [{
            "id": v.id, "version_no": v.version_no,
            "change_type": v.change_type, "change_note": v.change_note,
            "snapshot": json.loads(v.snapshot),
            "created_at": (v.created_at.strftime("%Y-%m-%d %H:%M:%S")
                           if v.created_at else ""),
        } for v in rows]
        return ok({"items": items, "total": total, "page": page,
                   "page_size": page_size})
    finally:
        s.close()


@router.post("/metric-versions/{version_id}/rollback", tags=["治理"])
def rollback_version(version_id: int):
    """回滚：将实体恢复至指定版本快照，并生成一条 rollback 版本记录"""
    s = get_session()
    try:
        v = _get_or_404(s, MetricVersion, version_id, "版本")
        obj = _entity_by_type(s, v.entity_type, v.entity_id)
        if not obj:
            raise HTTPException(404, f"实体不存在: {v.entity_type} id={v.entity_id}")
        _snapshot_version(s, v.entity_type, obj, "rollback",
                          f"回滚至 {v.version_no}")
        snap = json.loads(v.snapshot)
        for k, val in snap.items():
            if hasattr(obj, k) and k not in ("id", "created_at", "updated_at"):
                setattr(obj, k, val)
        s.commit()
        return ok({"id": v.id, "entity_type": v.entity_type,
                   "entity_id": v.entity_id, "version_no": v.version_no})
    finally:
        s.close()


_VERSION_DIFF_SKIP = ("id", "created_at", "updated_at")


@router.get("/metric-versions/compare", tags=["治理"])
def compare_versions(entity_type: str, entity_id: int,
                     a: int, b: int):
    """口径对比：逐字段 diff 两个版本快照（可对比发布前后口径变化）"""
    s = get_session()
    try:
        va = s.query(MetricVersion).filter_by(id=a, entity_type=entity_type,
                                              entity_id=entity_id).first()
        vb = s.query(MetricVersion).filter_by(id=b, entity_type=entity_type,
                                              entity_id=entity_id).first()
        if not va or not vb:
            raise HTTPException(404, "版本不存在（或不属于该实体）")
        snap_a = json.loads(va.snapshot)
        snap_b = json.loads(vb.snapshot)
        fields = []
        for k in sorted(set(snap_a) | set(snap_b)):
            if k in _VERSION_DIFF_SKIP:
                continue
            old_v, new_v = snap_a.get(k), snap_b.get(k)
            if isinstance(old_v, (list, dict)) or isinstance(new_v, (list, dict)):
                changed = old_v != new_v
                old_s = json.dumps(old_v, ensure_ascii=False) if old_v is not None else ""
                new_s = json.dumps(new_v, ensure_ascii=False) if new_v is not None else ""
            else:
                changed = old_v != new_v
                old_s = "" if old_v is None else str(old_v)
                new_s = "" if new_v is None else str(new_v)
            if changed:
                fields.append({"key": k, "old": old_s, "new": new_s})
        return ok({
            "entity_type": entity_type, "entity_id": entity_id,
            "a": {"id": va.id, "version_no": va.version_no,
                  "change_type": va.change_type, "change_note": va.change_note,
                  "created_at": va.created_at.strftime("%Y-%m-%d %H:%M:%S")
                  if va.created_at else ""},
            "b": {"id": vb.id, "version_no": vb.version_no,
                  "change_type": vb.change_type, "change_note": vb.change_note,
                  "created_at": vb.created_at.strftime("%Y-%m-%d %H:%M:%S")
                  if vb.created_at else ""},
            "changed_fields": fields,
            "changed_count": len(fields),
            "snapshot_a": snap_a, "snapshot_b": snap_b,
        })
    finally:
        s.close()


@router.post("/approvals", tags=["治理"])
def submit_approval(body: ApprovalIn):
    """提交发布审批：DRAFT 状态的指标提交后进入审批流（同实体重复提交 409）"""
    s = get_session()
    try:
        if body.entity_type not in APPROVAL_ENTITY_TYPES:
            raise HTTPException(400, f"仅指标支持审批发布: {', '.join(APPROVAL_ENTITY_TYPES)}")
        obj = _entity_by_type(s, body.entity_type, body.entity_id)
        if not obj:
            raise HTTPException(404, f"实体不存在: {body.entity_type} id={body.entity_id}")
        if getattr(obj, "status", "") == STATUS_PUBLISHED:
            raise HTTPException(400, f"{obj.name} 已是发布状态，无需重复提交")
        if getattr(obj, "status", "") == "ARCHIVED":
            raise HTTPException(400, f"{obj.name} 已归档停用，不能提交发布")
        pending = (s.query(Approval).filter_by(
            entity_type=body.entity_type, entity_id=body.entity_id,
            status="PENDING").first())
        if pending:
            raise HTTPException(409, f"存在待审批的发布申请（单号 #{pending.id}），请勿重复提交")
        a = Approval(entity_type=body.entity_type, entity_id=body.entity_id,
                     entity_code=obj.code, entity_name=obj.name,
                     action="publish", status="PENDING",
                     comment=body.comment)
        s.add(a)
        s.commit()
        _new_alert(s, "info", "approval", a.id,
                   f"待办：{obj.name}（{obj.code}）提交发布审批")
        s.commit()
        return ok({"id": a.id, "status": a.status})
    finally:
        s.close()


@router.get("/approvals", tags=["治理"])
def list_approvals(status: Optional[str] = None,
                   page: int = 1, page_size: int = 20):
    """审批单列表：status=PENDING 待办 / APPROVED+REJECTED 历史（默认全部）"""
    s = get_session()
    try:
        q = s.query(Approval).order_by(Approval.id.desc())
        if status:
            q = q.filter_by(status=status)
        total = q.count()
        rows = (q.offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        items = [{
            "id": a.id, "entity_type": a.entity_type, "entity_id": a.entity_id,
            "entity_code": a.entity_code, "entity_name": a.entity_name,
            "action": a.action, "status": a.status, "comment": a.comment,
            "created_at": (a.created_at.strftime("%Y-%m-%d %H:%M:%S")
                           if a.created_at else ""),
            "reviewed_at": (a.reviewed_at.strftime("%Y-%m-%d %H:%M:%S")
                            if a.reviewed_at else ""),
        } for a in rows]
        return ok({"items": items, "total": total, "page": page,
                   "page_size": page_size})
    finally:
        s.close()


@router.post("/approvals/{approval_id}/approve", tags=["治理"])
def approve_approval(approval_id: int, body: ApprovalReviewIn = None):
    """同意发布：实体置 PUBLISHED + 变更前存档版本 + 写告警"""
    s = get_session()
    try:
        a = _get_or_404(s, Approval, approval_id, "审批单")
        if a.status != "PENDING":
            raise HTTPException(409, f"审批单已处理（{a.status}），不可重复操作")
        obj = _entity_by_type(s, a.entity_type, a.entity_id)
        if not obj:
            raise HTTPException(404, f"实体不存在: {a.entity_type} id={a.entity_id}")
        _snapshot_version(s, a.entity_type, obj, "approve", "审批通过发布")
        obj.status = STATUS_PUBLISHED
        a.status = "APPROVED"
        a.reviewed_at = dt.datetime.now()
        a.comment = (body.comment if body else "") or a.comment
        s.commit()
        _new_alert(s, "info", "approval", a.id,
                   f"已发布：{a.entity_name}（{a.entity_code}）审批通过上线")
        s.commit()
        return ok({"id": a.id, "status": a.status})
    finally:
        s.close()


@router.post("/approvals/{approval_id}/reject", tags=["治理"])
def reject_approval(approval_id: int, body: ApprovalReviewIn = None):
    """驳回发布：实体状态不变，审批单置 REJECTED"""
    s = get_session()
    try:
        a = _get_or_404(s, Approval, approval_id, "审批单")
        if a.status != "PENDING":
            raise HTTPException(409, f"审批单已处理（{a.status}），不可重复操作")
        a.status = "REJECTED"
        a.reviewed_at = dt.datetime.now()
        a.comment = (body.comment if body else "") or a.comment
        s.commit()
        _new_alert(s, "warning", "approval", a.id,
                   f"被驳回：{a.entity_name}（{a.entity_code}）发布申请未通过")
        s.commit()
        return ok({"id": a.id, "status": a.status})
    finally:
        s.close()


@router.get("/impact-report", tags=["治理"])
def impact_report(object_type: str, object_id: int):
    """变更影响评估：统计对象变更波及的下游模型/派生/复合/数据集/应用 + 血缘链"""
    s = get_session()
    try:
        object_node, hits = _find_impacted_downstreams(s, object_type, object_id)
        # 派生指标引用（含间接：原子被派生用、维度被派生用）
        derived = []
        if object_type == "atomic_metric":
            derived = [{"id": d.id, "code": d.code, "name": d.name}
                       for d in s.query(DerivedMetric)
                       .filter_by(atomic_id=object_id).all()]
        elif object_type == "derived_metric":
            d = s.query(DerivedMetric).filter_by(code=object_node["code"]).first()
            if d:
                derived = [{"id": d.id, "code": d.code, "name": d.name}]
        elif object_type == "dimension":
            derived = [{"id": d.id, "code": d.code, "name": d.name}
                       for d in s.query(DerivedMetric).all()
                       if object_node["code"] in (d.dim_codes or [])]
        # 复合指标引用（引用上述派生指标，或直接引用原子指标）
        composite = []
        for c in s.query(CompositeMetric).all():
            refs = c.ref_codes or []
            hit = any(dm["code"] in refs for dm in derived)
            if object_type == "atomic_metric" and object_node["code"] in refs:
                hit = True
            if hit and c.code not in [x["code"] for x in composite]:
                composite.append({"id": c.id, "code": c.code, "name": c.name})
        # 数据集 + 授权应用
        ds_codes = set()
        for h in hits:
            ds_codes.update(_refs_in_datasets(s, downstream_id=h["dm"].id))
        for d in s.query(Dataset).all():
            if object_node["code"] in (d.metric_codes or []):
                ds_codes.add(d.code)
        datasets = []
        for code in sorted(ds_codes):
            d = s.query(Dataset).filter_by(code=code).first()
            if not d:
                continue
            apps = (s.query(DownstreamApp)
                    .join(AppDatasetGrant, AppDatasetGrant.app_id == DownstreamApp.id)
                    .filter(AppDatasetGrant.dataset_id == d.id).all())
            datasets.append({
                "id": d.id, "code": d.code, "name": d.name,
                "source_type": d.source_type,
                "granted_apps": [{"id": a.id, "code": a.code, "name": a.name}
                                 for a in apps],
            })
        apps = [a for d in datasets for a in d["granted_apps"]]
        return ok({
            "object": object_node,
            "summary": {
                "downstream_models": len(hits),
                "derived_metrics": len(derived),
                "composite_metrics": len(composite),
                "datasets": len(datasets),
                "granted_apps": len(apps),
            },
            "downstreams": [{
                "id": h["dm"].id, "code": h["dm"].code, "name": h["dm"].name,
                "source_model_code": h["dm"].source_model.code
                if h["dm"].source_model else None,
                "materialized": bool(h["dm"].materialized),
                "physical_table": h["dm"].physical_table,
                "row_count": h["dm"].row_count,
                "chain": h["chain"],
            } for h in hits],
            "derived": derived, "composite": composite, "datasets": datasets,
        })
    finally:
        s.close()


@router.post("/entity-tags", tags=["治理"])
def set_entity_tags(body: EntityTagsIn):
    """批量设置实体标签（全量替换）"""
    s = get_session()
    try:
        if body.entity_type not in ENTITY_MODELS:
            raise HTTPException(400, f"不支持的实体类型: {body.entity_type}")
        if not _entity_by_type(s, body.entity_type, body.entity_id):
            raise HTTPException(404, f"实体不存在: {body.entity_type} id={body.entity_id}")
        _set_tags(s, body.entity_type, body.entity_id, body.tags)
        s.commit()
        return ok({"entity_type": body.entity_type, "entity_id": body.entity_id,
                   "tags": body.tags})
    finally:
        s.close()


@router.delete("/entity-tags/{tag_id}", tags=["治理"])
def delete_entity_tag(tag_id: int):
    s = get_session()
    try:
        t = _get_or_404(s, EntityTag, tag_id, "标签")
        s.delete(t)
        s.commit()
        return ok({"deleted": tag_id})
    finally:
        s.close()


@router.get("/tags", tags=["治理"])
def list_tags(keyword: str = ""):
    """全部标签及使用计数（资产检索用）"""
    s = get_session()
    try:
        q = (s.query(EntityTag.tag, func.count(EntityTag.id).label("cnt"))
             .group_by(EntityTag.tag).order_by(func.count(EntityTag.id).desc()))
        if keyword:
            q = q.filter(EntityTag.tag.like(f"%{_like_escape(keyword)}%", escape="\\"))
        return ok([{"tag": tag, "count": cnt} for tag, cnt in q.all()])
    finally:
        s.close()


@router.get("/tags/entities", tags=["治理"])
def tags_entities(tag: str):
    """按标签过滤实体：返回各类型下打该标签的实体（id/code/name）"""
    s = get_session()
    try:
        rows = s.query(EntityTag).filter_by(tag=tag).all()
        out = {t: [] for t in ENTITY_MODELS}
        for r in rows:
            obj = _entity_by_type(s, r.entity_type, r.entity_id)
            if obj:
                out[r.entity_type].append({"id": obj.id, "code": obj.code,
                                           "name": obj.name})
        return ok(out)
    finally:
        s.close()


# ===========================================================================
# 数据质量与监控：质量规则校验、健康度总览、站内告警
# ===========================================================================

class QualityRuleIn(BaseModel):
    entity_id: int
    rule_type: str
    params: dict = {}
    severity: str = "warning"
    enabled: int = 1


QUALITY_RULE_TYPES = ("row_count_min", "row_count_change", "non_null_rate", "fresh_days")
QUALITY_SEVERITIES = ("info", "warning", "critical")


def _parse_bucket(s):
    """解析 date_bucket 为 date：day(YYYY-MM-DD) / month(YYYY-MM) / week(YYYY-Www)"""
    try:
        return dt.date.fromisoformat(str(s))
    except ValueError:
        pass
    m = re.match(r"^(\d{4})-(\d{2})$", str(s))
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), 1)
    m = re.match(r"^(\d{4})-W(\d{2})$", str(s))
    if m:
        return (dt.date(int(m.group(1)), 1, 1)
                + dt.timedelta(weeks=int(m.group(2)) - 1))
    return None


def _check_quality_rule(s, rule: QualityRule):
    """执行单条质量规则，更新 last_* 并返回 (result, value, message)；
    result ∈ ok/fail/error，非 ok 由调用方写告警"""
    m = s.query(DownstreamModel).get(rule.entity_id)
    if not m:
        rule.last_result, rule.last_message = "error", "下游模型不存在，无法校验"
        return rule.last_result, "", rule.last_message
    if not m.materialized or not m.physical_table:
        rule.last_result, rule.last_message = "error", "下游模型未物化，无法校验"
        return rule.last_result, "", rule.last_message
    tbl = _safe_ident(m.physical_table)
    params = rule.params or {}
    rule_type = rule.rule_type
    try:
        if rule_type == "row_count_min":
            n = s.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
            min_rows = int(params.get("min_rows", 0))
            val = str(n)
            if n >= min_rows:
                result, msg = "ok", f"行数 {n} 达标（下限 {min_rows}）"
            else:
                result, msg = "fail", f"行数 {n} 低于下限 {min_rows}"
        elif rule_type == "row_count_change":
            n = s.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
            prev, val = rule.last_value, str(n)
            if prev and prev.isdigit() and int(prev) > 0:
                pct = abs(n - int(prev)) / int(prev) * 100
                val = f"{pct:.1f}%"
                max_pct = float(params.get("max_change_pct", 20))
                if pct <= max_pct:
                    result, msg = "ok", f"行数波动 {pct:.1f}% 在阈值 {max_pct}% 内"
                else:
                    result, msg = "fail", f"行数波动 {pct:.1f}% 超过阈值 {max_pct}%"
            else:
                result, msg = "ok", f"首次检查，记录基准行数 {n}"
        elif rule_type == "non_null_rate":
            col = _safe_ident(str(params.get("column", "")))
            if not col:
                rule.last_result, rule.last_message = "error", "non_null_rate 规则需指定校验列 column"
                return rule.last_result, "", rule.last_message
            total = s.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
            nn = s.execute(
                text(f"SELECT COUNT(*) FROM {tbl} WHERE {col} IS NOT NULL")).scalar()
            rate = (nn / total * 100) if total else 100.0
            min_rate = float(params.get("min_rate", 95))
            val = f"{rate:.1f}%"
            if rate >= min_rate:
                result, msg = "ok", f"非空率 {rate:.1f}% 达标（阈值 {min_rate}%）"
            else:
                result, msg = "fail", f"非空率 {rate:.1f}% 低于阈值 {min_rate}%"
        elif rule_type == "fresh_days":
            latest = s.execute(text(f"SELECT MAX(date_bucket) FROM {tbl}")).scalar()
            if latest is None:
                result, val, msg = "fail", "无数据", "物化表无数据，新鲜度不满足"
            else:
                ld = _parse_bucket(latest)
                days = (dt.date.today() - ld).days if ld else 9999
                max_days = int(params.get("max_days", 7))
                val = f"{days}d"
                if days <= max_days:
                    result, msg = "ok", f"最新数据距今 {days} 天达标（阈值 {max_days} 天）"
                else:
                    result, msg = "fail", f"最新数据距今 {days} 天超过阈值 {max_days} 天"
        else:
            rule.last_result, rule.last_message = "error", f"未知规则类型: {rule_type}"
            return rule.last_result, "", rule.last_message
    except Exception as e:  # noqa: BLE001 - 校验执行异常按 error 处理
        rule.last_result, rule.last_message = "error", f"校验执行异常: {e}"
        return rule.last_result, "", rule.last_message
    rule.last_result = result
    # row_count_change 的 last_value 存原始行数（波动 % 仅作展示值），
    # 否则把百分比写回 last_value 会导致下次检查无法比较、每两次才校验一次
    rule.last_value = str(n) if rule_type == "row_count_change" else val
    rule.last_message = msg
    rule.last_check_at = dt.datetime.now()
    return result, val, msg


def _check_with_instance(s, rule: QualityRule, trigger: str = "manual"):
    """执行规则并记录任务实例（trigger=manual/auto），失败写告警"""
    m = s.query(DownstreamModel).get(rule.entity_id)
    inst = TaskInstance(task_type="quality_check", entity_type="downstream_model",
                        entity_id=rule.entity_id,
                        entity_code=m.code if m else "",
                        trigger=trigger, status="RUNNING")
    s.add(inst)
    result, val, msg = _check_quality_rule(s, rule)
    inst.status = "SUCCESS" if result == "ok" else "FAILED"
    inst.detail = {"rule_id": rule.id, "rule_type": rule.rule_type,
                   "result": result, "value": val}
    inst.error = "" if result == "ok" else msg
    inst.finished_at = dt.datetime.now()
    if result != "ok":
        # 告警级别跟随规则配置的 severity（error 状态恒为 error）
        level = "error" if result == "error" else (rule.severity or "warning")
        _new_alert(s, level, "quality", rule.id,
                   f"质量规则告警 [{rule.rule_type}] "
                   f"{m.code if m else rule.entity_id}: {msg}")
    return result, val, msg


def _run_quality_checks(s, entity_id: int):
    """物化/重导/调度成功后自动校验该下游模型全部启用规则（trigger=auto）。
    内部兜底不向上抛异常：单条规则校验失败由 _check_quality_rule 记 error，
    整体异常写 error 告警，避免质量检查失败误伤物化/重导任务本身"""
    try:
        for rule in s.query(QualityRule).filter_by(entity_id=entity_id,
                                                   enabled=1).all():
            _check_with_instance(s, rule, trigger="auto")
    except Exception as e:  # noqa: BLE001
        try:
            _new_alert(s, "error", "quality", entity_id,
                       f"自动质量检查异常: {e}")
        except Exception:  # noqa: BLE001
            pass


@router.post("/quality-rules", tags=["质量"])
def create_quality_rule(body: QualityRuleIn):
    s = get_session()
    try:
        if body.rule_type not in QUALITY_RULE_TYPES:
            raise HTTPException(400, f"非法规则类型: {body.rule_type}，可选 {QUALITY_RULE_TYPES}")
        if body.severity not in QUALITY_SEVERITIES:
            raise HTTPException(400, f"非法严重级别: {body.severity}")
        m = _get_or_404(s, DownstreamModel, body.entity_id, "下游模型")
        r = QualityRule(entity_type="downstream_model", entity_id=body.entity_id,
                        rule_type=body.rule_type, params=body.params,
                        severity=body.severity, enabled=1 if body.enabled else 0)
        s.add(r)
        s.commit()
        return ok({"id": r.id, "entity_code": m.code})
    finally:
        s.close()


@router.get("/quality-rules", tags=["质量"])
def list_quality_rules(entity_id: Optional[int] = None):
    s = get_session()
    try:
        q = s.query(QualityRule).order_by(QualityRule.id.desc())
        if entity_id:
            q = q.filter_by(entity_id=entity_id)
        items = []
        for r in q.all():
            m = s.query(DownstreamModel).get(r.entity_id)
            items.append({
                "id": r.id, "entity_type": r.entity_type, "entity_id": r.entity_id,
                "entity_code": m.code if m else "",
                "entity_name": m.name if m else "",
                "rule_type": r.rule_type, "params": r.params or {},
                "severity": r.severity, "enabled": bool(r.enabled),
                "last_check_at": (r.last_check_at.strftime("%Y-%m-%d %H:%M:%S")
                                  if r.last_check_at else ""),
                "last_result": r.last_result, "last_value": r.last_value,
                "last_message": r.last_message,
            })
        return ok({"items": items})
    finally:
        s.close()


@router.put("/quality-rules/{rule_id}", tags=["质量"])
def update_quality_rule(rule_id: int, body: QualityRuleIn):
    s = get_session()
    try:
        r = _get_or_404(s, QualityRule, rule_id, "质量规则")
        if body.rule_type not in QUALITY_RULE_TYPES:
            raise HTTPException(400, f"非法规则类型: {body.rule_type}")
        if body.severity not in QUALITY_SEVERITIES:
            raise HTTPException(400, f"非法严重级别: {body.severity}")
        _get_or_404(s, DownstreamModel, body.entity_id, "下游模型")
        r.entity_id = body.entity_id
        r.rule_type = body.rule_type
        r.params = body.params
        r.severity = body.severity
        r.enabled = 1 if body.enabled else 0
        s.commit()
        return ok({"id": r.id})
    finally:
        s.close()


@router.delete("/quality-rules/{rule_id}", tags=["质量"])
def delete_quality_rule(rule_id: int):
    s = get_session()
    try:
        r = _get_or_404(s, QualityRule, rule_id, "质量规则")
        s.delete(r)
        s.commit()
        return ok({"deleted": rule_id})
    finally:
        s.close()


@router.post("/quality-rules/{rule_id}/toggle", tags=["质量"])
def toggle_quality_rule(rule_id: int):
    s = get_session()
    try:
        r = _get_or_404(s, QualityRule, rule_id, "质量规则")
        r.enabled = 0 if r.enabled else 1
        s.commit()
        return ok({"id": r.id, "enabled": bool(r.enabled)})
    finally:
        s.close()


@router.post("/quality-rules/{rule_id}/check", tags=["质量"])
def check_quality_rule(rule_id: int):
    """手动执行质量规则校验（记录任务实例；fail/error 自动写告警）"""
    s = get_session()
    try:
        r = _get_or_404(s, QualityRule, rule_id, "质量规则")
        result, val, msg = _check_with_instance(s, r, trigger="manual")
        s.commit()
        return ok({"rule_id": r.id, "result": result, "value": val,
                   "message": msg})
    finally:
        s.close()


@router.get("/quality/health", tags=["质量"])
def quality_health():
    """健康度总览：每个下游模型的物化/行数/最新数据/新鲜度/规则结果 → 三色等级
    green 健康 / yellow 关注（未物化、有 fail、新鲜度>3 天）/ red 告警（error、新鲜度>7 天）"""
    s = get_session()
    try:
        rows = []
        for m in s.query(DownstreamModel).order_by(DownstreamModel.id).all():
            latest, fresh_days = None, None
            if m.materialized and m.physical_table:
                try:
                    latest = s.execute(text(
                        f"SELECT MAX(date_bucket) FROM {_safe_ident(m.physical_table)}"
                    )).scalar()
                    ld = _parse_bucket(latest) if latest is not None else None
                    fresh_days = (dt.date.today() - ld).days if ld else None
                except Exception:  # noqa: BLE001 - 物化表异常不阻断总览
                    pass
            # 健康度只统计启用规则，禁用规则的旧失败结果不拖累等级
            rules = s.query(QualityRule).filter_by(entity_id=m.id,
                                                   enabled=1).all()
            failed = [r for r in rules if r.last_result == "fail"]
            error = [r for r in rules if r.last_result == "error"]
            if error or (fresh_days is not None and fresh_days > 7):
                level = "red"
            elif failed or not m.materialized or (fresh_days is not None and fresh_days > 3):
                level = "yellow"
            else:
                level = "green"
            rows.append({
                "id": m.id, "code": m.code, "name": m.name,
                "materialized": bool(m.materialized),
                "physical_table": m.physical_table, "row_count": m.row_count,
                "latest_date": str(latest) if latest is not None else None,
                "fresh_days": fresh_days,
                "rule_total": len(rules), "rule_failed": len(failed),
                "rule_error": len(error), "level": level,
            })
        return ok({
            "summary": {"total": len(rows),
                        "green": sum(1 for r in rows if r["level"] == "green"),
                        "yellow": sum(1 for r in rows if r["level"] == "yellow"),
                        "red": sum(1 for r in rows if r["level"] == "red")},
            "items": rows,
        })
    finally:
        s.close()


# ---- 站内告警（通知铃铛） ---------------------------------------------------

@router.get("/alerts", tags=["质量"])
def list_alerts(unread_only: Optional[bool] = None,
                page: int = 1, page_size: int = 20):
    s = get_session()
    try:
        q = s.query(Alert).order_by(Alert.id.desc())
        if unread_only:
            q = q.filter_by(read=0)
        total = q.count()
        rows = (q.offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        items = [{
            "id": a.id, "level": a.level, "source_type": a.source_type,
            "source_id": a.source_id, "message": a.message,
            "read": bool(a.read),
            "created_at": (a.created_at.strftime("%Y-%m-%d %H:%M:%S")
                           if a.created_at else ""),
        } for a in rows]
        return ok({"items": items, "total": total, "page": page,
                   "page_size": page_size})
    finally:
        s.close()


@router.get("/alerts/unread-count", tags=["质量"])
def alerts_unread_count():
    s = get_session()
    try:
        n = s.query(Alert).filter_by(read=0).count()
        return ok({"unread": n})
    finally:
        s.close()


@router.post("/alerts/{alert_id}/read", tags=["质量"])
def read_alert(alert_id: int):
    s = get_session()
    try:
        a = _get_or_404(s, Alert, alert_id, "告警")
        a.read = 1
        s.commit()
        return ok({"id": a.id, "read": True})
    finally:
        s.close()


@router.post("/alerts/read-all", tags=["质量"])
def read_all_alerts():
    s = get_session()
    try:
        n = s.query(Alert).filter_by(read=0).update({"read": 1})
        s.commit()
        return ok({"updated": n})
    finally:
        s.close()


# ===========================================================================
# 调度与运维：周期调度（零依赖线程调度器）、任务实例、失败重试
# ===========================================================================

class ScheduleIn(BaseModel):
    entity_id: int
    schedule_type: str = "daily"          # daily / interval
    hour: int = 2                         # daily：每天几点（0-23）
    minute: int = 0                       # daily：几分（0-59）
    interval_minutes: int = 60            # interval：每 N 分钟
    action: str = "materialize"           # materialize / reimport
    enabled: int = 1


SCHEDULE_TYPES = ("daily", "interval")
SCHEDULE_ACTIONS = ("materialize", "reimport")


def _check_schedule_body(body: ScheduleIn):
    """调度参数边界校验：小时/分钟/间隔分钟必须在合法范围"""
    if body.schedule_type not in SCHEDULE_TYPES:
        raise HTTPException(400, f"非法调度类型: {body.schedule_type}，可选 {SCHEDULE_TYPES}")
    if body.action not in SCHEDULE_ACTIONS:
        raise HTTPException(400, f"非法调度动作: {body.action}，可选 {SCHEDULE_ACTIONS}")
    if not (0 <= body.hour <= 23):
        raise HTTPException(400, f"非法小时: {body.hour}，必须 0-23")
    if not (0 <= body.minute <= 59):
        raise HTTPException(400, f"非法分钟: {body.minute}，必须 0-59")
    if body.interval_minutes < 1:
        raise HTTPException(400, f"非法间隔分钟: {body.interval_minutes}，必须 >= 1")


def _calc_next_run(sch: Schedule) -> Optional[dt.datetime]:
    """下次运行时间：daily=下一到达时刻 / interval=当前+间隔分钟"""
    now = dt.datetime.now()
    if sch.schedule_type == "daily":
        nxt = now.replace(hour=sch.hour or 0, minute=sch.minute or 0,
                          second=0, microsecond=0)
        if nxt <= now:
            nxt += dt.timedelta(days=1)
        return nxt
    if sch.schedule_type == "interval":
        return now + dt.timedelta(minutes=sch.interval_minutes or 60)
    return None


def _execute_schedule_run(s, sch: Schedule, trigger: str = "schedule") -> str:
    """执行一次调度任务（物化/重导），写任务实例；失败回滚并写 FAILED 实例 + 告警。
    返回 "SUCCESS"/"FAILED"；失败同样推进 next_run_at，避免每 30s 无限重试刷告警；
    成功后由调用方触发自动质量检查"""
    sch_id = sch.id
    m = s.query(DownstreamModel).get(sch.entity_id)
    if not m:
        # 模型已被删除：记录失败并推进下次执行时间，不再空转重试
        s.rollback()
        sch = s.query(Schedule).get(sch_id)
        inst = TaskInstance(task_type=sch.action or "materialize",
                            entity_type="downstream_model",
                            entity_id=sch.entity_id, entity_code="",
                            trigger=trigger, status="FAILED",
                            error="下游模型不存在，调度已跳过",
                            finished_at=dt.datetime.now())
        s.add(inst)
        sch.next_run_at = _calc_next_run(sch)
        return "FAILED"
    m_id, m_code = m.id, m.code
    inst = TaskInstance(task_type=sch.action, entity_type="downstream_model",
                        entity_id=m_id, entity_code=m_code,
                        trigger=trigger, status="RUNNING")
    s.add(inst)
    try:
        if sch.action == "materialize":
            sql, params = gen.generate_downstream_sql(m)
            tbl = f"dl_{_safe_ident(m.code)}"
            m.definition_sql = sql
            with engine.begin() as conn:
                conn.execute(text(f"DROP TABLE IF EXISTS {tbl}"))
                conn.execute(text(f"CREATE TABLE {tbl} AS {sql}"), params)
            with engine.connect() as conn:
                n = conn.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
            m.materialized, m.physical_table, m.row_count = 1, tbl, n
            inst.detail = {"physical_table": tbl, "row_count": n}
        elif sch.action == "reimport":
            if not m.materialized or not m.physical_table:
                raise ValueError("请先物化，再执行数据重导")
            start, end = _reimport_range()
            fmt = GRANULARITY_FMT.get(m.granularity, "%Y-%m-%d")
            sb, eb = start.strftime(fmt), end.strftime(fmt)
            deleted, inserted, total = _do_reimport(s, m, sb, eb)
            inst.detail = {"start_date": start.isoformat(),
                           "end_date": end.isoformat(),
                           "deleted": deleted, "inserted": inserted,
                           "total_rows": total}
        else:
            raise ValueError(f"未知调度动作: {sch.action}")
        inst.status = "SUCCESS"
        inst.finished_at = dt.datetime.now()
        sch.last_run_at = dt.datetime.now()
        sch.next_run_at = _calc_next_run(sch)
        return "SUCCESS"
    except Exception as e:  # noqa: BLE001 - 调度失败写 FAILED 实例 + 告警
        s.rollback()
        inst = TaskInstance(task_type=sch.action, entity_type="downstream_model",
                            entity_id=m_id, entity_code=m_code,
                            trigger=trigger, status="FAILED",
                            error=str(e), finished_at=dt.datetime.now())
        s.add(inst)
        s.flush()  # 先落 id，供告警 source_id 引用
        _new_alert(s, "error", "task", inst.id,
                   f"调度任务失败 [{sch.action}] {m_code}: {e}")
        # 失败也推进下次执行时间，避免每 30s 无限重试 + 告警刷屏
        sch = s.query(Schedule).get(sch_id)
        sch.next_run_at = _calc_next_run(sch)
        return "FAILED"


def _scheduler_tick():
    """扫描一轮：到点的启用调度执行；成功后自动质量检查"""
    s = get_session()
    try:
        now = dt.datetime.now()
        for sch in s.query(Schedule).filter_by(enabled=1).all():
            if sch.next_run_at and sch.next_run_at > now:
                continue
            status = _execute_schedule_run(s, sch)
            if status == "SUCCESS":
                _run_quality_checks(s, sch.entity_id)
            s.commit()
    finally:
        s.close()


def _scheduler_loop(stop_event: threading.Event):
    """调度器线程：每 30s 扫描一轮，stop_event 置位后退出"""
    while not stop_event.is_set():
        try:
            _scheduler_tick()
        except Exception as e:  # noqa: BLE001 - 单轮异常不终止线程
            print(f"[scheduler] tick error: {e}", flush=True)
        stop_event.wait(30)


@router.post("/schedules", tags=["运维"])
def create_schedule(body: ScheduleIn):
    s = get_session()
    try:
        m = _get_or_404(s, DownstreamModel, body.entity_id, "下游模型")
        _check_schedule_body(body)
        sch = Schedule(entity_id=body.entity_id, schedule_type=body.schedule_type,
                       hour=body.hour, minute=body.minute,
                       interval_minutes=body.interval_minutes, action=body.action,
                       enabled=1 if body.enabled else 0)
        sch.next_run_at = _calc_next_run(sch)
        s.add(sch)
        s.commit()
        return ok({"id": sch.id, "entity_code": m.code,
                   "next_run_at": (sch.next_run_at.strftime("%Y-%m-%d %H:%M:%S")
                                   if sch.next_run_at else "")})
    finally:
        s.close()


@router.get("/schedules", tags=["运维"])
def list_schedules():
    s = get_session()
    try:
        items = []
        for sch in s.query(Schedule).order_by(Schedule.id).all():
            m = s.query(DownstreamModel).get(sch.entity_id)
            items.append({
                "id": sch.id, "entity_id": sch.entity_id,
                "entity_code": m.code if m else "",
                "entity_name": m.name if m else "",
                "schedule_type": sch.schedule_type, "hour": sch.hour,
                "minute": sch.minute, "interval_minutes": sch.interval_minutes,
                "action": sch.action, "enabled": bool(sch.enabled),
                "last_run_at": (sch.last_run_at.strftime("%Y-%m-%d %H:%M:%S")
                                if sch.last_run_at else ""),
                "next_run_at": (sch.next_run_at.strftime("%Y-%m-%d %H:%M:%S")
                                if sch.next_run_at else ""),
            })
        return ok({"items": items})
    finally:
        s.close()


@router.put("/schedules/{schedule_id}", tags=["运维"])
def update_schedule(schedule_id: int, body: ScheduleIn):
    s = get_session()
    try:
        sch = _get_or_404(s, Schedule, schedule_id, "调度")
        _get_or_404(s, DownstreamModel, body.entity_id, "下游模型")
        _check_schedule_body(body)
        sch.entity_id = body.entity_id
        sch.schedule_type = body.schedule_type
        sch.hour = body.hour
        sch.minute = body.minute
        sch.interval_minutes = body.interval_minutes
        sch.action = body.action
        sch.enabled = 1 if body.enabled else 0
        sch.next_run_at = _calc_next_run(sch)
        s.commit()
        return ok({"id": sch.id})
    finally:
        s.close()


@router.delete("/schedules/{schedule_id}", tags=["运维"])
def delete_schedule(schedule_id: int):
    s = get_session()
    try:
        sch = _get_or_404(s, Schedule, schedule_id, "调度")
        s.delete(sch)
        s.commit()
        return ok({"deleted": schedule_id})
    finally:
        s.close()


@router.post("/schedules/{schedule_id}/toggle", tags=["运维"])
def toggle_schedule(schedule_id: int):
    s = get_session()
    try:
        sch = _get_or_404(s, Schedule, schedule_id, "调度")
        sch.enabled = 0 if sch.enabled else 1
        if sch.enabled and not sch.next_run_at:
            sch.next_run_at = _calc_next_run(sch)
        s.commit()
        return ok({"id": sch.id, "enabled": bool(sch.enabled)})
    finally:
        s.close()


@router.post("/schedules/{schedule_id}/run", tags=["运维"])
def run_schedule(schedule_id: int):
    """立即执行一次调度任务（手动触发，供前端/测试使用，不依赖线程时序）；
    成功自动跑质量检查"""
    s = get_session()
    try:
        sch = _get_or_404(s, Schedule, schedule_id, "调度")
        status = _execute_schedule_run(s, sch, trigger="manual")
        if status == "SUCCESS":
            _run_quality_checks(s, sch.entity_id)
        s.commit()
        inst = (s.query(TaskInstance)
                .filter_by(task_type=sch.action, entity_id=sch.entity_id,
                           trigger="manual")
                .order_by(TaskInstance.id.desc()).first())
        return ok({"schedule_id": sch.id, "status": status,
                   "task_instance_id": inst.id if inst else None,
                   "entity_code": inst.entity_code if inst else ""})
    finally:
        s.close()


@router.get("/task-instances", tags=["运维"])
def list_task_instances(task_type: Optional[str] = None,
                        status: Optional[str] = None,
                        page: int = 1, page_size: int = 20):
    s = get_session()
    try:
        q = s.query(TaskInstance).order_by(TaskInstance.id.desc())
        if task_type:
            q = q.filter_by(task_type=task_type)
        if status:
            q = q.filter_by(status=status)
        total = q.count()
        rows = (q.offset((_page_clamped(page) - 1) * _page_size_clamped(page_size))
                .limit(_page_size_clamped(page_size)).all())
        items = [{
            "id": i.id, "task_type": i.task_type, "entity_type": i.entity_type,
            "entity_id": i.entity_id, "entity_code": i.entity_code,
            "status": i.status, "trigger": i.trigger, "detail": i.detail or {},
            "error": i.error,
            "started_at": (i.started_at.strftime("%Y-%m-%d %H:%M:%S")
                           if i.started_at else ""),
            "finished_at": (i.finished_at.strftime("%Y-%m-%d %H:%M:%S")
                            if i.finished_at else ""),
        } for i in rows]
        return ok({"items": items, "total": total, "page": page,
                   "page_size": page_size})
    finally:
        s.close()


@router.get("/task-instances/{instance_id}", tags=["运维"])
def get_task_instance(instance_id: int):
    s = get_session()
    try:
        i = _get_or_404(s, TaskInstance, instance_id, "任务实例")
        return ok({
            "id": i.id, "task_type": i.task_type, "entity_type": i.entity_type,
            "entity_id": i.entity_id, "entity_code": i.entity_code,
            "status": i.status, "trigger": i.trigger, "detail": i.detail or {},
            "error": i.error,
            "started_at": (i.started_at.strftime("%Y-%m-%d %H:%M:%S")
                           if i.started_at else ""),
            "finished_at": (i.finished_at.strftime("%Y-%m-%d %H:%M:%S")
                            if i.finished_at else ""),
        })
    finally:
        s.close()


@router.post("/task-instances/{instance_id}/retry", tags=["运维"])
def retry_task_instance(instance_id: int):
    """失败任务重试：按实例类型重放（物化重建 / 重导按原日期范围 / 质量检查重跑规则）"""
    s = get_session()
    try:
        inst = _get_or_404(s, TaskInstance, instance_id, "任务实例")
        if inst.status == "RUNNING":
            raise HTTPException(409, "实例执行中，暂不可重试")
        m = _get_or_404(s, DownstreamModel, inst.entity_id, "下游模型")
        m_id, m_code = m.id, m.code
        if inst.task_type == "quality_check":
            # 重试原失败的那条规则（detail.rule_id）；兼容旧实例回退到第一条
            d = inst.detail or {}
            rule = (s.query(QualityRule).get(d.get("rule_id"))
                    if d.get("rule_id") else None)
            if not rule:
                rule = s.query(QualityRule).filter_by(entity_id=m_id).first()
            if not rule:
                raise HTTPException(400, "该实体没有质量规则，无需重试")
            result, val, msg = _check_with_instance(s, rule, trigger="manual")
            s.commit()
            return ok({"status": "ok", "result": result, "value": val,
                       "message": msg})
        new_inst = TaskInstance(task_type=inst.task_type,
                                entity_type="downstream_model",
                                entity_id=m_id, entity_code=m_code,
                                trigger="manual", status="RUNNING")
        s.add(new_inst)
        try:
            if inst.task_type == "materialize":
                sql, params = gen.generate_downstream_sql(m)
                tbl = f"dl_{_safe_ident(m.code)}"
                m.definition_sql = sql
                with engine.begin() as conn:
                    conn.execute(text(f"DROP TABLE IF EXISTS {tbl}"))
                    conn.execute(text(f"CREATE TABLE {tbl} AS {sql}"), params)
                with engine.connect() as conn:
                    n = conn.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
                m.materialized, m.physical_table, m.row_count = 1, tbl, n
                new_inst.detail = {"physical_table": tbl, "row_count": n}
            elif inst.task_type == "reimport":
                if not m.materialized or not m.physical_table:
                    raise ValueError("请先物化，再执行数据重导")
                d = inst.detail or {}
                start, end = _reimport_range(d.get("start_date"),
                                             d.get("end_date"))
                fmt = GRANULARITY_FMT.get(m.granularity, "%Y-%m-%d")
                sb, eb = start.strftime(fmt), end.strftime(fmt)
                deleted, inserted, total = _do_reimport(s, m, sb, eb)
                new_inst.detail = {"start_date": start.isoformat(),
                                   "end_date": end.isoformat(),
                                   "deleted": deleted, "inserted": inserted,
                                   "total_rows": total}
            else:
                raise ValueError(f"未知任务类型: {inst.task_type}")
            new_inst.status = "SUCCESS"
            new_inst.finished_at = dt.datetime.now()
        except Exception as e:  # noqa: BLE001 - 重试失败写 FAILED 实例 + 告警
            s.rollback()
            new_inst = TaskInstance(task_type=inst.task_type,
                                    entity_type="downstream_model",
                                    entity_id=m_id, entity_code=m_code,
                                    trigger="manual", status="FAILED",
                                    error=str(e), finished_at=dt.datetime.now())
            s.add(new_inst)
            s.flush()  # 先落 id，供告警 source_id 引用
            _new_alert(s, "error", "task", new_inst.id,
                       f"任务重试失败 [{inst.task_type}] {m_code}: {e}")
        s.commit()
        return ok({"id": new_inst.id, "status": new_inst.status,
                   "detail": new_inst.detail, "error": new_inst.error})
    finally:
        s.close()


# ===========================================================================
# 元数据浏览（Dashboard 用）+ 静态资源
# ===========================================================================

@router.get("/overview", tags=["概览"])
def overview():
    s = get_session()
    try:
        return ok({
            "domains": s.query(SubjectDomain).count(),
            "processes": s.query(BusinessProcess).count(),
            "dimensions": s.query(Dimension).count(),
            "atomic": s.query(AtomicMetric).count(),
            "derived": s.query(DerivedMetric).count(),
            "composite": s.query(CompositeMetric).count(),
            "logical_models": s.query(LogicalModel).count(),
            "downstream_models": s.query(DownstreamModel).count(),
            "downstream_apps": s.query(DownstreamApp).count(),
            "datasets": s.query(Dataset).count(),
        })
    finally:
        s.close()


@router.get("/metrics", tags=["概览"])
def list_metrics():
    """全部指标元数据（按类型分组），供前端筛选展示"""
    s = get_session()
    try:
        atomic = [{"type": "atomic", "id": m.id, "code": m.code, "name": m.name,
                   "process": m.process.name, "agg": m.agg_function,
                   "field": m.physical_field, "table": m.process.physical_table,
                   "unit": m.unit, "status": m.status, "description": m.description}
                  for m in s.query(AtomicMetric).all()]
        derived = [{"type": "derived", "id": m.id, "code": m.code, "name": m.name,
                    "atomic": m.atomic.code, "time_period": m.time_period,
                    "dim_codes": m.dim_codes or [], "filters": m.filters or [],
                    "modifier_codes": m.modifier_codes or [],
                    "compare_type": m.compare_type or "none",
                    "owner": m.owner, "cert_level": m.cert_level,
                    "unit": m.atomic.unit, "status": m.status,
                    "description": m.description}
                   for m in s.query(DerivedMetric).all()]
        composite = [{"type": "composite", "id": m.id, "code": m.code, "name": m.name,
                      "expression": m.expression, "ref_codes": m.ref_codes or [],
                      "unit": m.unit, "status": m.status,
                      "owner": m.owner, "cert_level": m.cert_level,
                      "description": m.description}
                     for m in s.query(CompositeMetric).all()]
        return ok({"atomic": atomic, "derived": derived, "composite": composite})
    finally:
        s.close()


# ===========================================================================
# App 组装：统一响应包装 + 双前缀挂载（/api 与 /api/v1）
# ===========================================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    """启动内置调度线程（每 30s 扫描启用的调度执行），退出时优雅停止。
    TestClient 非上下文模式不触发，不影响测试"""
    stop_event = threading.Event()
    thread = threading.Thread(target=_scheduler_loop, args=(stop_event,),
                              name="schedule-worker", daemon=True)
    thread.start()
    print("[scheduler] started", flush=True)
    try:
        yield
    finally:
        stop_event.set()
        thread.join(timeout=5)
        print("[scheduler] stopped", flush=True)


app = FastAPI(title="统一指标维度管理平台 Demo", version="1.0.0",
              lifespan=lifespan)
# 仅放行本地前端来源，避免任意站点跨域读取平台数据
app.add_middleware(CORSMiddleware,
                   allow_origins=["http://127.0.0.1:8000",
                                   "http://localhost:8000",
                                   "http://127.0.0.1:5500",
                                   "http://localhost:5500"],
                   allow_methods=["GET", "POST", "PUT", "DELETE"],
                   allow_headers=["Content-Type"])


@app.exception_handler(HTTPException)
async def http_exc_handler(request, exc: HTTPException):
    return JSONResponse(
        status_code=exc.status_code or 400,
        content={"code": exc.status_code or 400, "message": str(exc.detail),
                 "data": None})


@app.exception_handler(Exception)
async def unhandled_exc_handler(request, exc: Exception):
    # 不向客户端暴露内部异常细节（路径/堆栈/表名等）
    return JSONResponse(
        status_code=500,
        content={"code": 500, "message": "服务器内部错误，请查看服务端日志",
                 "data": None})


app.include_router(router, prefix="/api")
app.include_router(router, prefix="/api/v1")
app.include_router(openapi_router, prefix="/openapi")


@app.get("/")
def index():
    return FileResponse(FRONTEND_DIR / "index.html")


app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)