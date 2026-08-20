"""核心逻辑单元测试（需求文档第 9 章：SQL 生成、指标派生需有单元测试）

覆盖：
  1. 派生规则引擎：时间周期解析、维度 JOIN、业务限定（筛选条件）-> SQL 生成
  2. 复合指标：表达式替换、防除零、子查询 JOIN
  3. 统一指标查询：多指标/多维度联合查询、日/周/月日期粒度、口径一致
  4. 下游模型：定义 SQL 生成、物化（幂等刷新）、表血缘
  5. 安全：标识符白名单防注入
  6. API：统一响应结构、引用校验（409）、编码冲突（409）、参数校验（400）
  7. 血缘：指标全链路 + 表级血缘（物理表 -> 逻辑模型 -> 下游模型 -> 物化表）
  8. 下游应用 / 数据集 / 开放 API：AppKey+AppSecret 认证、双源调用、调用日志、注入回归
"""
import datetime as dt
import re

import pytest

from sql_generator import SQLGenerator, MetricNotFoundError

gen = SQLGenerator()


# ===========================================================================
# 派生规则引擎：原子指标 + 修饰词 -> SQL
# ===========================================================================

def test_derived_sql_time_period(api):
    """最近7天：SQL 应含维度 JOIN、时间条件（动态参数）"""
    mtype, sql, params = gen.generate_sql_only("pay_amount_7d_city")
    assert mtype == "derived"
    assert "SUM(t.pay_amount)" in sql
    assert "FROM dwd_pay_detail t" in sql
    assert "LEFT JOIN dim_city" in sql
    assert "pay_date >= :_start" in sql and "pay_date <= :_end" in sql


def test_period_resolution():
    today = dt.date.today()
    # 7d
    start, end = gen.resolve_period("7d")
    assert end == today and (today - start).days == 7
    # ytd
    start, end = gen.resolve_period("ytd")
    assert start == dt.date(today.year, 1, 1) and end == today
    # custom 必须传 start_date
    with pytest.raises(ValueError):
        gen.resolve_period("custom")
    # 非法周期
    with pytest.raises(ValueError):
        gen.resolve_period("999d")


def test_derived_filter_business_limit(api):
    """业务限定：order_status IN ('PAID','SHIPPED') 应生成 IN 条件（值走绑定参数，防注入）"""
    res = api.get("/api/v1/derived-metrics/3/sql-preview")
    d = res.json()["data"]
    assert "order_status IN" in d["sql"]
    assert "PAID" in str(d["params"].values()) and "SHIPPED" in str(d["params"].values())

    mtype, sql, params = gen.generate_sql_only(
        "order_amount_30d_cat",
        None,
        (dt.date.today() - dt.timedelta(days=30)).isoformat(),
        dt.date.today().isoformat())
    assert "order_status IN" in sql
    assert "PAID" in params.values() and "SHIPPED" in params.values()


def test_composite_sql_expression(api):
    """客单价 = 支付金额/支付笔数：应生成子查询 JOIN + NULLIF 防除零"""
    d = api.get("/api/v1/composite-metrics/1").json()["data"]
    sql = d["generated_sql"]
    assert "NULLIF" in sql
    assert "CAST(" in sql
    # 两个引用派生指标均生成子查询
    for ref in ("pay_amount_7d_city", "pay_count_7d_city"):
        assert ref in d["ref_codes"]


def test_composite_execute_grouped(api):
    r = api.post("/api/v1/query", json={"metric_code": "avg_order_value",
                                        "dim_codes": ["dim_city"]})
    assert r.status_code == 200
    d = r.json()["data"]
    assert d["summary"]["metric_types"] == ["composite"]
    cols = d["columns"]
    # 新契约：首列日期桶，指标值列在维度列之后
    assert cols[0] == "date_bucket" and "dim_city" in cols
    vi = cols.index("avg_order_value")
    assert len(d["rows"]) > 0
    # 客单价为正且有限（行结构 [date_bucket, 维度..., 指标...]，值列按 columns 定位）
    for row in d["rows"]:
        assert row[vi] is not None and row[vi] > 0


def test_refund_rate_sanity(api):
    """退款率 = 退款金额/支付金额 ∈ [0, 1]，口径一致"""
    r = api.post("/api/v1/query", json={"metric_code": "refund_rate",
                                        "dim_codes": ["dim_city"]})
    d = r.json()["data"]
    vi = d["columns"].index("refund_rate")
    for row in d["rows"]:
        assert row[vi] is not None and 0 <= row[vi] <= 1, f"退款率异常: {row}"


# ===========================================================================
# 统一指标查询
# ===========================================================================

def test_query_atomic_with_dims(api):
    """原子指标 + 维度 + 时间范围 -> 返回一致口径结果（含日期桶首列）"""
    r = api.post("/api/v1/query", json={
        "metric_code": "pay_amount_sum", "dim_codes": ["dim_city"],
        "start_date": "2026-08-01", "end_date": "2026-08-19"})
    d = r.json()["data"]
    assert r.status_code == 200
    assert d["summary"]["row_count"] > 0
    assert d["columns"][0] == "date_bucket"
    assert "dim_city" in d["columns"] and "pay_amount_sum" in d["columns"]
    # 值列位于日期桶与维度列之后
    assert d["columns"].index("pay_amount_sum") == 2
    assert d["summary"]["granularity"] == "day"
    assert d["summary"]["metric_names"] == ["支付金额"]
    for row in d["rows"]:
        assert re.match(r"^\d{4}-\d{2}-\d{2}$", row[0])  # 日桶格式 YYYY-MM-DD


def test_query_metric_not_found(api):
    r = api.post("/api/v1/query", json={"metric_code": "not_exist"})
    assert r.status_code == 404
    assert r.json()["code"] == 404


def test_query_invalid_dim(api):
    """非法维度编码 -> 400，不执行 SQL"""
    r = api.post("/api/v1/query", json={"metric_code": "pay_amount_sum",
                                        "dim_codes": ["evil_dim"]})
    assert r.status_code == 400


def test_sql_preview_no_execution(api):
    r = api.get("/api/v1/sql-preview",
                params={"metric_codes": "pay_amount_7d_city,order_amount_sum",
                        "granularity": "week"})
    d = r.json()["data"]
    assert "SELECT" in d["sql"] and d["params"]
    assert d["metric_codes"] == ["pay_amount_7d_city", "order_amount_sum"]
    assert d["granularity"] == "week"
    # 周粒度桶格式写入 SQL（strftime 白名单映射）
    assert "strftime('%Y-W%W'" in d["sql"]


# ===========================================================================
# SQL 注入防护
# ===========================================================================

def test_injection_metric_code(api):
    """注入尝试：不合法编码被拒绝（400/404），且数据库不受影响"""
    r = api.post("/api/v1/query", json={"metric_code": "x'; DROP TABLE t;--"})
    assert r.status_code in (400, 404)
    # 验证注入未生效：平台元数据完好，查询仍可用
    r = api.get("/api/v1/metrics")
    assert r.status_code == 200 and len(r.json()["data"]["atomic"]) == 5
    r = api.post("/api/v1/query", json={"metric_code": "pay_amount_7d_city",
                                        "dim_codes": ["dim_city"]})
    assert r.status_code == 200


def test_injection_dim_code(api):
    r = api.post("/api/v1/query", json={"metric_code": "pay_amount_sum",
                                        "dim_codes": ["dim_city; DROP"]})
    assert r.status_code == 400


# ===========================================================================
# 接口契约：统一响应 / 引用校验 / 编码冲突 / 参数校验
# ===========================================================================

def test_unified_response_shape(api):
    r = api.get("/api/v1/domains")
    body = r.json()
    assert set(body.keys()) == {"code", "message", "data"}
    assert body["code"] == 0 and body["data"]["total"] == 1


def test_duplicate_code_409(api):
    r = api.post("/api/v1/domains", json={"code": "trade", "name": "重复"})
    assert r.status_code == 409
    assert "编码已存在" in r.json()["message"]


def test_delete_referenced_atomic_blocked(api):
    """pay_amount_sum 被派生指标引用 -> 409"""
    r = api.delete("/api/v1/atomic-metrics/3")
    assert r.status_code == 409


def test_delete_referenced_dim_blocked(api):
    r = api.delete("/api/v1/dimensions/1")
    assert r.status_code == 409
    assert "禁止删除" in r.json()["message"]


def test_create_derived_validation(api):
    # 非法时间周期
    r = api.post("/api/v1/derived-metrics", json={
        "code": "bad_period", "name": "错误周期", "atomic_code": "pay_amount_sum",
        "time_period": "999d"})
    assert r.status_code == 400
    # 不合法的筛选操作符
    r = api.post("/api/v1/derived-metrics", json={
        "code": "bad_filter", "name": "错误限定", "atomic_code": "pay_amount_sum",
        "time_period": "7d", "filters": [{"field": "x", "op": "HACK", "value": 1}]})
    assert r.status_code == 400


def test_status_change_flow(api):
    """原子指标状态：DRAFT -> PUBLISHED -> ARCHIVED"""
    r = api.post("/api/v1/atomic-metrics", json={
        "code": "tmp_status", "name": "临时", "process_id": 1,
        "agg_function": "COUNT", "physical_field": "order_id", "status": "DRAFT"})
    mid = r.json()["data"]["id"]
    r = api.post(f"/api/v1/atomic-metrics/{mid}/status",
                 json={"status": "ARCHIVED"})
    assert r.json()["data"]["status"] == "ARCHIVED"
    r = api.post(f"/api/v1/atomic-metrics/{mid}/status", json={"status": "BAD"})
    assert r.status_code == 400
    api.delete(f"/api/v1/atomic-metrics/{mid}")  # 清理


def test_derived_full_crud(api):
    """派生规则引擎端到端：界面配置 -> 元数据 -> 即刻查询"""
    today = dt.date.today()
    r = api.post("/api/v1/derived-metrics", json={
        "code": "pay_amount_30d_region", "name": "最近30天各大区支付金额",
        "atomic_code": "pay_amount_sum", "time_period": "30d",
        "dim_codes": ["dim_city"],
        "filters": [{"field": "pay_channel", "op": "IN", "value": ["WECHAT", "ALIPAY"]}]})
    assert r.status_code == 200, r.text
    mid = r.json()["data"]["id"]

    # SQL 预览：应含筛选条件
    r = api.get(f"/api/v1/derived-metrics/{mid}/sql-preview")
    assert "pay_channel" in r.json()["data"]["sql"]

    # 查询可用
    r = api.post("/api/v1/query", json={"metric_code": "pay_amount_30d_region",
                                        "dim_codes": ["dim_city"]})
    assert r.status_code == 200

    # 更新周期
    r = api.put(f"/api/v1/derived-metrics/{mid}", json={
        "code": "pay_amount_30d_region", "name": "最近30天各大区支付金额",
        "atomic_code": "pay_amount_sum", "time_period": "7d",
        "dim_codes": ["dim_city"], "filters": []})
    assert r.status_code == 200

    # 删除（未被复合引用）
    r = api.delete(f"/api/v1/derived-metrics/{mid}")
    assert r.status_code == 200
    r = api.get(f"/api/v1/derived-metrics/{mid}")
    assert r.status_code == 404


def test_logic_model_crud_and_sql(api):
    """逻辑模型：定义 JOIN 宽表 -> 生成 SELECT 预览"""
    r = api.post("/api/v1/logical-models", json={
        "code": "test_wide", "name": "测试宽表", "domain_id": 1,
        "physical_table": "dwd_order_detail", "join_type": "JOIN",
        "join_config": [{"table": "dim_city", "alias": "d0",
                         "on": "t.city_id = d0.city_id"}]})
    assert r.status_code == 200
    r = api.get("/api/v1/logical-models")
    model = [m for m in r.json()["data"] if m["code"] == "test_wide"][0]
    assert "LEFT JOIN dim_city d0" in model["generated_sql"]
    assert model["join_type"] == "JOIN"
    api.delete(f"/api/v1/logical-models/{model['id']}")


def test_dimension_attribute_crud(api):
    r = api.post("/api/v1/dimensions/1/attributes",
                 json={"code": "test_attr", "name": "测试属性",
                       "physical_field": "city_name"})
    assert r.status_code == 200
    attr_id = r.json()["data"]["id"]
    r = api.get("/api/v1/dimensions/1")
    assert any(a["id"] == attr_id for a in r.json()["data"]["attributes"])
    r = api.delete(f"/api/v1/dimension-attributes/{attr_id}")
    assert r.status_code == 200


# ===========================================================================
# Excel 导出
# ===========================================================================

def test_export_excel(api):
    r = api.get("/api/v1/query/export", params={
        "metric_codes": "pay_amount_7d_city,order_count", "dim_codes": "dim_city"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"  # zip 魔数
    from openpyxl import load_workbook
    wb = load_workbook(io_bytes(r.content))
    ws = wb.active
    # 表头 = 新查询契约：date_bucket + 维度 + 指标列
    assert [ws.cell(1, c).value for c in (1, 2, 3, 4)] == [
        "date_bucket", "dim_city", "pay_amount_7d_city", "order_count"]
    assert ws.max_row >= 2  # 表头 + 至少一行数据
    # 与查询接口行数一致（口径一致）
    q = api.post("/api/v1/query", json={
        "metric_codes": ["pay_amount_7d_city", "order_count"],
        "dim_codes": ["dim_city"]}).json()["data"]
    assert ws.max_row == len(q["rows"]) + 1


def io_bytes(b):
    import io
    return io.BytesIO(b)


# ===================================================================
# 血缘
# ===================================================================

def test_lineage_full_chain(api):
    r = api.get("/api/v1/lineage/avg_order_value")
    d = r.json()["data"]
    types = {n["type"] for n in d["nodes"]}
    # 契约：物理表=table / 物理字段=field / 指标=atomic|derived|composite
    assert {"composite", "derived", "atomic", "table", "field"} <= types
    # 物理层 -> 原子 -> 派生 -> 复合 的边
    assert any(e["to"].startswith("composite:") for e in d["edges"])
    assert any(e["from"].startswith("field:") for e in d["edges"])


def test_lineage_unknown_404(api):
    r = api.get("/api/v1/lineage/nope")
    assert r.status_code == 404


# ===================================================================
# 多指标联合查询 + 日期粒度（日/周/月）
# ===================================================================

def test_multi_metric_alignment(api):
    """多指标 + 多维度联合查询：按 date_bucket + 公共维度 LEFT JOIN 对齐"""
    r = api.post("/api/v1/query", json={
        "metric_codes": ["order_amount_sum", "order_count"],
        "dim_codes": ["dim_city", "dim_category"],
        "start_date": "2026-08-01", "end_date": "2026-08-19",
        "granularity": "day"})
    assert r.status_code == 200, r.text
    d = r.json()["data"]
    assert d["columns"] == ["date_bucket", "dim_city", "dim_category",
                            "order_amount_sum", "order_count"]
    assert d["summary"]["metric_names"] == ["下单金额", "下单次数"]
    assert d["summary"]["metric_types"] == ["atomic", "atomic"]
    assert d["summary"]["granularity"] == "day"
    assert len(d["rows"]) > 0
    for row in d["rows"]:
        assert len(row) == 5
        assert re.match(r"^\d{4}-\d{2}-\d{2}$", row[0])
        assert row[3] is not None and row[4] is not None  # 两指标均对齐非空


def test_multi_metric_mixed_types(api):
    """原子 + 派生 + 复合混合查询：值列按指标编码定位，LEFT JOIN 可产生空值"""
    r = api.post("/api/v1/query", json={
        "metric_codes": ["order_amount_sum", "pay_amount_7d_city"],
        "dim_codes": ["dim_city"], "granularity": "day"})
    assert r.status_code == 200, r.text
    d = r.json()["data"]
    assert set(d["summary"]["metric_types"]) == {"atomic", "derived"}
    assert "order_amount_sum" in d["columns"] and "pay_amount_7d_city" in d["columns"]


def test_granularity_week_month(api):
    """日期粒度桶格式：周 YYYY-Www / 月 YYYY-MM"""
    cases = (("week", r"^\d{4}-W\d{2}$"), ("month", r"^\d{4}-\d{2}$"))
    for g, pat in cases:
        r = api.post("/api/v1/query", json={
            "metric_codes": ["order_amount_sum"], "dim_codes": ["dim_city"],
            "granularity": g, "start_date": "2026-08-01", "end_date": "2026-08-19"})
        assert r.status_code == 200, r.text
        d = r.json()["data"]
        assert d["summary"]["granularity"] == g
        assert d["columns"][0] == "date_bucket"
        assert len(d["rows"]) > 0
        for row in d["rows"]:
            assert re.match(pat, row[0]), f"{g} 桶格式异常: {row[0]}"


def test_invalid_granularity_400(api):
    r = api.post("/api/v1/query", json={"metric_codes": ["order_amount_sum"],
                                        "granularity": "hour"})
    assert r.status_code == 400
    assert "粒度" in r.json()["message"]
    # sql-preview / export 同样拒绝
    r = api.get("/api/v1/sql-preview", params={"metric_codes": "order_amount_sum",
                                               "granularity": "hour"})
    assert r.status_code == 400


def test_dim_incompatible_metric_400(api):
    """维度关联字段不在指标物理表中 -> 400（防 500 崩溃）"""
    r = api.post("/api/v1/query", json={
        "metric_codes": ["refund_amount_sum"], "dim_codes": ["dim_user"]})
    assert r.status_code == 400
    assert "不支持维度" in r.json()["message"]
    # 元数据完好
    assert api.get("/api/v1/metrics").status_code == 200


def test_injection_metric_codes(api):
    """多指标参数注入：非法编码被拒，数据库不受影响"""
    r = api.post("/api/v1/query", json={
        "metric_codes": ["pay_amount_sum; DROP TABLE meta_atomic_metric;--"]})
    assert r.status_code in (400, 404)
    r = api.get("/api/v1/metrics")
    assert r.status_code == 200 and len(r.json()["data"]["atomic"]) == 5


# ===================================================================
# 下游模型：定义 SQL / 物化（幂等）/ 表血缘
# ===================================================================

def _trade_wide_lm(api):
    lms = api.get("/api/v1/logical-models").json()["data"]
    return [m for m in lms if m["code"] == "trade_wide_order"][0]


def test_downstream_crud_and_materialize(api):
    """下游模型：创建（定义 SQL 入库）-> 物化 -> 幂等刷新 -> 编辑重置 -> 删除"""
    lm = _trade_wide_lm(api)
    r = api.post("/api/v1/downstream-models", json={
        "code": "city_pay_daily_test", "name": "城市订单日汇总(测试)",
        "source_model_id": lm["id"], "granularity": "day",
        "metrics": [{"metric_code": "order_amount_sum", "dim_codes": ["dim_city"]},
                    {"metric_code": "order_count", "dim_codes": ["dim_city"]}]})
    assert r.status_code == 200, r.text
    mid = r.json()["data"]["id"]
    sql0 = r.json()["data"]["definition_sql"]
    assert len(sql0) > 0 and "date_bucket" in sql0 and "CREATE" not in sql0

    # 详情：定义 SQL 持久化，列表可见
    d = api.get(f"/api/v1/downstream-models/{mid}").json()["data"]
    assert d["definition_sql"] == sql0
    assert d["materialized"] is False
    items = api.get("/api/v1/downstream-models").json()["data"]["items"]
    assert any(x["id"] == mid and x["source_model_name"] == "订单交易宽表"
               for x in items)

    # 物化：CREATE TABLE dl_{code} AS <sql>；再次物化 = 重建刷新（幂等）
    r = api.post(f"/api/v1/downstream-models/{mid}/materialize")
    assert r.status_code == 200, r.text
    tbl, n1 = r.json()["data"]["physical_table"], r.json()["data"]["row_count"]
    assert tbl == "dl_city_pay_daily_test" and n1 > 0
    r2 = api.post(f"/api/v1/downstream-models/{mid}/materialize")
    assert r2.status_code == 200 and r2.json()["data"]["row_count"] == n1

    # 物化状态 + 物化表数据可查
    m = [x for x in api.get("/api/v1/downstream-models").json()["data"]["items"]
         if x["id"] == mid][0]
    assert m["materialized"] is True and m["physical_table"] == tbl
    assert m["row_count"] == n1
    dd = api.get(f"/api/v1/downstream-models/{mid}/data").json()["data"]
    assert dd["total"] == n1 and dd["columns"][0] == "date_bucket"
    assert dd["columns"][1] == "dim_city"

    # 编辑：定义变更后旧物化表被拆除、状态复位
    r = api.put(f"/api/v1/downstream-models/{mid}", json={
        "code": "city_pay_daily_test", "name": "城市订单日汇总(测试)",
        "source_model_id": lm["id"], "granularity": "month",
        "metrics": [{"metric_code": "order_amount_sum", "dim_codes": ["dim_city"]}]})
    assert r.status_code == 200
    m = api.get(f"/api/v1/downstream-models/{mid}").json()["data"]
    assert m["materialized"] is False and m["physical_table"] is None
    from models import engine
    from sqlalchemy import text
    exists = engine.connect().execute(text(
        "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name=:t"),
        {"t": tbl}).scalar()
    assert exists == 0

    # 删除：元数据与（已拆除的）物化表均清理
    assert api.delete(f"/api/v1/downstream-models/{mid}").status_code == 200
    assert api.get(f"/api/v1/downstream-models/{mid}").status_code == 404


def test_downstream_validation_400(api):
    lm = _trade_wide_lm(api)
    # 复合指标需先展开为派生指标
    r = api.post("/api/v1/downstream-models", json={
        "code": "bad_composite_ds", "name": "错误", "source_model_id": lm["id"],
        "metrics": [{"metric_code": "avg_order_value", "dim_codes": ["dim_city"]}]})
    assert r.status_code == 400
    assert "复合" in r.json()["message"]
    # 指标过程表不在逻辑模型表范围内（退款表未 JOIN 进宽表）
    r = api.post("/api/v1/downstream-models", json={
        "code": "bad_table_ds", "name": "错误2", "source_model_id": lm["id"],
        "metrics": [{"metric_code": "refund_amount_sum", "dim_codes": ["dim_city"]}]})
    assert r.status_code == 400
    assert "物理表" in r.json()["message"]
    # 编码冲突 409
    r = api.post("/api/v1/downstream-models", json={
        "code": "city_order_daily", "name": "重复", "source_model_id": lm["id"],
        "metrics": [{"metric_code": "order_count", "dim_codes": ["dim_city"]}]})
    assert r.status_code == 409


def test_downstream_preview_no_materialize(api):
    """预览执行定义 SQL 但不落地"""
    lm = _trade_wide_lm(api)
    r = api.post("/api/v1/downstream-models", json={
        "code": "preview_ds", "name": "预览", "source_model_id": lm["id"],
        "metrics": [{"metric_code": "order_count", "dim_codes": ["dim_city"]}]})
    mid = r.json()["data"]["id"]
    p = api.post(f"/api/v1/downstream-models/{mid}/preview", params={"limit": 50})
    assert p.status_code == 200
    d = p.json()["data"]
    assert d["columns"][0] == "date_bucket" and "order_count" in d["columns"]
    assert 0 < len(d["rows"]) <= 50
    # 未物化，data 接口拒绝
    assert api.get(f"/api/v1/downstream-models/{mid}/data").status_code == 400
    api.delete(f"/api/v1/downstream-models/{mid}")


# ===================================================================
# 表级血缘（物理表 -> 逻辑模型 -> 下游模型 -> 物化表）
# ===================================================================

def test_lineage_tables_full_chain(api):
    r = api.get("/api/v1/lineage/tables")
    assert r.status_code == 200
    d = r.json()["data"]
    ids = {n["id"] for n in d["nodes"]}
    types = {n["type"] for n in d["nodes"]}
    assert {"table", "logical_model", "downstream_model"} <= types
    # 物理表（业务过程 + 维度表）与逻辑模型、种子下游模型
    assert "table:dwd_order_detail" in ids
    assert "table:dim_city" in ids
    assert "model:trade_wide_order" in ids
    assert "downstream:city_order_daily" in ids
    # 边：物理表 -> 逻辑模型（含 join_config 表）-> 下游模型
    edges = {(e["from"], e["to"]) for e in d["edges"]}
    assert ("table:dwd_order_detail", "model:trade_wide_order") in edges
    assert ("table:dim_city", "model:trade_wide_order") in edges
    assert ("model:trade_wide_order", "downstream:city_order_daily") in edges
    # 指标血缘端点不受影响
    assert api.get("/api/v1/lineage/avg_order_value").status_code == 200


def test_lineage_tables_after_materialize(api):
    """物化后表血缘出现 下游模型 -> 物化表 链路"""
    lm = _trade_wide_lm(api)
    r = api.post("/api/v1/downstream-models", json={
        "code": "lineage_ds", "name": "血缘测试", "source_model_id": lm["id"],
        "metrics": [{"metric_code": "order_count", "dim_codes": ["dim_city"]}]})
    mid = r.json()["data"]["id"]
    assert api.post(f"/api/v1/downstream-models/{mid}/materialize").status_code == 200
    d = api.get("/api/v1/lineage/tables").json()["data"]
    ids = {n["id"] for n in d["nodes"]}
    assert "table:dl_lineage_ds" in ids
    edges = {(e["from"], e["to"]) for e in d["edges"]}
    assert ("downstream:lineage_ds", "table:dl_lineage_ds") in edges


# ===========================================================================
# 下游应用 / 数据集 / 开放 API（AppKey+AppSecret 认证、双源调用、调用日志）
# ===========================================================================

def _mk_app(api, code):
    """创建下游应用，返回 (id, appkey, appsecret)"""
    r = api.post("/api/v1/downstream-apps", json={
        "code": code, "name": f"{code} 应用", "description": "测试用"})
    assert r.status_code == 200, r.text
    d = r.json()["data"]
    return d["id"], d["appkey"], d["appsecret"]


def test_downstream_app_crud_and_secret(api):
    """下游应用：创建（自动生成密钥）-> 重复编码 409 / 非法状态 400 -> 重置密钥 -> 删除"""
    aid, key0, secret0 = _mk_app(api, "oa_app1")
    assert len(key0) == 20 and len(secret0) >= 24  # token_hex(10) / token_urlsafe(24)
    # 列表可见，密钥明文返回（管理端），累计调用为 0
    items = api.get("/api/v1/downstream-apps?page_size=100").json()["data"]["items"]
    app = [a for a in items if a["id"] == aid][0]
    assert app["code"] == "oa_app1" and app["appkey"] == key0
    assert app["status"] == "ENABLED" and app["call_count"] == 0
    assert app["dataset_count"] == 0
    # 重复编码 409、非法状态 400
    assert api.post("/api/v1/downstream-apps", json={
        "code": "oa_app1", "name": "重复"}).status_code == 409
    assert api.post("/api/v1/downstream-apps", json={
        "code": "oa_app2", "name": "x", "status": "BANNED"}).status_code == 400
    # 重置密钥：AppKey 不变、AppSecret 变更；新密钥可认证、旧密钥失效
    r = api.post(f"/api/v1/downstream-apps/{aid}/reset-secret")
    assert r.status_code == 200
    new_secret = r.json()["data"]["appsecret"]
    assert r.json()["data"]["appkey"] == key0 and new_secret != secret0
    h = {"X-App-Key": key0, "X-App-Secret": new_secret}
    assert api.get("/openapi/v1/datasets", headers=h).status_code == 200
    h_old = {"X-App-Key": key0, "X-App-Secret": secret0}
    assert api.get("/openapi/v1/datasets", headers=h_old).status_code == 401
    # 删除后消失
    assert api.delete(f"/api/v1/downstream-apps/{aid}").status_code == 200
    assert api.get(f"/api/v1/downstream-apps/{aid}").status_code == 404


def test_dataset_crud_and_validation(api):
    """数据集：metric_query 创建/编辑；非法参数 400；重复编码 409"""
    payload = {"code": "oa_ds_mq", "name": "实时指标数据集", "source_type": "metric_query",
               "metric_codes": ["order_amount_sum", "order_count"],
               "dim_codes": ["dim_city"], "granularity": "day"}
    r = api.post("/api/v1/datasets", json=payload)
    assert r.status_code == 200, r.text
    did = r.json()["data"]["id"]
    d = api.get(f"/api/v1/datasets/{did}").json()["data"]
    assert d["source_type"] == "metric_query" and d["granularity"] == "day"
    assert d["metric_codes"] == ["order_amount_sum", "order_count"]
    assert d["source_model_name"] == ""
    # 非法校验：数据源类型 / 日期粒度 / 指标编码 / 缺来源模型
    bads = [
        {"code": "oa_bad1", "name": "x", "source_type": "hive_table"},
        {"code": "oa_bad2", "name": "x", "source_type": "metric_query",
         "metric_codes": ["order_count"], "granularity": "hour"},
        {"code": "oa_bad3", "name": "x", "source_type": "metric_query",
         "metric_codes": ["no_such_metric"], "granularity": "day"},
        {"code": "oa_bad4", "name": "x", "source_type": "downstream_model"},
    ]
    for bad in bads:
        assert api.post("/api/v1/datasets", json=bad).status_code == 400, bad
    # 重复编码 409
    assert api.post("/api/v1/datasets", json=payload).status_code == 409
    # 编辑：改粒度
    payload["granularity"] = "month"
    assert api.put(f"/api/v1/datasets/{did}", json=payload).status_code == 200
    assert api.get(f"/api/v1/datasets/{did}").json()["data"]["granularity"] == "month"
    # 删除后消失
    assert api.delete(f"/api/v1/datasets/{did}").status_code == 200
    assert api.get(f"/api/v1/datasets/{did}").status_code == 404


def test_dataset_detail_grants(api):
    """数据集详情：授权应用列表可见；删除应用后授权记录级联清理"""
    aid, _, _ = _mk_app(api, "oa_app_det")
    r = api.post("/api/v1/datasets", json={
        "code": "oa_ds_det", "name": "详情", "source_type": "metric_query",
        "metric_codes": ["order_count"], "dim_codes": [], "granularity": "day"})
    did = r.json()["data"]["id"]
    api.post(f"/api/v1/datasets/{did}/grant", json={"app_id": aid})
    d = api.get(f"/api/v1/datasets/{did}").json()["data"]
    assert [a["code"] for a in d["granted_apps"]] == ["oa_app_det"]
    # 删除应用 -> 授权记录级联清理
    assert api.delete(f"/api/v1/downstream-apps/{aid}").status_code == 200
    d = api.get(f"/api/v1/datasets/{did}").json()["data"]
    assert d["granted_apps"] == [] and d["granted_app_ids"] == []
    api.delete(f"/api/v1/datasets/{did}")


def test_openapi_auth_guard(api):
    """开放 API 认证：缺头 401 / 密钥错 401 / 应用停用 401 / 数据集不存在 404"""
    aid, key, secret = _mk_app(api, "oa_app_auth")
    url = "/openapi/v1/datasets/oa_ds_none/data"
    assert api.get(url).status_code == 401  # 缺认证头
    assert api.get(url, headers={"X-App-Key": key}).status_code == 401
    assert api.get(url, headers={"X-App-Secret": secret}).status_code == 401
    h_bad = {"X-App-Key": key, "X-App-Secret": "wrong-secret"}
    assert api.get(url, headers=h_bad).status_code == 401
    h_bad2 = {"X-App-Key": "no-such-key", "X-App-Secret": secret}
    assert api.get(url, headers=h_bad2).status_code == 401
    # 停用应用：即使密钥正确也 401
    aid2, key2, secret2 = _mk_app(api, "oa_app_off")
    r = api.put(f"/api/v1/downstream-apps/{aid2}", json={
        "code": "oa_app_off", "name": "停用", "status": "DISABLED"})
    assert r.status_code == 200
    h_off = {"X-App-Key": key2, "X-App-Secret": secret2}
    assert api.get("/openapi/v1/datasets", headers=h_off).status_code == 401
    # 认证通过但数据集不存在 -> 404
    h = {"X-App-Key": key, "X-App-Secret": secret}
    assert api.get(url, headers=h).status_code == 404
    api.delete(f"/api/v1/downstream-apps/{aid}")
    api.delete(f"/api/v1/downstream-apps/{aid2}")


def test_openapi_grant_flow(api):
    """授权流程：未授权 403 -> grant（幂等）后可调 -> revoke 后恢复 403；列表仅含已授权"""
    aid, key, secret = _mk_app(api, "oa_app_grant")
    h = {"X-App-Key": key, "X-App-Secret": secret}
    r = api.post("/api/v1/datasets", json={
        "code": "oa_ds_grant", "name": "授权测试", "source_type": "metric_query",
        "metric_codes": ["order_count"], "dim_codes": ["dim_city"], "granularity": "day"})
    did = r.json()["data"]["id"]
    url = "/openapi/v1/datasets/oa_ds_grant/data"
    assert api.get(url, headers=h).status_code == 403
    assert api.get("/openapi/v1/datasets", headers=h).json()["data"]["datasets"] == []
    # grant 幂等：重复授权去重
    assert api.post(f"/api/v1/datasets/{did}/grant", json={"app_id": aid}).status_code == 200
    assert api.post(f"/api/v1/datasets/{did}/grant",
                    json={"app_id": aid}).json()["data"]["granted"] is False
    assert api.get(url, headers=h).status_code == 200
    ds_list = api.get("/openapi/v1/datasets", headers=h).json()["data"]["datasets"]
    assert [d["code"] for d in ds_list] == ["oa_ds_grant"]
    # revoke 后恢复 403
    assert api.delete(f"/api/v1/datasets/{did}/grant/{aid}").status_code == 200
    assert api.get(url, headers=h).status_code == 403
    api.delete(f"/api/v1/datasets/{did}")
    api.delete(f"/api/v1/downstream-apps/{aid}")


def test_openapi_metric_query_call(api):
    """metric_query 源：实时 SQL 计算、列/行正确、分页、日期过滤"""
    aid, key, secret = _mk_app(api, "oa_app_mq")
    h = {"X-App-Key": key, "X-App-Secret": secret}
    r = api.post("/api/v1/datasets", json={
        "code": "oa_ds_mq2", "name": "实时指标", "source_type": "metric_query",
        "metric_codes": ["order_amount_sum", "order_count"],
        "dim_codes": ["dim_city"], "granularity": "day"})
    did = r.json()["data"]["id"]
    api.post(f"/api/v1/datasets/{did}/grant", json={"app_id": aid})
    url = "/openapi/v1/datasets/oa_ds_mq2/data"
    d = api.get(url, headers=h).json()["data"]
    # 默认窗口：最近 7 天（含今日）x 6 城市 -> 48 行
    assert set(d["columns"]) == {"date_bucket", "dim_city",
                                 "order_amount_sum", "order_count"}
    assert d["total"] == 48 and len(d["rows"]) == 48
    assert d["sql"] and "order_amount" in d["sql"]
    row0 = dict(zip(d["columns"], d["rows"][0]))
    assert row0["order_amount_sum"] > 0 and row0["order_count"] > 0
    assert re.match(r"\d{4}-\d{2}-\d{2}$", row0["date_bucket"])
    # 分页
    d2 = api.get(url, headers=h, params={"page_size": 3, "page": 2}).json()["data"]
    assert len(d2["rows"]) == 3
    # 日期过滤：单日 = 6 个城市
    day = (dt.date.today() - dt.timedelta(days=10)).isoformat()
    d3 = api.get(url, headers=h, params={"start_date": day, "end_date": day}).json()["data"]
    assert d3["total"] == 6
    assert all(dict(zip(d3["columns"], r))["date_bucket"] == day for r in d3["rows"])
    api.delete(f"/api/v1/datasets/{did}")
    api.delete(f"/api/v1/downstream-apps/{aid}")


def test_openapi_downstream_model_source(api):
    """downstream_model 源：未物化 400 -> 物化后可调（物化表直读 + 分页）"""
    lm = _trade_wide_lm(api)
    r = api.post("/api/v1/downstream-models", json={
        "code": "oa_dm1", "name": "开放API物化源", "source_model_id": lm["id"],
        "granularity": "day",
        "metrics": [{"metric_code": "order_amount_sum", "dim_codes": ["dim_city"]},
                    {"metric_code": "order_count", "dim_codes": ["dim_city"]}]})
    mid = r.json()["data"]["id"]
    aid, key, secret = _mk_app(api, "oa_app_dm")
    h = {"X-App-Key": key, "X-App-Secret": secret}
    r = api.post("/api/v1/datasets", json={
        "code": "oa_ds_dm", "name": "物化表数据集", "source_type": "downstream_model",
        "source_model_id": mid, "granularity": "day"})
    did = r.json()["data"]["id"]
    api.post(f"/api/v1/datasets/{did}/grant", json={"app_id": aid})
    url = "/openapi/v1/datasets/oa_ds_dm/data"
    # 未物化：400 提示
    r = api.get(url, headers=h)
    assert r.status_code == 400 and "物化" in r.json()["message"]
    # 物化后可调：物化表直读（无动态 sql 字段）
    assert api.post(f"/api/v1/downstream-models/{mid}/materialize").status_code == 200
    d = api.get(url, headers=h).json()["data"]
    assert d["columns"] == ["date_bucket", "dim_city", "order_amount_sum", "order_count"]
    assert d["total"] > 0 and len(d["rows"]) == min(d["total"], 100)
    assert "sql" not in d
    # 分页
    d2 = api.get(url, headers=h, params={"page_size": 5}).json()["data"]
    assert len(d2["rows"]) == 5 and d2["total"] == d["total"]
    api.delete(f"/api/v1/datasets/{did}")
    api.delete(f"/api/v1/downstream-apps/{aid}")


# ===================================================================
# 重导：上游逻辑模型指标/维度更新上线后，下游模型按时间范围重导数据
# 默认范围 = 近 3 个月（3 个月前当月 1 日 ~ 今天），可传 start_date/end_date 覆盖
# ===================================================================

def _mk_ds_model(api, code):
    """创建日粒度下游模型（dim_city + 订单金额/订单数）"""
    lm = _trade_wide_lm(api)
    r = api.post("/api/v1/downstream-models", json={
        "code": code, "name": code, "source_model_id": lm["id"],
        "granularity": "day",
        "metrics": [{"metric_code": "order_amount_sum", "dim_codes": ["dim_city"]},
                    {"metric_code": "order_count", "dim_codes": ["dim_city"]}]})
    assert r.status_code == 200, r.text
    return r.json()["data"]["id"]


def test_reimport_requires_materialize(api):
    """未物化不可重导 -> 400 提示先物化"""
    mid = _mk_ds_model(api, "reimp_nom")
    r = api.post(f"/api/v1/downstream-models/{mid}/reimport")
    assert r.status_code == 400 and "物化" in r.json()["message"]
    api.delete(f"/api/v1/downstream-models/{mid}")


def test_reimport_default_range(api):
    """默认重导近 3 个月：种子数据窗口（近 30 天）全部落在区间内 -> 全量重算，行数幂等"""
    mid = _mk_ds_model(api, "reimp_def")
    n0 = api.post(f"/api/v1/downstream-models/{mid}/materialize").json()["data"]["row_count"]
    r = api.post(f"/api/v1/downstream-models/{mid}/reimport")
    assert r.status_code == 200, r.text
    d = r.json()["data"]
    assert d["physical_table"] == "dl_reimp_def"
    assert d["start_date"] < d["end_date"]           # 默认区间 = 近 3 个月
    assert d["deleted"] == n0 and d["inserted"] == n0 and d["total_rows"] == n0
    dd = api.get(f"/api/v1/downstream-models/{mid}/data",
                 params={"page_size": 1000}).json()["data"]
    assert dd["total"] == n0                         # 重导后物化表数据完整
    api.delete(f"/api/v1/downstream-models/{mid}")


def test_reimport_custom_range(api):
    """自定义时间范围重导：仅重建区间内行（删除+重算），区间外数据保留；重复重导幂等"""
    mid = _mk_ds_model(api, "reimp_rng")
    n0 = api.post(f"/api/v1/downstream-models/{mid}/materialize").json()["data"]["row_count"]
    d0 = api.get(f"/api/v1/downstream-models/{mid}/data",
                 params={"page_size": 1000}).json()["data"]
    num_days = len({r[0] for r in d0["rows"]})       # 种子覆盖的自然日数（30）
    per_day = n0 // num_days                         # 日粒度 x 6 城市
    today = dt.date.today()
    start, end = today - dt.timedelta(days=10), today - dt.timedelta(days=6)
    expect = ((end - start).days + 1) * per_day      # 5 天 x 6 城市 = 30
    r = api.post(f"/api/v1/downstream-models/{mid}/reimport",
                 params={"start_date": start.isoformat(), "end_date": end.isoformat()})
    assert r.status_code == 200, r.text
    d = r.json()["data"]
    assert d["deleted"] == expect and d["inserted"] == expect
    assert d["total_rows"] == n0                     # 区间外行保留
    # 区间内日期桶全部重建、区间外日期桶仍在
    dd = api.get(f"/api/v1/downstream-models/{mid}/data",
                 params={"page_size": 1000}).json()["data"]
    buckets = {r[0] for r in dd["rows"]}
    assert start.isoformat() in buckets and end.isoformat() in buckets
    assert {b for b in buckets if start.isoformat() <= b <= end.isoformat()} == \
        {(start + dt.timedelta(days=i)).isoformat() for i in range(5)}
    # 重复重导同一区间：幂等（删除/写入数一致，总量不变）
    r2 = api.post(f"/api/v1/downstream-models/{mid}/reimport",
                  params={"start_date": start.isoformat(), "end_date": end.isoformat()}).json()["data"]
    assert r2["deleted"] == expect and r2["inserted"] == expect and r2["total_rows"] == n0
    api.delete(f"/api/v1/downstream-models/{mid}")


def test_reimport_validation_and_week(api):
    """重导校验：起止倒置 / 日期格式非法 -> 400；周粒度模型默认范围重导正常"""
    lm = _trade_wide_lm(api)
    r = api.post("/api/v1/downstream-models", json={
        "code": "reimp_val", "name": "重导校验", "source_model_id": lm["id"],
        "granularity": "week",
        "metrics": [{"metric_code": "order_amount_sum", "dim_codes": ["dim_city"]}]})
    mid = r.json()["data"]["id"]
    api.post(f"/api/v1/downstream-models/{mid}/materialize")
    assert api.post(f"/api/v1/downstream-models/{mid}/reimport",
                    params={"start_date": "2026-08-20", "end_date": "2026-08-01"}).status_code == 400
    assert api.post(f"/api/v1/downstream-models/{mid}/reimport",
                    params={"start_date": "2026/08/01"}).status_code == 400
    # 默认近 3 个月 -> 覆盖全部种子周桶，删除=写入=总量
    d = api.post(f"/api/v1/downstream-models/{mid}/reimport").json()["data"]
    assert d["deleted"] == d["inserted"] > 0 and d["total_rows"] == d["deleted"]
    api.delete(f"/api/v1/downstream-models/{mid}")


def test_count_distinct_agg(api):
    """COUNT_DISTINCT 聚合：翻译为 COUNT(DISTINCT x)（SQLite 无原生 COUNT_DISTINCT 函数）
    —— 统一查询与下游模型物化/重导均可用"""
    r = api.post("/api/v1/atomic-metrics", json={
        "code": "cd_order_cnt", "name": "去重订单数", "process_id": 1,
        "agg_function": "COUNT_DISTINCT", "physical_field": "order_id"})
    assert r.status_code == 200, r.text
    amid = r.json()["data"]["id"]
    try:
        # 统一查询路径：SQL 中应为 COUNT(DISTINCT ...) 而非 COUNT_DISTINCT(...)
        q = api.post("/api/v1/query", json={
            "metric_codes": ["cd_order_cnt"], "dim_codes": ["dim_city"],
            "granularity": "day"}).json()["data"]
        assert q["summary"]["row_count"] > 0
        assert "COUNT(DISTINCT" in q["sql"] and "COUNT_DISTINCT(" not in q["sql"]
        # 下游模型：物化 + 默认范围重导（此前 COUNT_DISTINCT 原样拼接导致 SQLite 报错）
        lm = _trade_wide_lm(api)
        r = api.post("/api/v1/downstream-models", json={
            "code": "cd_ds_model", "name": "去重计数", "source_model_id": lm["id"],
            "granularity": "day",
            "metrics": [{"metric_code": "cd_order_cnt", "dim_codes": ["dim_city"]},
                        {"metric_code": "order_count", "dim_codes": ["dim_city"]}]})
        mid = r.json()["data"]["id"]
        assert "COUNT(DISTINCT" in api.get(f"/api/v1/downstream-models/{mid}").json()["data"]["definition_sql"]
        n0 = api.post(f"/api/v1/downstream-models/{mid}/materialize").json()["data"]["row_count"]
        assert n0 > 0
        d = api.post(f"/api/v1/downstream-models/{mid}/reimport").json()["data"]
        assert d["deleted"] == n0 and d["inserted"] == n0 and d["total_rows"] == n0
        api.delete(f"/api/v1/downstream-models/{mid}")
    finally:
        api.delete(f"/api/v1/atomic-metrics/{amid}")


def test_openapi_call_logging_stats(api):
    """调用日志：每次成功调用记录日志，stats 计数/行数累加"""
    aid, key, secret = _mk_app(api, "oa_app_log")
    h = {"X-App-Key": key, "X-App-Secret": secret}
    r = api.post("/api/v1/datasets", json={
        "code": "oa_ds_log", "name": "日志", "source_type": "metric_query",
        "metric_codes": ["order_count"], "dim_codes": ["dim_city"], "granularity": "day"})
    did = r.json()["data"]["id"]
    api.post(f"/api/v1/datasets/{did}/grant", json={"app_id": aid})
    url = "/openapi/v1/datasets/oa_ds_log/data"
    s0 = api.get("/api/v1/openapi/stats").json()["data"]
    app0 = [x for x in s0["by_app"] if x["app_id"] == aid][0]
    assert app0["calls"] == 0  # 尚无调用
    for _ in range(2):
        assert api.get(url, headers=h).status_code == 200
    s1 = api.get("/api/v1/openapi/stats").json()["data"]
    assert s1["total_calls"] - s0["total_calls"] == 2
    app_stat = [x for x in s1["by_app"] if x["app_id"] == aid][0]
    ds_stat = [x for x in s1["by_dataset"] if x["dataset_id"] == did][0]
    assert app_stat["calls"] == 2 and app_stat["app_code"] == "oa_app_log"
    assert ds_stat["calls"] == 2 and ds_stat["dataset_code"] == "oa_ds_log"
    assert app_stat["rows"] > 0 and app_stat["rows"] == ds_stat["rows"]
    # 日志：含应用/数据集/行数/耗时/状态
    logs = api.get("/api/v1/openapi/logs", params={"page_size": 5}).json()["data"]["items"]
    top = [x for x in logs if x["app_code"] == "oa_app_log"][0]
    assert top["dataset_code"] == "oa_ds_log" and top["status"] == "success"
    assert top["row_count"] > 0 and top["duration_ms"] >= 0
    api.delete(f"/api/v1/datasets/{did}")
    api.delete(f"/api/v1/downstream-apps/{aid}")


def test_openapi_dataset_listing_per_app(api):
    """各应用仅能看到被授权的数据集（列表隔离）"""
    aid1, k1, s1 = _mk_app(api, "oa_app_l1")
    aid2, k2, s2 = _mk_app(api, "oa_app_l2")
    h1, h2 = {"X-App-Key": k1, "X-App-Secret": s1}, {"X-App-Key": k2, "X-App-Secret": s2}
    did1 = api.post("/api/v1/datasets", json={
        "code": "oa_ds_l1", "name": "L1", "source_type": "metric_query",
        "metric_codes": ["order_count"], "dim_codes": [], "granularity": "day"})
    did1 = did1.json()["data"]["id"]
    did2 = api.post("/api/v1/datasets", json={
        "code": "oa_ds_l2", "name": "L2", "source_type": "metric_query",
        "metric_codes": ["order_count"], "dim_codes": [], "granularity": "day"})
    did2 = did2.json()["data"]["id"]
    api.post(f"/api/v1/datasets/{did1}/grant", json={"app_id": aid1})
    api.post(f"/api/v1/datasets/{did2}/grant", json={"app_id": aid2})
    codes1 = {d["code"] for d in
              api.get("/openapi/v1/datasets", headers=h1).json()["data"]["datasets"]}
    codes2 = {d["code"] for d in
              api.get("/openapi/v1/datasets", headers=h2).json()["data"]["datasets"]}
    assert "oa_ds_l1" in codes1 and "oa_ds_l2" not in codes1
    assert "oa_ds_l2" in codes2 and "oa_ds_l1" not in codes2
    api.delete(f"/api/v1/datasets/{did1}")
    api.delete(f"/api/v1/datasets/{did2}")
    api.delete(f"/api/v1/downstream-apps/{aid1}")
    api.delete(f"/api/v1/downstream-apps/{aid2}")


def test_openapi_injection_guard(api):
    """注入回归：伪造 AppKey / 数据集编码含 SQL 片段 -> 401/404，库不受影响"""
    aid, key, secret = _mk_app(api, "oa_app_sec")
    h = {"X-App-Key": key, "X-App-Secret": secret}
    r = api.post("/api/v1/datasets", json={
        "code": "oa_ds_sec", "name": "安全", "source_type": "metric_query",
        "metric_codes": ["order_count"], "dim_codes": ["dim_city"], "granularity": "day"})
    did = r.json()["data"]["id"]
    api.post(f"/api/v1/datasets/{did}/grant", json={"app_id": aid})
    evil_key = {"X-App-Key": "x' OR '1'='1", "X-App-Secret": "x"}
    assert api.get("/openapi/v1/datasets", headers=evil_key).status_code == 401
    evil_code = "oa_ds_sec'; DROP TABLE meta_dataset;--"
    assert api.get(f"/openapi/v1/datasets/{evil_code}/data", headers=h).status_code == 404
    assert api.get("/openapi/v1/datasets/oa_ds_sec/data;--", headers=h).status_code == 404
    # 库未受损：数据集仍在、正常调用仍成功
    assert api.get("/api/v1/datasets?page_size=100").json()["data"]["total"] >= 3
    assert api.get("/openapi/v1/datasets/oa_ds_sec/data", headers=h).status_code == 200
    api.delete(f"/api/v1/datasets/{did}")
    api.delete(f"/api/v1/downstream-apps/{aid}")

# ===================================================================
# 任务重导：对象下游血缘（impact）+ 执行计划（plan）+ 确认执行（execute）
# ===================================================================

def _mk_order_derived(api, code):
    """创建基于订单过程的派生指标（order_amount_sum + 7d + dim_city）"""
    r = api.post("/api/v1/derived-metrics", json={
        "code": code, "name": code, "atomic_code": "order_amount_sum",
        "time_period": "7d", "dim_codes": ["dim_city"]})
    assert r.status_code == 200, r.text
    return r.json()["data"]["id"]


def _impact(api, otype, oid):
    r = api.get("/api/v1/reimport/impact",
                params={"object_type": otype, "object_id": oid})
    assert r.status_code == 200, r.text
    return r.json()["data"]


def test_reimport_impact_atomic_and_lm(api):
    """原子指标/逻辑模型 -> 受影响下游模型及血缘链（对象 -> 逻辑模型 -> 下游）"""
    atom = [m for m in api.get("/api/v1/atomic-metrics").json()["data"]["items"]
            if m["code"] == "order_count"][0]
    # 种子下游模型未物化：先物化，验证 impact 如实返回物化状态与行数
    seed = [x for x in api.get("/api/v1/downstream-models",
                               params={"page_size": 100}).json()["data"]["items"]
            if x["code"] == "city_order_daily"][0]
    api.post(f"/api/v1/downstream-models/{seed['id']}/materialize")
    d = _impact(api, "atomic_metric", atom["id"])
    assert d["object"]["code"] == "order_count"
    ds = [x for x in d["downstreams"] if x["code"] == "city_order_daily"][0]
    assert ds["materialized"] and ds["row_count"] > 0
    assert [n["type"] for n in ds["chain"]] == \
        ["atomic_metric", "logical_model", "downstream"]

    lm = _trade_wide_lm(api)
    d = _impact(api, "logical_model", lm["id"])
    # 其他测试可能残留指向同一逻辑模型的下游模型，断言至少含种子模型
    assert {"city_order_daily"} <= {x["code"] for x in d["downstreams"]}
    ds0 = [x for x in d["downstreams"] if x["code"] == "city_order_daily"][0]
    assert [n["type"] for n in ds0["chain"]] == ["logical_model", "downstream"]


def test_reimport_impact_derived_and_dimension(api):
    """派生指标/维度 -> 下游血缘；原子经派生间接命中（链含派生中介）"""
    did = _mk_order_derived(api, "ri_amt_7d")
    lm = _trade_wide_lm(api)
    r = api.post("/api/v1/downstream-models", json={
        "code": "ri_ds_der", "name": "派生下游", "source_model_id": lm["id"],
        "granularity": "day",
        "metrics": [{"metric_code": "ri_amt_7d", "dim_codes": ["dim_city"]}]})
    mid = r.json()["data"]["id"]
    try:
        # 派生指标直接命中
        d = _impact(api, "derived_metric", did)
        assert {x["code"] for x in d["downstreams"]} == {"ri_ds_der"}
        assert [n["type"] for n in d["downstreams"][0]["chain"]] == \
            ["derived_metric", "logical_model", "downstream"]
        # 其原子指标经派生间接命中（chain 含派生中介节点）
        atom = [m for m in api.get("/api/v1/atomic-metrics").json()["data"]["items"]
                if m["code"] == "order_amount_sum"][0]
        ds = [x for x in _impact(api, "atomic_metric", atom["id"])["downstreams"]
              if x["code"] == "ri_ds_der"][0]
        assert [n["type"] for n in ds["chain"]] == \
            ["atomic_metric", "derived_metric", "logical_model", "downstream"]
        # 维度：直接命中（city_order_daily 的 dim_codes）+ 经派生命中（ri_ds_der）
        dims = api.get("/api/v1/dimensions").json()["data"]
        dim = [x for x in (dims if isinstance(dims, list) else dims["items"])
               if x["code"] == "dim_city"][0]
        codes = {x["code"] for x in _impact(api, "dimension", dim["id"])["downstreams"]}
        assert {"city_order_daily", "ri_ds_der"} <= codes
    finally:
        api.delete(f"/api/v1/downstream-models/{mid}")
        api.delete(f"/api/v1/derived-metrics/{did}")


def test_reimport_impact_validation(api):
    """对象类型非法 -> 400；对象不存在 -> 404"""
    r = api.get("/api/v1/reimport/impact",
                params={"object_type": "composite_metric", "object_id": 1})
    assert r.status_code == 400 and "对象类型" in r.json()["message"]
    assert api.get("/api/v1/reimport/impact",
                   params={"object_type": "atomic_metric", "object_id": 99999}).status_code == 404


def test_reimport_plan_estimated(api):
    """执行计划：预估删除行数 = 物化表区间内行数；未物化不可预估；downstream_ids 过滤"""
    mid_ok = _mk_ds_model(api, "ri_plan_ok")
    n0 = api.post(f"/api/v1/downstream-models/{mid_ok}/materialize").json()["data"]["row_count"]
    lm = _trade_wide_lm(api)
    r = api.post("/api/v1/downstream-models", json={
        "code": "ri_plan_nm", "name": "未物化", "source_model_id": lm["id"],
        "granularity": "day",
        "metrics": [{"metric_code": "order_amount_sum", "dim_codes": ["dim_city"]}]})
    mid_nm = r.json()["data"]["id"]
    atom = [m for m in api.get("/api/v1/atomic-metrics").json()["data"]["items"]
            if m["code"] == "order_amount_sum"][0]
    try:
        d = api.post("/api/v1/reimport/plan", json={
            "object_type": "atomic_metric", "object_id": atom["id"]}).json()["data"]
        assert d["start_date"] < d["end_date"]        # 默认近 3 个月
        items = {x["id"]: x for x in d["items"]}
        assert items[mid_ok]["estimated_deleted"] == n0    # 区间覆盖全部种子数据
        assert items[mid_ok]["materialized"] is True
        assert items[mid_nm]["estimated_deleted"] is None  # 未物化不可预估
        # downstream_ids 过滤：只计划选中模型
        d2 = api.post("/api/v1/reimport/plan", json={
            "object_type": "atomic_metric", "object_id": atom["id"],
            "downstream_ids": [mid_ok]}).json()["data"]
        assert [x["id"] for x in d2["items"]] == [mid_ok]
        # 自定义区间：预估 = 区间内现有行数（5 天 x 6 城市）
        today = dt.date.today()
        start, end = today - dt.timedelta(days=10), today - dt.timedelta(days=6)
        d3 = api.post("/api/v1/reimport/plan", json={
            "object_type": "atomic_metric", "object_id": atom["id"],
            "downstream_ids": [mid_ok],
            "start_date": start.isoformat(), "end_date": end.isoformat()}).json()["data"]
        assert d3["items"][0]["estimated_deleted"] == 5 * (n0 // 30)
    finally:
        api.delete(f"/api/v1/downstream-models/{mid_ok}")
        api.delete(f"/api/v1/downstream-models/{mid_nm}")


def test_reimport_plan_execute_batch(api):
    """确认执行：批量重导 ok；未物化 skipped；单模型失败不阻断其余"""
    m1 = _mk_ds_model(api, "ri_ex_1")
    m2 = _mk_ds_model(api, "ri_ex_2")
    n1 = api.post(f"/api/v1/downstream-models/{m1}/materialize").json()["data"]["row_count"]
    n2 = api.post(f"/api/v1/downstream-models/{m2}/materialize").json()["data"]["row_count"]
    lm = _trade_wide_lm(api)
    r = api.post("/api/v1/downstream-models", json={
        "code": "ri_ex_nm", "name": "未物化", "source_model_id": lm["id"],
        "granularity": "day",
        "metrics": [{"metric_code": "order_amount_sum", "dim_codes": ["dim_city"]}]})
    m3 = r.json()["data"]["id"]
    try:
        # 全量默认区间：ok，删除/写入/总量一致
        d = api.post("/api/v1/reimport/plan/execute", json={
            "downstream_ids": [m1, m2]}).json()["data"]
        by_id = {x["id"]: x for x in d["results"]}
        assert by_id[m1]["status"] == "ok" and by_id[m1]["deleted"] == n1
        assert by_id[m2]["status"] == "ok" and by_id[m2]["inserted"] == n2
        assert by_id[m2]["total_rows"] == n2
        # 自定义区间 + 未物化 skipped（不阻断 ok 模型）
        today = dt.date.today()
        start, end = today - dt.timedelta(days=3), today - dt.timedelta(days=1)
        d2 = api.post("/api/v1/reimport/plan/execute", json={
            "downstream_ids": [m1, m3],
            "start_date": start.isoformat(), "end_date": end.isoformat()}).json()["data"]
        by_id2 = {x["id"]: x for x in d2["results"]}
        assert by_id2[m1]["status"] == "ok"
        assert by_id2[m1]["deleted"] == 3 * (n1 // 30)
        assert by_id2[m3]["status"] == "skipped" and "物化" in by_id2[m3]["message"]
        # 单模型失败不阻断其余：DROP 物化表 -> DELETE 报错 -> error，另一个正常
        from models import engine
        from sqlalchemy import text
        with engine.connect() as conn:
            conn.execute(text("DROP TABLE dl_ri_ex_2"))
        d3 = api.post("/api/v1/reimport/plan/execute", json={
            "downstream_ids": [m1, m2]}).json()["data"]
        by_id3 = {x["id"]: x for x in d3["results"]}
        assert by_id3[m1]["status"] == "ok"
        assert by_id3[m2]["status"] == "error"
        # 空列表 / 起止倒置 -> 400
        assert api.post("/api/v1/reimport/plan/execute",
                        json={"downstream_ids": []}).status_code == 400
        assert api.post("/api/v1/reimport/plan/execute",
                        json={"downstream_ids": [m1],
                              "start_date": "2026-08-20",
                              "end_date": "2026-08-01"}).status_code == 400
    finally:
        api.delete(f"/api/v1/downstream-models/{m1}")
        api.delete(f"/api/v1/downstream-models/{m2}")
        api.delete(f"/api/v1/downstream-models/{m3}")
